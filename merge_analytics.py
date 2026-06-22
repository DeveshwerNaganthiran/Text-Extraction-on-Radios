import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill
import time

def categorize_root_cause(error_msg):
    """Analyzes the AI error explanation to determine the exact Root Cause of the failure."""
    msg = str(error_msg).lower()
    if msg == 'nan' or msg.strip() == '': return 'Passed (No Error)'
    if 'truncat' in msg: return 'Screen Truncation (Text too long)'
    if 'roll' in msg: return 'Rolling Text Sync Issue'
    if 'space' in msg: return 'Spacing or Punctuation Mismatch'
    if 'extra' in msg: return 'Extra Words/Prefixes on Screen'
    if 'typo' in msg or 'misread' in msg or 'incorrect character' in msg: return 'OCR Character Confusion / Font Issue'
    if 'completely different' in msg: return 'Completely Wrong Text / Translation Bug'
    if 'missing' in msg: return 'Missing Words/Letters'
    return 'Other Mismatch'

def calculate_stats(df, group_col):
    """Calculates Total, Pass, Fail, Warn, and their percentages."""
    if df.empty:
        return pd.DataFrame(columns=[group_col, 'Total', 'Passes', 'Fails', 'Warnings', 'Pass_Pct', 'Fail_Pct', 'Warn_Pct'])
    
    stats = df.groupby(group_col)['Verdict'].value_counts().unstack(fill_value=0)
    
    for col in ['PASS', 'FAIL', 'WARN']:
        if col not in stats.columns:
            stats[col] = 0
            
    stats['Total'] = stats['PASS'] + stats['FAIL'] + stats['WARN']
    stats['Passes'] = stats['PASS']
    stats['Fails'] = stats['FAIL']
    stats['Warnings'] = stats['WARN']
    
    stats['Pass_Pct'] = (stats['Passes'] / stats['Total']) * 100
    stats['Fail_Pct'] = (stats['Fails'] / stats['Total']) * 100
    stats['Warn_Pct'] = (stats['Warnings'] / stats['Total']) * 100
    
    stats = stats.reset_index()
    stats = stats.sort_values(by='Total', ascending=False)
    
    if not stats.empty:
        grand_total = pd.DataFrame({
            group_col: ['GRAND TOTAL'],
            'Total': [stats['Total'].sum()],
            'Passes': [stats['Passes'].sum()],
            'Fails': [stats['Fails'].sum()],
            'Warnings': [stats['Warnings'].sum()]
        })
        
        t_tot = grand_total['Total'].iloc[0]
        grand_total['Pass_Pct'] = (grand_total['Passes'] / t_tot * 100) if t_tot > 0 else 0
        grand_total['Fail_Pct'] = (grand_total['Fails'] / t_tot * 100) if t_tot > 0 else 0
        grand_total['Warn_Pct'] = (grand_total['Warnings'] / t_tot * 100) if t_tot > 0 else 0
        
        stats = pd.concat([stats, grand_total], ignore_index=True)
    
    return stats[[group_col, 'Total', 'Passes', 'Fails', 'Warnings', 'Pass_Pct', 'Fail_Pct', 'Warn_Pct']]

def autofit_columns(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns):
        series = df[col]
        max_len = max((series.astype(str).map(len).max(), len(str(col)))) + 2
        max_len = min(max_len, 50)
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

def inject_dashboard(writer, df_raw, sheet_name, title):
    if df_raw.empty: return
    
    verdicts = df_raw['Verdict'].astype(str).str.strip().str.upper()
    passes = (verdicts == 'PASS').sum()
    fails = (verdicts == 'FAIL').sum()
    warns = (verdicts == 'WARN').sum()
    total_valid = passes + fails + warns
    
    if total_valid == 0: return
    
    ws = writer.sheets[sheet_name]
    
    ws.cell(row=1, column=10, value="Verdict").font = Font(bold=True)
    ws.cell(row=1, column=11, value="Count").font = Font(bold=True)
    ws.cell(row=2, column=10, value="PASS")
    ws.cell(row=2, column=11, value=passes)
    ws.cell(row=3, column=10, value="FAIL")
    ws.cell(row=3, column=11, value=fails)
    ws.cell(row=4, column=10, value="WARN")
    ws.cell(row=4, column=11, value=warns)
    
    manual_review_count = fails + warns
    manual_review_pct = (manual_review_count / total_valid) * 100 if total_valid > 0 else 0
    
    ws.cell(row=6, column=10, value="QA MANUAL REVIEW REQUIRED:").font = Font(bold=True, color="9C0006")
    ws.cell(row=6, column=11, value=f"{manual_review_count} items ({manual_review_pct:.1f}%)").font = Font(bold=True, color="9C0006")
    
    pie = PieChart()
    labels = Reference(ws, min_col=10, min_row=2, max_row=4)
    data = Reference(ws, min_col=11, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = f"{title} Verdict Breakdown"
    ws.add_chart(pie, "M1")

    max_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=max_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

def inject_bar_chart(writer, df_reasons, sheet_name):
    if df_reasons.empty: return
    ws = writer.sheets[sheet_name]
    
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Root Cause Failure Analysis"
    bar.y_axis.title = 'Number of Errors'
    bar.x_axis.title = 'Error Category'

    data = Reference(ws, min_col=2, min_row=1, max_row=len(df_reasons)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_reasons)+1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 20
    bar.height = 10
    ws.add_chart(bar, "D2")

def generate_historical_snapshot(df_merged, normal_lang_stats):
    """Creates a flat 1-row snapshot for run-to-run trend tracking."""
    verdicts = df_merged['Verdict'].astype(str).str.strip().str.upper()
    passes = (verdicts == 'PASS').sum()
    fails = (verdicts == 'FAIL').sum()
    warns = (verdicts == 'WARN').sum()
    total = passes + fails + warns
    
    total_langs = df_merged['Language'].nunique() if 'Language' in df_merged.columns else 0
    pass_pct = round((passes / total) * 100, 1) if total > 0 else 0
    fail_pct = round((fails / total) * 100, 1) if total > 0 else 0
    
    actual_mins = 0.0
    avg_sec_per_string = 0.0
    
    if 'Timestamp' in df_merged.columns:
        dates = pd.to_datetime(df_merged['Timestamp'], errors='coerce').dropna().sort_values()
        if not dates.empty and len(dates) > 1:
            diffs = dates.diff().dt.total_seconds().dropna()
            diffs = diffs.clip(upper=90.0)
            total_active_seconds = diffs.sum() + 30.0 
            
            actual_mins = round(total_active_seconds / 60.0, 1)
            if total > 0:
                avg_sec_per_string = round(total_active_seconds / total, 1)
    
    worst_lang = "None"
    if not normal_lang_stats.empty and 'Language' in normal_lang_stats.columns:
        lang_only = normal_lang_stats[normal_lang_stats['Language'] != 'GRAND TOTAL']
        if not lang_only.empty:
            worst_lang_row = lang_only.sort_values(by='Fail_Pct', ascending=False).iloc[0]
            worst_lang = f"{worst_lang_row['Language']} ({worst_lang_row['Fail_Pct']:.1f}% Fail)"
    
    snapshot = pd.DataFrame({
        "Run_Date": [time.strftime("%Y-%m-%d %H:%M")],
        "Total_Languages_Tested": [total_langs],
        "Total_Strings_Verified": [total],
        "Overall_Pass_%": [pass_pct],
        "Overall_Fail_%": [fail_pct],
        "Worst_Language": [worst_lang],
        "Avg_Time_Per_String_Sec": [avg_sec_per_string],
        "Actual_Total_Time_Mins": [actual_mins]
    })
    return snapshot

def main():
    root = Tk()
    root.withdraw()
    
    print("==================================================")
    print(" WALKIE-TRACKER EXECUTIVE ANALYTICS GENERATOR")
    print("==================================================")
    
    file_paths = filedialog.askopenfilenames(title="Select Excel Parts", filetypes=[("Excel Files", "*.xlsx *.xls")])
    if not file_paths: return
        
    print(f"\nLoading {len(file_paths)} files...")
    all_dfs = []
    for fp in file_paths:
        df = pd.read_excel(fp, engine='openpyxl')
        if 'ROI Image' in df.columns: df = df.drop(columns=['ROI Image'])
        df = df[~df.iloc[:, 0].astype(str).str.contains("SUMMARY", na=False, case=False)]
        df = df[df.iloc[:, 0] != "Timestamp"]
        all_dfs.append(df)
        
    df_merged = pd.concat(all_dfs, ignore_index=True).dropna(how='all')
    df_merged['Verdict'] = df_merged['Verdict'].astype(str).str.strip().str.upper()
    
    # Purge "SKIP" Data
    df_merged = df_merged[df_merged['Verdict'] != 'SKIP']
    
    def get_type(error_msg):
        msg = str(error_msg).lower()
        if "rolling" in msg: return "Rolling"
        elif "truncated" in msg: return "Truncation"
        else: return "Normal"
            
    df_merged['Type'] = df_merged['Error Message'].apply(get_type) if 'Error Message' in df_merged.columns else "Normal"
    
    # Generate Root Cause Data
    df_merged['Failure_Root_Cause'] = df_merged['Error Message'].apply(categorize_root_cause)
    df_errors_only = df_merged[df_merged['Verdict'].isin(['FAIL', 'WARN'])]
    root_cause_stats = df_errors_only['Failure_Root_Cause'].value_counts().reset_index()
    root_cause_stats.columns = ['Error Category', 'Number of Occurrences']
    
    print("Generating Executive Statistics...")
    
    df_normal = df_merged[df_merged['Type'] == 'Normal']
    df_rolling = df_merged[df_merged['Type'] == 'Rolling']
    df_trunc = df_merged[df_merged['Type'] == 'Truncation']
    
    all_lang_stats = calculate_stats(df_merged, 'Language')
    normal_lang_stats = calculate_stats(df_normal, 'Language')
    device_stats = calculate_stats(df_merged, 'Device')
    
    snapshot_df = generate_historical_snapshot(df_merged, normal_lang_stats)
    
    output_file = "Executive_Verification_Analytics.xlsx"
    output_path = os.path.join(os.path.dirname(file_paths[0]), output_file)
    print(f"\nSaving Executive Report to: {output_path}")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        snapshot_df.to_excel(writer, sheet_name='Historical_Snapshot', index=False)
        autofit_columns(writer, snapshot_df, 'Historical_Snapshot')
        
        all_lang_stats.to_excel(writer, sheet_name='Lang_Stats_All', index=False)
        autofit_columns(writer, all_lang_stats, 'Lang_Stats_All')
        inject_dashboard(writer, df_merged, 'Lang_Stats_All', "Overall Language")
        
        # 🌟 NEW TAB: ROOT CAUSE ANALYSIS 🌟
        root_cause_stats.to_excel(writer, sheet_name='Error_Root_Cause_Analysis', index=False)
        autofit_columns(writer, root_cause_stats, 'Error_Root_Cause_Analysis')
        inject_bar_chart(writer, root_cause_stats, 'Error_Root_Cause_Analysis')
        
        device_stats.to_excel(writer, sheet_name='Device_Hardware_Stats', index=False)
        autofit_columns(writer, device_stats, 'Device_Hardware_Stats')
        inject_dashboard(writer, df_merged, 'Device_Hardware_Stats', "Device Hardware")
        
        df_normal.to_excel(writer, sheet_name='Raw_Data_Normal', index=False)
        autofit_columns(writer, df_normal, 'Raw_Data_Normal')
        
        if not df_rolling.empty:
            df_rolling.to_excel(writer, sheet_name='Raw_Data_Rolling', index=False)
            autofit_columns(writer, df_rolling, 'Raw_Data_Rolling')
            
        if not df_trunc.empty:
            df_trunc.to_excel(writer, sheet_name='Raw_Data_Truncation', index=False)
            autofit_columns(writer, df_trunc, 'Raw_Data_Truncation')
            
    print("\n✅ Done! Executive Dashboard with Root Cause Analysis generated successfully.")

if __name__ == "__main__":
    main()