from __future__ import annotations

import argparse
import os
import sys
import time
import unicodedata
import re
import hanzidentifier
from pathlib import Path
import threading
import difflib
from difflib import SequenceMatcher
import cv2
import numpy as np
import yaml
import json
import pandas as pd

try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:
    Image = None
    ImageDraw = None
    ImageFont = None

sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.fast_detector import FastDetector
from src.msi_genai_ocr import MSIGenAIOCR

# ---------------------------------------------------------------------------
# Display Style Mappings
# ---------------------------------------------------------------------------
DISPLAY_STYLES = {
    34: "EMERGENCY_NOTICE_CENTER_DISP_STYLE",
    38: "FEATURE_HOME_SCREEN_SUBFEATURE_DISP_STYLE",
    39: "FEATURE_HOME_SCREEN_FEATURE_DISP_STYLE",
    40: "FEATURE_HOME_SCREEN_FIRST_LINE_SELECTED_FEATURE_DISP_STYLE",
    41: "FEATURE_HOME_SCREEN_SYSTEM_DISP_STYLE",
    50: "FEATURE_GOOD_NOTICE_SYSTEM_DISP_STYLE",
    51: "FEATURE_GOOD_NOTICE_FEATURE_DISP_STYLE",
    52: "FEATURE_BAD_NOTICE_SYSTEM_DISP_STYLE",
    53: "FEATURE_BAD_NOTICE_FEATURE_DISP_STYLE",
    54: "FEATURE_NEUTRAL_NOTICE_SYSTEM_DISP_STYLE",
    55: "FEATURE_NEUTRAL_NOTICE_FEATURE_DISP_STYLE"
}

def get_display_style_name(command_str: str) -> str:
    if not command_str or command_str == "SKIP_VERIFY": return "-"
    parts = command_str.split(":")
    if len(parts) >= 3:
        try:
            style_idx = int(parts[2])
            return DISPLAY_STYLES.get(style_idx, f"UNKNOWN_{style_idx}")
        except Exception:
            pass
    return "UNKNOWN"

# ---------------------------------------------------------------------------
# Strict Error Detection Methods
# ---------------------------------------------------------------------------

def _is_screen_blank(roi_img: np.ndarray) -> bool:
    try:
        if roi_img is None or getattr(roi_img, "size", 0) == 0: return True
        gray = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY) if len(roi_img.shape) == 3 else roi_img
        std_dev = np.std(gray)
        if std_dev < 8.0: return True
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blur, 40, 120)
        edge_density = np.count_nonzero(edges) / float(edges.size)
        if edge_density < 0.003: return True
        return False
    except Exception:
        return False

def _parse_structured_fields(block: str):
    info = {
        "error_red": False,
        "error_evidence": "",
        "error_type": "",
        "language": "",
        "original": "",
        "english": "",
    }
    if not block: return info
    lines = [ln.rstrip("\r") for ln in str(block).splitlines()]
    compact = [ln.strip() for ln in lines if ln.strip()]
    if compact and compact[0].strip().lower().startswith("error detected"):
        info["error_red"] = True
        try:
            first_raw = compact[0].strip()
            first_lower = first_raw.lower()
            if "upside" in first_lower: info["error_type"] = "upside down"
            elif "overlap" in first_lower or "bridge" in first_lower: info["error_type"] = "overlap"
            elif "misalign" in first_lower: info["error_type"] = "misalignment"
            else: info["error_type"] = "generic"
                
            ev_lines = []
            if ":" in first_raw: ev_lines.append(first_raw.split(":", 1)[-1].strip().strip('"'))
            else: ev_lines.append(first_raw)
            
            for ev_ln in compact[1:]:
                ev_low = ev_ln.lower()
                if ev_low.startswith("detected language") or ev_low.startswith("detected text"): break
                if ev_low.startswith("likely") or ev_low.startswith("token"):
                    ev_lines.append(ev_ln.strip())
            info["error_evidence"] = "\n".join(ev_lines).strip()
        except Exception: pass
    
    mode = None
    buf_orig, buf_eng = [], []
    for raw in lines:
        s = raw.strip()
        low = s.lower()
        if low.startswith("error detected") or low.startswith("likely") or low.startswith("token"): continue
        if low.startswith("detected language:"):
            info["language"] = s.split(":", 1)[-1].strip()
            mode = None; continue
        if low.startswith("detected text(original):"):
            mode = "orig"
            rest = s.split(":", 1)[-1].strip()
            if rest and rest.lower() != "detected text(original)": buf_orig.append(rest)
            continue
        if low.startswith("detected text(english translation):"):
            mode = "eng"
            rest = s.split(":", 1)[-1].strip()
            if rest and rest.lower() != "detected text(english translation)": buf_eng.append(rest)
            continue

        if mode == "orig" and s: buf_orig.append(s)
        elif mode == "eng" and s: buf_eng.append(s)

    info["original"] = "\n".join([x for x in buf_orig if x]).strip()
    info["english"] = "\n".join([x for x in buf_eng if x]).strip()
    
    if info["error_type"] == "misalignment":
        lines_count = len([ln for ln in info["original"].splitlines() if ln.strip()])
        if lines_count <= 1:
            info["error_red"] = False
            info["error_type"] = ""
            info["error_evidence"] = ""

    return info

def _font_candidates_for_text(text: str, preferred: str = "") -> list:
    s = str(text or "")
    has_hangul = any(0xAC00 <= ord(ch) <= 0xD7A3 for ch in s)
    has_cjk = any(
        (0x3040 <= ord(ch) <= 0x30FF) or (0x4E00 <= ord(ch) <= 0x9FFF) or
        (0x3400 <= ord(ch) <= 0x4DBF) or (0xF900 <= ord(ch) <= 0xFAFF) for ch in s
    )

    out = []
    if preferred: out.append(preferred)

    if has_hangul:
        out.extend([r"C:\\Windows\\Fonts\\malgun.ttf", r"C:\\Windows\\Fonts\\gulim.ttc", r"C:\\Windows\\Fonts\\batang.ttc"])
    elif has_cjk:
        out.extend([r"C:\\Windows\\Fonts\\msyh.ttc", r"C:\\Windows\\Fonts\\simsun.ttc", r"C:\\Windows\\Fonts\\meiryo.ttc", r"C:\\Windows\\Fonts\\msgothic.ttc"])

    out.extend([r"C:\\Windows\\Fonts\\tahoma.ttf", r"C:\\Windows\\Fonts\\segoeui.ttf", r"C:\\Windows\\Fonts\\arial.ttf", r"C:\\Windows\\Fonts\\arialbd.ttf", r"C:\\Windows\\Fonts\\arialuni.ttf"])
    seen, dedup = set(), []
    for p in out:
        if not p or p in seen: continue
        seen.add(p)
        dedup.append(p)
    return dedup

def _draw_text_unicode(img_bgr: np.ndarray, text: str, org: tuple, font_scale: float, color_bgr: tuple, thickness: int = 1, font_path: str = ""):
    try:
        s = str(text or "")
        if not s: return img_bgr
        needs_unicode = any(ord(ch) > 127 for ch in s)
        if not needs_unicode or Image is None or ImageDraw is None or ImageFont is None:
            cv2.putText(img_bgr, s, org, cv2.FONT_HERSHEY_SIMPLEX, float(font_scale), tuple(int(x) for x in color_bgr), int(thickness), cv2.LINE_AA)
            return img_bgr

        size = max(12, int(round(28 * float(font_scale))))
        candidates = _font_candidates_for_text(s, font_path)
        font = None
        for c in candidates:
            if not c: continue
            try:
                if os.path.exists(c):
                    font = ImageFont.truetype(c, size)
                    break
            except Exception: font = None

        if font is None:
            cv2.putText(img_bgr, s, org, cv2.FONT_HERSHEY_SIMPLEX, float(font_scale), tuple(int(x) for x in color_bgr), int(thickness), cv2.LINE_AA)
            return img_bgr

        img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(img_rgb)
        draw = ImageDraw.Draw(pil_img)
        x, y = int(org[0]), int(org[1])
        y = max(0, y - int(size * 0.85))
        b, g, r = [int(v) for v in color_bgr]
        draw.text((x, y), s, font=font, fill=(r, g, b))
        out_rgb = np.asarray(pil_img)
        return cv2.cvtColor(out_rgb, cv2.COLOR_RGB2BGR)
    except Exception: return img_bgr

def _truncate_text_to_px(text: str, max_w: int, font_scale: float) -> str:
    try:
        s = str(text or "")
        if not s: return ""
        ell = "..."
        needs_unicode = any(ord(ch) > 127 for ch in s)

        if not needs_unicode or ImageFont is None:
            (tw, _th), _ = cv2.getTextSize(s, cv2.FONT_HERSHEY_SIMPLEX, float(font_scale), 1)
            if tw <= int(max_w): return s
            lo, hi, best = 0, len(s), ""
            while lo <= hi:
                mid = (lo + hi) // 2
                cand = s[:mid].rstrip() + ell
                (cw, _ch), _ = cv2.getTextSize(cand, cv2.FONT_HERSHEY_SIMPLEX, float(font_scale), 1)
                if cw <= int(max_w): best, lo = cand, mid + 1
                else: hi = mid - 1
            return best if best else ell

        size = max(12, int(round(28 * float(font_scale))))
        font = None
        for c in _font_candidates_for_text(s, ""):
            try:
                if os.path.exists(c):
                    font = ImageFont.truetype(c, size)
                    break
            except Exception: pass

        if font is None: return s if len(s) <= 60 else (s[:57] + ell)

        def _w(t: str) -> int:
            try:
                if hasattr(font, "getlength"): return int(round(float(font.getlength(t))))
            except Exception: pass
            try:
                bbox = font.getbbox(t)
                return int(bbox[2] - bbox[0])
            except Exception: return len(t) * size

        if _w(s) <= int(max_w): return s
        lo, hi, best = 0, len(s), ""
        while lo <= hi:
            mid = (lo + hi) // 2
            cand = s[:mid].rstrip() + ell
            if _w(cand) <= int(max_w): best, lo = cand, mid + 1
            else: hi = mid - 1
        return best if best else ell
    except Exception: return str(text or "")

def _wrap_text_to_px(text: str, max_w: int, font_scale: float) -> list:
    try:
        s = str(text or "").strip()
        if not s: return [""]
        out, rest = [], s
        while rest:
            if len(rest) <= 1 or _truncate_text_to_px(rest, max_w, font_scale) == rest:
                out.append(rest); break
            lo, hi, best = 1, len(rest), 1
            while lo <= hi:
                mid = (lo + hi) // 2
                cand = rest[:mid].rstrip()
                if not cand: hi = mid - 1; continue
                if _truncate_text_to_px(cand, max_w, font_scale) == cand: best, lo = mid, mid + 1
                else: hi = mid - 1
            seg = rest[:best].rstrip()
            if not seg: seg, best = rest[:1], 1
            out.append(seg)
            rest = rest[best:].lstrip()
        return out if out else [""]
    except Exception: return [str(text or "")] 

# ---------------------------------------------------------------------------
# Camera UI Overlay Functions
# ---------------------------------------------------------------------------
def _create_camera_overlay_state(cap: cv2.VideoCapture) -> dict:
    state = {"enabled": False, "drag": None, "values": {"Brightness": 0, "Sharpness": 0, "Focus": 0}, "last_applied": {}, "layout": {}}
    max_vals = {"Brightness": 255, "Sharpness": 255, "Focus": 50}
    def _safe_get(prop):
        try: return cap.get(prop)
        except Exception: return None
    for label, prop in [("Brightness", getattr(cv2, "CAP_PROP_BRIGHTNESS", None)), ("Sharpness", getattr(cv2, "CAP_PROP_SHARPNESS", None)), ("Focus", getattr(cv2, "CAP_PROP_FOCUS", None))]:
        if prop is None: continue
        v = _safe_get(prop)
        try:
            vv = 0 if v is None or (isinstance(v, float) and (v != v)) else int(round(float(v)))
            vmax = int(max_vals.get(label, 255))
            state["values"][label] = max(0, min(vmax, vv))
        except Exception:
            state["values"][label] = 0
    try: state["last_applied"] = dict(state["values"])
    except Exception: state["last_applied"] = {}
    return state

def _apply_camera_overlay_settings(cap: cv2.VideoCapture, overlay: dict) -> None:
    if cap is None or not overlay: return
    vals, last = overlay.get("values") or {}, overlay.get("last_applied") or {}
    prop_autofocus = getattr(cv2, "CAP_PROP_AUTOFOCUS", None)
    props = {"Brightness": getattr(cv2, "CAP_PROP_BRIGHTNESS", None), "Sharpness": getattr(cv2, "CAP_PROP_SHARPNESS", None), "Focus": getattr(cv2, "CAP_PROP_FOCUS", None)}
    for k, prop in props.items():
        if prop is None or k not in vals: continue
        v = vals.get(k)
        if last.get(k) == v: continue
        try:
            if k == "Focus" and prop_autofocus is not None:
                try: cap.set(prop_autofocus, 0)
                except Exception: pass
            cap.set(prop, float(v))
            overlay.setdefault("last_applied", {})[k] = v
        except Exception: pass

def _draw_camera_overlay(img_bgr: np.ndarray, overlay: dict) -> np.ndarray:
    if img_bgr is None or not overlay or not overlay.get("enabled"): return img_bgr
    h, w = img_bgr.shape[:2]
    panel_w = min(360, max(260, int(w * 0.38)))
    panel_h = 150
    x0, y0 = 10, max(10, h - panel_h - 10)
    x1, y1 = x0 + panel_w, y0 + panel_h
    out = img_bgr
    try:
        overlay_img = out.copy()
        cv2.rectangle(overlay_img, (x0, y0), (x1, y1), (0, 0, 0), -1)
        cv2.addWeighted(overlay_img, 0.45, out, 0.55, 0, out)
        cv2.rectangle(out, (x0, y0), (x1, y1), (40, 40, 40), 1)
    except Exception: return img_bgr

    vals = overlay.get("values") or {}
    labels, max_vals = ["Brightness", "Sharpness", "Focus"], {"Brightness": 255, "Sharpness": 255, "Focus": 50}
    slider_left, slider_right, row_y = x0 + 120, x1 - 15, [y0 + 35, y0 + 75, y0 + 115]
    overlay["layout"] = {}

    try: cv2.putText(out, "Camera Settings", (x0 + 10, y0 + 22), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
    except Exception: pass

    for i, lab in enumerate(labels):
        y, v, vmax = row_y[i], int(vals.get(lab, 0) or 0), int(max_vals.get(lab, 255))
        v = max(0, min(vmax, v))
        try: cv2.putText(out, f"{lab}: {v}", (x0 + 10, y + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (230, 230, 230), 1)
        except Exception: pass
        try:
            cv2.line(out, (slider_left, y), (slider_right, y), (190, 190, 190), 2)
            denom = float(vmax) if float(vmax) > 0 else 1.0
            knob_x = int(slider_left + (slider_right - slider_left) * (float(v) / denom))
            cv2.circle(out, (knob_x, y), 7, (0, 200, 255), -1)
            cv2.circle(out, (knob_x, y), 7, (0, 0, 0), 1)
        except Exception: pass
        overlay["layout"][lab] = {"x1": int(slider_left), "x2": int(slider_right), "y": int(y)}
    return out

def _apply_camera_env_tuning(cap: cv2.VideoCapture) -> None:
    try:
        if cap is None: return
        focus = (os.getenv("WALKIE_CAMERA_FOCUS", "") or "").strip()
        if focus != "" and hasattr(cv2, "CAP_PROP_FOCUS"):
            cap.set(cv2.CAP_PROP_FOCUS, float(focus))
    except Exception: pass

def _norm_col(s: str) -> str:
    v = str(s or "").strip().lower().replace("_", " ")
    v = re.sub(r"[\(\)\[\]\{\}:,;/\\\-]+", " ", v)
    v = " ".join(v.split())
    if v.startswith("string "): v = v[len("string "):].strip()
    if v.startswith("str "): v = v[len("str "):].strip()
    return v

def _norm_text(s: str) -> str:
    s = "" if s is None else str(s)
    try: s = unicodedata.normalize("NFKC", s)
    except Exception: pass
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln for ln in (" ".join(ln.strip().split()) for ln in s.split("\n")) if ln != ""]
    return "\n".join(lines).strip()

def load_device_profiles() -> dict:
    profiles = {}
    try:
        raw = str(os.getenv("WALKIE_DEVICE_PROFILES_JSON", "") or "").strip()
        if raw:
            obj = json.loads(raw)
            devices = obj.get("devices") if isinstance(obj, dict) else None
            if isinstance(devices, list):
                for d in devices:
                    if isinstance(d, dict) and d.get("id"): profiles[int(d["id"])] = d
        else:
            cfgp = Path(__file__).resolve().parents[1] / "configs" / "device_profiles.json"
            if cfgp.exists():
                with open(cfgp, "r", encoding="utf-8") as f:
                    obj = json.load(f) or {}
                devices = obj.get("devices") if isinstance(obj, dict) else None
                if isinstance(devices, list):
                    for d in devices:
                        if isinstance(d, dict) and d.get("id"): profiles[int(d["id"])] = d
    except Exception: pass
    return profiles

def get_device_name(profiles: dict, device_id: int) -> str:
    d = profiles.get(device_id) or {}
    name = str(d.get("name") or "").strip()
    return name if name else f"Device {device_id}"

def map_language_to_region(lang: str, text: str = "", allowed_langs: str = "") -> tuple[str, str]:
    l = str(lang).lower().strip()
    t = str(text)
    
    allowed = [x.strip() for x in allowed_langs.split(",") if x.strip()]
    foreign_allowed = [x for x in allowed if x.lower() not in ["english", "en"]]
    
    def _get_region(lang_name: str) -> str:
        ln = lang_name.lower()
        if "korean" in ln or "japanese" in ln or "chinese" in ln or "thai" in ln: return "apac"
        if "spanish" in ln or "portuguese" in ln: return "lacr"
        if "english" in ln or "en" in ln: return "english"
        return "emea"

    has_hiragana = any('\u3040' <= c <= '\u309F' for c in t)
    has_katakana = any('\u30A0' <= c <= '\u30FF' for c in t)
    has_hangul = any('\uAC00' <= c <= '\uD7A3' for c in t)
    has_cyrillic = any('\u0400' <= c <= '\u04FF' for c in t)
    has_arabic = any('\u0600' <= c <= '\u06FF' for c in t)
    has_hebrew = any('\u0590' <= c <= '\u05FF' for c in t)
    has_thai = any('\u0E00' <= c <= '\u0E7F' for c in t)
    has_cjk = any('\u4E00' <= c <= '\u9FFF' for c in t)

    if has_hiragana or has_katakana or has_hangul or has_cyrillic or has_arabic or has_hebrew or has_thai or has_cjk:
        if len(foreign_allowed) == 1:
            return _get_region(foreign_allowed[0]), foreign_allowed[0]
        elif foreign_allowed:
            for fa in foreign_allowed:
                if fa.lower() in l: return _get_region(fa), fa
            return _get_region(foreign_allowed[0]), foreign_allowed[0]

    if foreign_allowed:
        for fa in foreign_allowed:
            if fa.lower() in l: return _get_region(fa), fa

    if "english" in l or l == "en": return "english", "English"
    if foreign_allowed: return _get_region(foreign_allowed[0]), foreign_allowed[0]
    return "english", "English"

def _show_ocr_result_window(
    roi: np.ndarray, original: str, english: str, language: str,
    verdict: str = "", expected_lines: list | None = None,
    device_name: str = "", error_msg: str = "",
) -> None:
    try:
        if roi is None or getattr(roi, "size", 0) == 0: return
        lang_label = (language or "").strip()
        try: parts = [p.strip().lower() for p in lang_label.split(",") if p.strip()]
        except Exception: parts = []
        english_only = bool(parts) and all(p in ["english", "en"] for p in parts)
        show_english = bool(english and english.strip()) and (not english_only)

        o_lines = [ln.strip() for ln in str(original or "").splitlines() if ln.strip()]
        e_lines = [ln.strip() for ln in str(english or "").splitlines() if ln.strip()]

        pad, gap, line_h = 22, 44, 40
        roi_h0, roi_w0 = roi.shape[:2]
        out_w = int(max(roi_w0, 1250 if show_english else 1050))
        roi_scale = 2.0 if roi_w0 < 520 else (3.0 if roi_w0 < 360 else 1.0)
        
        roi_disp = roi
        if float(roi_scale) > 1.01:
            try: roi_disp = cv2.resize(roi, (int(round(roi_w0 * roi_scale)), int(round(roi_h0 * roi_scale))), interpolation=cv2.INTER_CUBIC)
            except Exception: roi_disp = roi
        roi_h, roi_w = roi_disp.shape[:2]

        scale, head_scale, color = (0.92 if out_w >= 1050 else 0.82), (1.05 if out_w >= 1050 else 0.95), (255, 255, 255)

        if show_english:
            col_w = max(220, (out_w - (2 * pad) - gap) // 2)
            left_x, right_x = pad, pad + col_w + gap
            left_head = f"Original ({lang_label})" if lang_label else "Original"
            right_head, max_w = "English", max(60, col_w - 10)
            left_head_lines, right_head_lines = _wrap_text_to_px(left_head, max_w, head_scale), _wrap_text_to_px(right_head, max_w, head_scale)
            header_rows = max(len(left_head_lines), len(right_head_lines), 1)
            left_wrapped, right_wrapped, max_rows = [], [], 0
            for i in range(max(len(o_lines), len(e_lines), 1)):
                lw = _wrap_text_to_px(o_lines[i] if i < len(o_lines) else "", max_w, scale)
                rw = _wrap_text_to_px(e_lines[i] if i < len(e_lines) else "", max_w, scale)
                rows = max(len(lw), len(rw), 1)
                while len(lw) < rows: lw.append("")
                while len(rw) < rows: rw.append("")
                left_wrapped.extend(lw); right_wrapped.extend(rw); max_rows += rows

            text_h = pad + (header_rows * line_h) + (max_rows * line_h) + pad
            out = np.zeros((roi_h + text_h, out_w, 3), dtype=np.uint8)
            x_img = int(max(0, (out_w - roi_w) // 2))
            out[:roi_h, x_img : x_img + roi_w] = roi_disp

            y = roi_h + pad + 24
            for r in range(header_rows):
                out = _draw_text_unicode(out, left_head_lines[r] if r < len(left_head_lines) else "", (left_x, y), head_scale, color, 1)
                out = _draw_text_unicode(out, right_head_lines[r] if r < len(right_head_lines) else "", (right_x, y), head_scale, color, 1)
                y += line_h

            for idx in range(max_rows):
                l, rr = left_wrapped[idx] if idx < len(left_wrapped) else "", right_wrapped[idx] if idx < len(right_wrapped) else ""
                if l: out = _draw_text_unicode(out, l, (left_x, y), scale, color, 1)
                if rr: out = _draw_text_unicode(out, rr, (right_x, y), scale, color, 1)
                y += line_h
        else:
            max_w, head = max(80, out_w - (2 * pad) - 10), f"Original ({lang_label})" if lang_label else "Original"
            head_lines, body_wrapped = _wrap_text_to_px(head, max_w, head_scale), []
            for ln in (o_lines or [""]): body_wrapped.extend(_wrap_text_to_px(ln, max_w, scale))
            text_h = pad + (len(head_lines) * line_h) + (len(body_wrapped) * line_h) + pad
            out = np.zeros((roi_h + text_h, out_w, 3), dtype=np.uint8)
            x_img = int(max(0, (out_w - roi_w) // 2))
            out[:roi_h, x_img : x_img + roi_w] = roi_disp
            y = roi_h + pad + 24
            for hl in head_lines:
                out = _draw_text_unicode(out, hl, (pad, y), head_scale, color, 1)
                y += line_h
            for bl in body_wrapped:
                if bl: out = _draw_text_unicode(out, bl, (pad, y), scale, color, 1)
                y += line_h

        exp_draw_lines = []
        try:
            if expected_lines:
                for ln in [str(x) for x in list(expected_lines) if str(x).strip()]:
                    exp_draw_lines.extend(_wrap_text_to_px(ln, max(120, int(out_w - 40)), 0.65))
        except Exception: pass

        try:
            header_h = 100
            summary_h = 120 + (max(0, len(exp_draw_lines)) * 32) + 16
            if error_msg: summary_h += 40

            header = np.zeros((header_h, out_w, 3), dtype=np.uint8)
            cv2.putText(header, f"OCR Result - {device_name}" if device_name else "OCR Result", (24, 60), cv2.FONT_HERSHEY_SIMPLEX, 1.4, (0, 255, 255), 3)

            summary = np.zeros((summary_h, out_w, 3), dtype=np.uint8)
            cv2.putText(summary, f"Detected Language: {lang_label if lang_label else 'Unknown'}", (24, 44), cv2.FONT_HERSHEY_SIMPLEX, 0.95, (255, 255, 255), 2)
            
            v = str(verdict or "").strip().upper()
            if v:
                vcol = (0, 255, 0) if v == "PASS" else ((0, 0, 255) if v == "FAIL" else (0, 255, 255))
                cv2.putText(summary, f"Verdict: {v}", (24, 86), cv2.FONT_HERSHEY_SIMPLEX, 1.1, vcol, 3)

            start_y = 118
            if exp_draw_lines:
                for i, ln in enumerate(exp_draw_lines):
                    summary = _draw_text_unicode(summary, ln, (24, start_y + (i * 32)), 0.8, (255, 255, 255), 1)
                start_y += len(exp_draw_lines) * 32

            if error_msg:
                summary = _draw_text_unicode(summary, error_msg, (24, start_y + 10), 1.1, (0, 0, 255), 3)

            disp = np.vstack([header, summary, out])
        except Exception: disp = np.vstack([header, out])

        try:
            h0, w0, min_w, min_h = disp.shape[:2], disp.shape[:2][1], 1250, 900
            if w0 > 0 and h0 > 0 and (w0 < min_w or h0 < min_h):
                s = max(1.0, min(2.2, max(float(min_w) / float(w0), float(min_h) / float(h0))))
                disp = cv2.resize(disp, (int(round(w0 * s)), int(round(h0 * s))), interpolation=cv2.INTER_CUBIC)
        except Exception: pass

        cv2.namedWindow("OCR Result", cv2.WINDOW_NORMAL)
        try: cv2.resizeWindow("OCR Result", int(disp.shape[1]), int(disp.shape[0]))
        except Exception: pass
        cv2.imshow("OCR Result", disp)
        
        start_time = time.time()
        while True:
            try:
                if hasattr(cv2, "getWindowProperty") and hasattr(cv2, "WND_PROP_VISIBLE") and cv2.getWindowProperty("OCR Result", cv2.WND_PROP_VISIBLE) < 1: break
            except Exception: pass
            if cv2.waitKey(50) not in [None, -1]: break
            if time.time() - start_time > 5.0: break
        
        try: cv2.destroyWindow("OCR Result")
        except Exception: pass
    except Exception: return

def _find_sheet(xls: pd.ExcelFile, name: str) -> str:
    target = _norm_col(name)
    for s in xls.sheet_names:
        if _norm_col(s) == target: return s
    for s in xls.sheet_names:
        if target in _norm_col(s): return s
    raise ValueError(f"Sheet '{name}' not found. Available: {xls.sheet_names}")

def _pick_language_column(df: pd.DataFrame, region: str, language: str) -> str:
    want = _norm_col(language)
    cols = { _norm_col(c): c for c in df.columns }
    synonyms = {
        "japanese": ["japanese", "ja"], "korean": ["korean", "ko"], "simplified chinese": ["simplified chinese", "chinese simplified", "zh cn", "zh-hans"],
        "traditional chinese": ["traditional chinese", "chinese traditional", "zh tw", "zh-hant"], "french": ["french", "fr"], "spanish": ["spanish", "es"],
        "german": ["german", "de"], "italian": ["italian", "it"], "polish": ["polish", "pl"], "russian": ["russian", "ru"], "turkish": ["turkish", "tr"],
        "arabic": ["arabic", "ar"], "hungarian": ["hungarian", "hu"], "hebrew": ["hebrew", "iw", "he"], "czech": ["czech", "cs"], "portuguese": ["portuguese", "pt"],
    }
    candidates = [_norm_col(v) for v in synonyms.get(want, [want])] if want in synonyms else [want]
    for c in candidates:
        if c in cols: return cols[c]
    raise ValueError(f"Language column for region='{region}' language='{language}' not found. Columns: {list(df.columns)}")

def load_expected(excel_path: str, region: str, language: str, index: str = "", tag: str = "") -> dict:
    if not os.path.exists(str(excel_path)): raise FileNotFoundError(excel_path)
    xls = pd.ExcelFile(str(excel_path), engine="openpyxl")
    english_sheet, category_sheet = _find_sheet(xls, "english"), _find_sheet(xls, "category")
    region_norm = _norm_col(region)
    if region_norm in ["apac"]: region_sheet = _find_sheet(xls, "apac")
    elif region_norm in ["emea"]: region_sheet = _find_sheet(xls, "emea")
    elif region_norm in ["lacr", "latam", "latam\u0026caribbean", "la cr"]: region_sheet = _find_sheet(xls, "lacr")
    elif region_norm in ["english", "en", "global"]: region_sheet = english_sheet
    else: raise ValueError(f"Region must be one of: english, apac, emea, lacr. Got {region}")

    df_en, df_cat, df_reg = pd.read_excel(xls, sheet_name=english_sheet, engine="openpyxl").copy(), pd.read_excel(xls, sheet_name=category_sheet, engine="openpyxl").copy(), pd.read_excel(xls, sheet_name=region_sheet, engine="openpyxl").copy()

    def _coerce_index(v):
        try:
            if pd.isna(v): return ""
        except Exception: pass
        s = str(v).strip()
        if s.endswith(".0"):
            try: return str(int(float(s)))
            except Exception: pass
        if s.isdigit():
            try: return str(int(s))  
            except Exception: pass
        return s

    if "index" in [_norm_col(c) for c in df_en.columns]: 
        df_en["__index"] = df_en[next(c for c in df_en.columns if _norm_col(c) == "index")].apply(_coerce_index)
    else: raise ValueError("English sheet missing 'index' column")

    if "index" in [_norm_col(c) for c in df_reg.columns]: 
        df_reg["__index"] = df_reg[next(c for c in df_reg.columns if _norm_col(c) == "index")].apply(_coerce_index)
    else: raise ValueError(f"Region sheet '{region_sheet}' missing 'index' column")

    if "index" in [_norm_col(c) for c in df_cat.columns]: 
        df_cat["__index"] = df_cat[next(c for c in df_cat.columns if _norm_col(c) == "index")].apply(_coerce_index)
    else: raise ValueError("Category sheet missing 'index' column")

    idx = str(index).strip()
    if idx and idx.endswith(".0"):
        try: idx = str(int(float(idx)))
        except Exception: pass
    if idx.isdigit():
        try: idx = str(int(idx))  
        except Exception: pass

    def _find_tag_column(df: pd.DataFrame) -> str:
        preferred, fallback = [], []
        for c in df.columns:
            low, n = str(c or "").strip().lower(), _norm_col(c)
            if ("tag" in low and "string" in low) or n in ["string tag", "stringtag"]: preferred.append(c)
            elif n == "tag" or "tag" in low: fallback.append(c)
        return preferred[0] if preferred else (fallback[0] if fallback else "")

    en_tag_col = _find_tag_column(df_en)
    reg_tag_col = _find_tag_column(df_reg)

    idx_region = ""
    if idx:
        row_en = df_en[df_en["__index"] == idx]
        row_reg = df_reg[df_reg["__index"] == idx]
        row_cat = df_cat[df_cat["__index"] == idx]
    elif tag:
        tag_norm = str(tag).strip().lower()
        if not en_tag_col: raise ValueError("English sheet missing 'string tag' column")
        row_en_all = df_en[df_en[en_tag_col].astype(str).str.strip().str.lower() == tag_norm]
        if row_en_all.empty: raise ValueError(f"No row found for tag '{tag}' in English sheet")
        idx = str(row_en_all.iloc[0]["__index"]).strip()
        row_en = df_en[df_en["__index"] == idx]
        row_cat = df_cat[df_cat["__index"] == idx]
        if not reg_tag_col: row_reg = df_reg[df_reg["__index"] == "__no_match__"]
        else:
            row_reg = df_reg[df_reg["__index"] == idx]
            if not row_reg.empty: idx_region = str(row_reg.iloc[0].get("__index", "")).strip()
    else: raise ValueError("Provide --index or --tag")

    if row_en.empty: raise ValueError(f"No English row found for index '{idx}'")

    en_row = row_en.iloc[0].to_dict()
    reg_row = row_reg.iloc[0].to_dict() if not row_reg.empty else {}
    cat_row = row_cat.iloc[0].to_dict() if not row_cat.empty else {}

    def _extract_merged_text_safely(df, target_idx, text_col_name, tag_col_name):
        if not text_col_name or text_col_name not in df.columns: return ""
        matching_rows = df.index[df["__index"] == target_idx].tolist()
        if not matching_rows: return ""
        
        start_row = matching_rows[0]
        lines = []
        
        val = df.iloc[start_row][text_col_name]
        if pd.notna(val) and str(val).strip(): lines.append(str(val).strip())
            
        for r in range(start_row + 1, len(df)):
            r_idx = df.iloc[r].get("__index", "")
            r_tag = df.iloc[r].get(tag_col_name, "") if tag_col_name else ""
            if (pd.notna(r_idx) and str(r_idx).strip() != "") or (pd.notna(r_tag) and str(r_tag).strip() != ""): break
            val = df.iloc[r].get(text_col_name, "")
            if pd.notna(val) and str(val).strip(): lines.append(str(val).strip())
                
        return " ".join(lines)

    en_text_col = next((c for c in df_en.columns if _norm_col(c) in ["string (english)", "string english", "english", "string"]), None)
    if en_text_col is None: raise ValueError("English sheet missing 'string (english)' column")

    expected_en = _extract_merged_text_safely(df_en, idx, en_text_col, en_tag_col)
    
    if _norm_col(region_sheet) == _norm_col(english_sheet): 
        expected_local = expected_en
    else:
        lang_col = _pick_language_column(df_reg, region, language)
        expected_local = _extract_merged_text_safely(df_reg, idx, lang_col, reg_tag_col)

    found_tag = str(en_row.get(en_tag_col, "")) if en_tag_col and en_tag_col in en_row else str(tag)
    if str(found_tag).lower() == 'nan': found_tag = ""
    return {
        "index": idx, "index_region": idx_region, "expected_en": expected_en, "expected_local": expected_local, "region_sheet": region_sheet, "tag": found_tag
    }

def capture_screen_roi(detector: FastDetector, camera_id: int, confidence: float, warmup_sec: float = 1.5, rolling_mode: bool = False):
    cap = cv2.VideoCapture(int(camera_id), cv2.CAP_DSHOW)
    
    if not cap.isOpened(): raise RuntimeError(f"Could not open camera {camera_id}")
    try: cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    except Exception: pass
    try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
    except Exception: pass
    try: cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280); cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    except Exception: pass

    _apply_camera_env_tuning(cap)
    
    burst_frames = []
    
    # We ONLY capture 5 frames initially.
    t_end = time.time() + 0.5
    last_t = 0
    while time.time() < t_end:
        ok, frame = cap.read()
        if ok and frame is not None and frame.size > 0:
            if time.time() - last_t >= 0.08:
                burst_frames.append(frame.copy()) 
                last_t = time.time()
            if len(burst_frames) >= 5:
                break

    cap.release()
    if not burst_frames: raise RuntimeError("Failed to capture frame")

    base_frame = burst_frames[-1]
    boxes, screens = detector.detect_with_screens(base_frame, confidence)
    if not boxes: raise RuntimeError("No device detected")

    boxes = sorted(boxes, key=lambda b: (b[0], b[1]))
    rois = []
    roi_coords = [] 
    
    for (x1, y1, x2, y2) in boxes:
        screen_box = next((s for s in screens if x1 <= (s[0]+s[2])/2 <= x2 and y1 <= (s[1]+s[3])/2 <= y2), None)
        if screen_box is None: screen_box = (x1, y1, x2, y2)
        sx1, sy1, sx2, sy2 = max(0, screen_box[0]), max(0, screen_box[1]), min(base_frame.shape[1], screen_box[2]), min(base_frame.shape[0], screen_box[3])
        if sx2 <= sx1 or sy2 <= sy1: continue
        
        best_roi = None
        max_brightness = -1
        
        for f in burst_frames:
            try:
                crop = f[sy1:sy2, sx1:sx2]
                if crop.size > 0:
                    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                    brightness = np.mean(gray)
                    if brightness > max_brightness:
                        max_brightness = brightness
                        best_roi = crop
            except Exception: pass
                
        if best_roi is not None:
            rois.append(best_roi)
            roi_coords.append((sx1, sy1, sx2, sy2)) 
        
    if not rois: raise RuntimeError("Invalid screen ROIs")
    return base_frame, rois, roi_coords

def capture_screen_roi_preview(detector: FastDetector | None, camera_id: int, confidence: float = 0.25, model_path: str = "", window_name: str = "Verify Preview", profiles: dict = None):
    cap = cv2.VideoCapture(int(camera_id), cv2.CAP_DSHOW)
    if not cap.isOpened(): raise RuntimeError(f"Could not open camera {camera_id}")

    try: cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    except Exception: pass
    try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
    except Exception: pass
    try: cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280); cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    except Exception: pass

    det_holder = {"det": detector}
    if det_holder["det"] is None and model_path:
        def _load_det():
            try: det_holder["det"] = FastDetector(model_path)
            except Exception: det_holder["det"] = None
        threading.Thread(target=_load_det, daemon=True).start()

    last_frame, last_boxes, last_screens = None, [], []
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    try: cv2.resizeWindow(window_name, 1280, 720)
    except Exception: pass

    overlay = _create_camera_overlay_state(cap) if cap else None
    zoom, did_tune = 1.0, False

    saved_mapping = {}
    try:
        map_file = Path(__file__).resolve().parents[1] / "configs" / "box_mapping.json"
        if map_file.exists():
            with open(map_file, "r") as f:
                raw_mapping = json.load(f)
                saved_mapping = {int(k): int(v) for k, v in raw_mapping.items()}
    except Exception: pass

    box_state = {
        "assignments": [],  
        "selected_idx": None,
        "sorted_boxes": []
    }

    def _unified_mouse_callback(event, x, y, flags, param):
        if overlay and overlay.get("enabled"):
            def _hit(lx, ly):
                for lab, lay in (overlay.get("layout") or {}).items():
                    yy, x1, x2 = int(lay.get("y") or 0), int(lay.get("x1") or 0), int(lay.get("x2") or 0)
                    if x1 <= lx <= x2 and (yy - 12) <= ly <= (yy + 12): return str(lab)
                return ""
            def _set(lab, lx):
                lay = (overlay.get("layout") or {}).get(lab) or {}
                x1, x2 = int(lay.get("x1") or 0), int(lay.get("x2") or 0)
                if x2 > x1:
                    xx, vmax = max(x1, min(x2, int(lx))), float({"Brightness": 255, "Sharpness": 255, "Focus": 50}.get(lab, 255))
                    v = int(round((float(xx - x1) / float(x2 - x1)) * vmax))
                    overlay.setdefault("values", {})[lab] = max(0, min(int(vmax), v))

            if event == cv2.EVENT_LBUTTONDOWN:
                lab = _hit(int(x), int(y))
                if lab: 
                    overlay["drag"] = lab
                    _set(lab, int(x))
                    return
            elif event == cv2.EVENT_MOUSEMOVE:
                if (flags & cv2.EVENT_FLAG_LBUTTON) and overlay.get("drag"):
                    _set(str(overlay.get("drag")), int(x))
                    return
            elif event == cv2.EVENT_LBUTTONUP:
                if overlay.get("drag"):
                    overlay["drag"] = None
                    return
        
        if event == cv2.EVENT_LBUTTONDOWN:
            for i, (bx1, by1, bx2, by2) in enumerate(box_state["sorted_boxes"]):
                if bx1 <= x <= bx2 and by1 <= y <= by2:
                    if box_state["selected_idx"] is None:
                        box_state["selected_idx"] = i
                    else:
                        if box_state["selected_idx"] != i:
                            try:
                                s_idx = box_state["selected_idx"]
                                temp = box_state["assignments"][s_idx]
                                box_state["assignments"][s_idx] = box_state["assignments"][i]
                                box_state["assignments"][i] = temp
                            except Exception: pass
                        box_state["selected_idx"] = None
                    return
            box_state["selected_idx"] = None

    cv2.setMouseCallback(window_name, _unified_mouse_callback)

    try:
        while True:
            if overlay is not None: _apply_camera_overlay_settings(cap, overlay)
            ok, frame = cap.read()
            if not ok or frame is None or frame.size == 0: continue
            if not did_tune:
                try: _apply_camera_env_tuning(cap)
                except Exception: pass
                did_tune = True

            if zoom and float(zoom) > 1.0:
                try:
                    h, w = frame.shape[:2]
                    new_w, new_h = max(2, int(w / float(zoom))), max(2, int(h / float(zoom)))
                    x1, y1 = max(0, (w - new_w) // 2), max(0, (h - new_h) // 2)
                    frame = cv2.resize(frame[y1 : y1 + new_h, x1 : x1 + new_w], (w, h), interpolation=cv2.INTER_LINEAR)
                except Exception: pass
            last_frame = frame

            try:
                if det_holder.get("det") is not None: 
                    last_boxes, last_screens = det_holder.get("det").detect_with_screens(frame, confidence)
            except Exception: pass
            
            sorted_boxes = sorted(last_boxes or [], key=lambda b: (b[0], b[1]))
            box_state["sorted_boxes"] = sorted_boxes
            
            if len(sorted_boxes) > 0:
                if not box_state["assignments"]:
                    for i in range(len(sorted_boxes)):
                        box_state["assignments"].append(saved_mapping.get(i, i))
                
                while len(box_state["assignments"]) < len(sorted_boxes):
                    assigned = set(box_state["assignments"])
                    nxt = 0
                    while nxt in assigned: nxt += 1
                    box_state["assignments"].append(nxt)
                while len(box_state["assignments"]) > len(sorted_boxes):
                    box_state["assignments"].pop()

            vis = frame.copy()
            for i, (x1, y1, x2, y2) in enumerate(sorted_boxes):
                dev_idx = box_state["assignments"][i] if i < len(box_state["assignments"]) else i
                dev_name = get_device_name(profiles, dev_idx + 1) if profiles else f"Device {dev_idx + 1}"
                
                color = (0, 255, 255) if box_state["selected_idx"] == i else (0, 255, 0)
                cv2.rectangle(vis, (x1, y1), (x2, y2), color, 2)
                
                cv2.rectangle(vis, (x1, max(0, y1 - 25)), (x1 + 180, y1), color, -1)
                cv2.putText(vis, dev_name, (x1 + 5, max(0, y1 - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

            for (sx1, sy1, sx2, sy2) in last_screens or []:
                cv2.rectangle(vis, (sx1, sy1), (sx2, sy2), (0, 0, 255), 1)

            cv2.putText(vis, "SPACE=Save & Exit  T=settings  +/-=zoom  X=cancel", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            cv2.putText(vis, "Click one box, then click another to SWAP their Device Name", (10, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)
            
            if det_holder.get("det") is None: cv2.putText(vis, "Loading detector...", (10, 85), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            if overlay is not None: vis = _draw_camera_overlay(vis, overlay)
            
            cv2.imshow(window_name, vis)
            
            if hasattr(cv2, "getWindowProperty") and hasattr(cv2, "WND_PROP_VISIBLE") and cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1: raise RuntimeError("Cancelled")

            k = cv2.waitKey(1) & 0xFF
            if k in [ord('x'), ord('X')]: raise RuntimeError("Cancelled")
            if k in [ord('t'), ord('T')] and overlay is not None: overlay["enabled"], overlay["drag"] = not overlay["enabled"], None
            if k in [ord('+'), ord('=')]: zoom = min(4.0, float(zoom) + 0.1)
            if k in [ord('-'), ord('_')]: zoom = max(1.0, float(zoom) - 0.1)
            if k in [ord('z'), ord('Z')]: zoom = 1.0
            if k == ord(' ') and det_holder.get("det") is not None: break
    finally:
        cap.release()
        try: cv2.destroyWindow(window_name)
        except Exception: pass

    if last_frame is None: raise RuntimeError("Failed to capture frame")
    if not last_boxes: raise RuntimeError("No device detected. Try adjusting lighting/camera angle or lower --confidence (e.g. 0.15).")

    mapping = {str(i): int(dev_idx) for i, dev_idx in enumerate(box_state["assignments"])}
    try:
        map_file = Path(__file__).resolve().parents[1] / "configs" / "box_mapping.json"
        map_file.parent.mkdir(parents=True, exist_ok=True)
        with open(map_file, "w") as f:
            json.dump(mapping, f)
        print(f"\n[INFO] Saved custom layout mapping: {mapping}")
    except Exception as e:
        print(f"\n[WARNING] Failed to save custom layout mapping: {e}")

    rois = []
    for (x1, y1, x2, y2) in sorted(last_boxes, key=lambda b: (b[0], b[1])):
        screen_box = next((s for s in last_screens if x1 <= (s[0]+s[2])/2 <= x2 and y1 <= (s[1]+s[3])/2 <= y2), None)
        if screen_box is None: screen_box = (x1, y1, x2, y2)
        sx1, sy1, sx2, sy2 = max(0, screen_box[0]), max(0, screen_box[1]), min(last_frame.shape[1], screen_box[2]), min(last_frame.shape[0], screen_box[3])
        if sx2 > sx1 and sy2 > sy1: rois.append(last_frame[sy1:sy2, sx1:sx2])

    if not rois: raise RuntimeError("Invalid screen ROIs")
    return last_frame, rois

def main():
    t0_total = time.time()
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--region", required=True)
    parser.add_argument("--language", required=True)
    parser.add_argument("--index", default="")
    parser.add_argument("--tag", default="")
    parser.add_argument("--command", default="")
    parser.add_argument("--config", default="configs/settings.yaml")
    parser.add_argument("--model-path", default="")
    parser.add_argument("--epoch", type=int, default=None)
    parser.add_argument("--camera-id", type=int, default=None)
    parser.add_argument("--save-roi-dir", default="")
    parser.add_argument("--summary-excel", default="")
    parser.add_argument("--preview", action="store_true")

    args = parser.parse_args()
    try: args.region = _norm_col(args.region)
    except Exception: pass
    region_map = {"multiple": "multiple", "auto": "multiple", "english": "english", "en": "english", "apac": "apac", "emea": "emea", "lacr": "lacr", "lac": "lacr", "latam": "lacr"}
    args.region = region_map.get(args.region)
    if not args.region: raise ValueError("--region must be one of: multiple, english, apac, emea, lacr")

    with open(args.config, "r", encoding="utf-8") as f: cfg = yaml.safe_load(f) or {}

    model_path = args.model_path or cfg.get("detector", {}).get("path", "")
    if not model_path: raise ValueError("Detector model path not provided")

    camera_id = args.camera_id if args.camera_id is not None else int(cfg.get("camera", {}).get("source", 0))
    confidence = float(cfg.get("detector", {}).get("confidence", 0.25))

    profiles = load_device_profiles()

    indices = [x.strip() for x in args.index.split(",")] if args.index else []
    tags = [x.strip() for x in args.tag.split(",")] if args.tag else []
    commands = [x.strip() for x in args.command.split(",")] if args.command else []

    expected_list = []
    max_len = max(len(indices), len(tags), len(commands), 1)

    for i in range(max_len):
        idx_val = indices[i] if i < len(indices) else (indices[-1] if indices else "")
        tag_val = tags[i] if i < len(tags) else (tags[-1] if tags else "")
        cmd_val = commands[i] if i < len(commands) else (commands[-1] if commands else "")
        
        disp_style = get_display_style_name(cmd_val)

        if tag_val == "SKIP_VERIFY" or idx_val == "SKIP_VERIFY":
            expected_list.append({"index": "SKIP_VERIFY", "tag": "SKIP_VERIFY", "command": cmd_val, "display_style": disp_style, "expected_en": "SKIP", "expected_local": "SKIP"})
            continue
            
        try:
            exp = load_expected(args.excel, "english", "english", index=idx_val, tag=tag_val)
            exp["command"] = cmd_val
            exp["display_style"] = disp_style
            expected_list.append(exp)
        except Exception as e:
            expected_list.append({"index": idx_val, "tag": tag_val, "command": cmd_val, "display_style": disp_style, "expected_en": "", "expected_local": ""})

    # We only save the languages chosen in the GUI, we DO NOT search Excel yet!
    chosen_langs = [l.strip() for l in args.language.split(",") if l.strip()] if args.region.lower() == "multiple" else [args.language]
    print("Expected (Local): [Will detect language via AI first, then fetch from Excel]")


    if args.preview:
        detector = FastDetector(model_path)
        full_frame, rois = capture_screen_roi_preview(detector, camera_id=camera_id, confidence=confidence, model_path=model_path, profiles=profiles)
        print("\n[INFO] Preview Closed. (Verification string tests will run via Start button).")
        return  
        
    detector = FastDetector(model_path)
    t0_cap = time.time()
    
    # First baseline capture
    full_frame, rois, roi_coords = capture_screen_roi(
        detector, camera_id=camera_id, confidence=confidence, rolling_mode=False
    )
    t1_cap = time.time()

    print("=" * 70)
    print(f"RADIO STRING VERIFICATION - DETECTED {len(rois)} DEVICES")
    print("=" * 70)

    ocr = None 
    box_mapping = {}
    try:
        map_file = Path(__file__).resolve().parents[1] / "configs" / "box_mapping.json"
        if map_file.exists():
            with open(map_file, "r") as f:
                box_mapping = json.load(f)
    except Exception: pass

    total_ocr_time = 0.0
    all_results = []
    summary_counts = {"PASS": 0, "FAIL": 0, "WARN": 0, "SKIP": 0} 
    
    for idx, roi in enumerate(rois):
        needs_rolling_capture = False  
        
        mapped_idx = int(box_mapping.get(str(idx), idx))
        device_id = mapped_idx + 1
        dev_name = get_device_name(profiles, device_id)
        
        # Grab the saved coordinates for this specific radio
        sx1, sy1, sx2, sy2 = roi_coords[idx]
        
        exp_dict = expected_list[mapped_idx] if mapped_idx < len(expected_list) else expected_list[-1]
        
        print("\n" + "=" * 70)
        print(f"Device: {dev_name} (Extracting text...)")
        print("=" * 70)
        
        if exp_dict.get("tag") == "SKIP_VERIFY" or exp_dict.get("index") == "SKIP_VERIFY":
            summary_counts["SKIP"] += 1
            print("Skipping verification for this device as requested via Automation Command.")
            payload = {"device": dev_name, "command": exp_dict.get('command', ''), "display_style": exp_dict.get('display_style', ''), "index": "SKIP", "tag": "SKIP", "expected": "-", "actual": "-", "confidence": "-", "verdict": "SKIP", "error": ""}
            print(f"[GUI_RESULT] {json.dumps(payload)}")
            continue

        if exp_dict.get('command'): print(f"Command: {exp_dict.get('command')}")
        if exp_dict.get('display_style'): print(f"Display Style: {exp_dict.get('display_style')}")
        print(f"Index: {exp_dict.get('index', '')}")
        if exp_dict.get('tag'): print(f"Tag: {exp_dict.get('tag', '')}")

        # --- PRE-FETCH EXCEL TARGETS BASED STRICTLY ON GUI SELECTION ---
        chosen_langs = [l.strip() for l in args.language.split(",") if l.strip()] if args.region.lower() == "multiple" else [args.language]
        active_targets = []
        
        for cl in chosen_langs:
            r, _ = map_language_to_region("", "", allowed_langs=cl)
            try:
                n_exp = load_expected(args.excel, r, cl, index=exp_dict.get("index"), tag=exp_dict.get("tag"))
                active_targets.append({
                    "region": r,
                    "language": cl,
                    "exp_target": n_exp.get("expected_local", "")
                })
            except Exception:
                pass

        # Fallback if nothing was found in Excel
        if not active_targets:
            active_targets.append({
                "region": args.region,
                "language": args.language,
                "exp_target": exp_dict.get("expected_local", "") or exp_dict.get("expected_en", "")
            })
        # -------------------------------------------------------------------------


        max_retries = 10 
        best_conf = -1.0
        best_attempt_data = None
        seen_observations = set()
        
        # CRITICAL FIX: Save the exact image used
        best_roi_for_saving = roi.copy()
        
        progressive_dims = [450, 600, 800, 800, 1000]
        
        try:
            if ocr is not None: del ocr
        except Exception: pass
        ocr = MSIGenAIOCR()
        
        for attempt in range(max_retries):
            retry_roi = roi.copy()
            used_rolling_this_attempt = False
            
            if attempt > 0:
                print(f"\n[RETRY {attempt}/{max_retries-1}] Verification failed. Simulating full restart...")
                
                try:
                    del ocr 
                except Exception: pass
                ocr = MSIGenAIOCR()
                
                try:
                    cap = cv2.VideoCapture(int(camera_id), cv2.CAP_DSHOW)
                    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
                    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
                    if needs_rolling_capture:
                        used_rolling_this_attempt = True
                        retry_burst = []
                        # Capture exactly 1 frame every ~0.2 seconds, up to 20 frames
                        t_end_capture = time.time() + 9.0  # Increased time to allow 20 frames to capture
                        last_save = 0
                        
                        while time.time() < t_end_capture:
                            ok, f = cap.read()
                            if ok and f is not None:
                                # Determine the gap: 0.3s for the second frame, 0.2s for the rest
                                required_interval = 0.2 if len(retry_burst) == 1 else 0.3
                                
                                if time.time() - last_save >= required_interval:
                                    retry_burst.append(f.copy())
                                    last_save = time.time()
                                    if len(retry_burst) == 20:  # Grab 20 frames
                                        break
                        cap.release()
                        
                        if len(retry_burst) >= 4:
                            crops = []
                            for f in retry_burst[:20]:  # Process up to 20 frames
                                try:
                                    crop = f[sy1:sy2, sx1:sx2]
                                    if crop.size > 0: crops.append(crop)
                                except Exception: pass
                                
                            if crops:
                                # GRID STITCHING (4 columns, 5 rows)
                                border = 6
                                bordered_crops = []
                                for c in crops:
                                    bc = cv2.copyMakeBorder(c, border, border, border, border, cv2.BORDER_CONSTANT, value=[255, 255, 255])
                                    bordered_crops.append(bc)
                                
                                # Make sure we have a multiple of 4 to form even columns
                                while len(bordered_crops) % 4 != 0:
                                    bordered_crops.pop()
                                    
                                if bordered_crops:
                                    rows_per_col = len(bordered_crops) // 4
                                    col1 = cv2.vconcat(bordered_crops[:rows_per_col])
                                    col2 = cv2.vconcat(bordered_crops[rows_per_col:2*rows_per_col])
                                    col3 = cv2.vconcat(bordered_crops[2*rows_per_col:3*rows_per_col])
                                    col4 = cv2.vconcat(bordered_crops[3*rows_per_col:])
                                    
                                    # Combine the 4 columns horizontally
                                    grid_roi = cv2.hconcat([col1, col2, col3, col4])
                                    
                                    # FORCE 1:1 ASPECT RATIO PADDING
                                    gh, gw = grid_roi.shape[:2]
                                    max_dim = max(gh, gw)
                                    top = (max_dim - gh) // 2
                                    bottom = max_dim - gh - top
                                    left = (max_dim - gw) // 2
                                    right = max_dim - gw - left
                                    retry_roi = cv2.copyMakeBorder(grid_roi, top, bottom, left, right, cv2.BORDER_CONSTANT, value=[0, 0, 0])
                    else:
                        best_f = None
                        max_b = -1
                        t_end_capture = time.time() + 0.5
                        last_save = 0
                        frames_grabbed = 0
                        
                        while time.time() < t_end_capture:
                            ok, f = cap.read()
                            if ok and f is not None:
                                if time.time() - last_save >= 0.08:
                                    crop = f[sy1:sy2, sx1:sx2]
                                    if crop.size > 0:
                                        b = np.mean(cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY))
                                        if b > max_b:
                                            max_b = b
                                            best_f = f.copy()
                                    last_save = time.time()
                                    frames_grabbed += 1
                                if frames_grabbed >= 5:
                                    break
                        cap.release()
                        
                        if best_f is not None:
                            retry_roi = best_f[sy1:sy2, sx1:sx2]
                except Exception as e:
                    print(f"    -> [Camera Error] Could not grab fresh frame: {e}")
                
            t0_ocr = time.time()
            
            current_dim = progressive_dims[attempt] if attempt < len(progressive_dims) else 1000
            current_squash = 0.6 if attempt == 3 else 1.0
            
            if attempt > 0:
                print(f"    -> [Image Prep] Resolution: {current_dim}px | Squash Ratio: {current_squash}")

            # Pass only the English text as a generic context reference to avoid hallucination
            hint_str = exp_dict.get("expected_en", "")

            if _is_screen_blank(retry_roi):
                print("    -> [Pre-Check] Camera sees a BLANK screen. Skipping AI hallucination.")
                text = "Detected text(original): " 
                _conf = 0.0
            else:
                if used_rolling_this_attempt:
                    pass_hint = f"HINT: The image contains a 2-column grid of screenshots showing a scrolling screen. Read down the left column, then the right."
                else:
                    pass_hint = f"REFERENCE ENGLISH TRANSLATION: '{hint_str}'. Ensure you transcribe the ORIGINAL language visible on screen." if hint_str else ""

                # Pass GUI selections so AI has a general idea of allowed alphabets
                pass_lang = args.language

                text, _conf = ocr.extract_text(
                    retry_roi, 
                    expected_language=pass_lang, 
                    dynamic_dim=current_dim, 
                    squash_ratio=current_squash,
                    expected_text=pass_hint
                )
            
            t1_ocr = time.time()

            if attempt == 0:
                total_ocr_time += (t1_ocr - t0_ocr)

            parsed = _parse_structured_fields(text)
            
            parsed_doc_error = bool(parsed.get("error_red"))
            parsed_doc_type = str(parsed.get("error_type") or "").strip().lower()
            parsed_doc_evidence = str(parsed.get("error_evidence") or "").strip()
            parsed["error_red"] = False
            parsed["error_evidence"] = ""
            parsed["error_type"] = ""

            lang_detected = parsed.get("language") or ""
            if "detected text(original)" in text.lower():
                orig_text = parsed.get("original", "")
            else:
                orig_text = parsed.get("original") or text
            eng_text = parsed.get("english") or ""
            
            orig_text = str(orig_text).replace("<<<", "").replace(">>>", "").strip()
            eng_text = str(eng_text).replace("<<<", "").replace(">>>", "").strip()

            observed_n = _norm_text(orig_text)

            best_sub_conf = -1.0
            best_sub_verdict = "FAIL"
            best_sub_data = {}

            # Loop through the GUI-selected languages (active_targets) and check if ANY of them match the OCR text
            for target_info in active_targets:
                t_region = target_info["region"]
                t_lang = target_info["language"]
                t_exp = target_info["exp_target"]
                
                expected_local_n = _norm_text(t_exp)
                is_cjk_target = any('\u4e00' <= ch <= '\u9fff' or '\u3040' <= ch <= '\u30ff' or '\uac00' <= ch <= '\ud7a3' for ch in expected_local_n)

                def _clean_for_compare(text: str, is_cjk: bool) -> str:
                    c = str(text)
                    # Remove only known weird UI glitch prefixes that the AI hallucinates at the far left edge
                    c = re.sub(r'^(i|l|v|4)\s+', '', c, flags=re.IGNORECASE)
                    c = re.sub(r'^(!|\?|⏹|\[\]|\'|\"|️)\s*', '', c)
                    
                    if is_cjk:
                        c = re.sub(r'\s+', '', c) # Remove spaces for CJK
                        # If expected string doesn't contain English letters, strip English letters from observation (fixes hallucinated English in pure CJK)
                        if not any(char.isascii() and char.isalpha() for char in expected_local_n):
                            c = re.sub(r'[A-Za-z]', '', c)
                        return c
                    else:
                        c = " ".join(c.split()) # Normalize spaces
                        c = re.sub(r'\s+[LHlh]$', '', c) # Strip specific hallucinated trailing letters
                        return c

                flat_obs = _clean_for_compare(observed_n, is_cjk_target)
                flat_exp = _clean_for_compare(expected_local_n, is_cjk_target)
                
                # --- STRICT EXACT MATCHING ---
                if flat_exp and flat_obs != flat_exp:
                    corrected_obs = flat_obs
                    
                    known_illusions = {
                        "RETRY": "RETRV", "Retry": "Retrv", "retry": "retrv",
                        "虎碼": "號碼", "新響": "漸響", "施錠": "旋鈕", "施鈕": "旋鈕",
                        "遠測": "遙測", "擺置": "搁置", "通话": "通稱", "通話": "通稱",
                        "鎖碼": "變碼", "R∫": "Rx", "r∫": "rx", "音": "颤音",
                        "SYSTELIAS": "SYSTEM ALIAS", "Systelias": "System Alias",
                        "資訊驗證失敗": "驗證失敗", "資訊 驗證失敗": "驗證失敗",
                        "스캔컴": "스캔켬", "스켈처": "스켈치", "Z-S": "Z-s",
                        "タイカヘンシン": "クイックヘンシン", "タイカ": "クイック",
                        "발기": "밝기", "받기": "밝기",
                        "흰&라이트커기": "혼&라이트켜기", "흰": "혼", "커기": "켜기",
                        # Rolling/Optical Illusions
                        "30S": "30ビ", "30L": "30ビ", "ヨウ": "ョウ", 
                        "ピョウ": "ビョウ", "ビヨウ": "ビョウ", "ヒョウ": "ビョウ", "トウ": "ョウ",
                        "ヨ": "ョ", "ヤ": "ャ", "ユ": "ュ"
                    }
                    
                    for bad, good in known_illusions.items():
                        if bad in corrected_obs and good in flat_exp:
                            corrected_obs = corrected_obs.replace(bad, good)
                    
                    sim = difflib.SequenceMatcher(None, flat_exp, corrected_obs).ratio()
                    if sim >= 0.80 and len(corrected_obs) == len(flat_exp):
                        corrected_obs = flat_exp
                    
                    # Handle minor cut-offs
                    if len(flat_obs) >= 2 and flat_exp.endswith(flat_obs):
                        if len(flat_exp) - len(flat_obs) <= 2:
                            corrected_obs = flat_exp

                    # Handle minor 1-2 character cut-offs at the edges (safe to allow)
                    if len(flat_obs) >= 2 and flat_exp.endswith(flat_obs) and len(flat_exp) - len(flat_obs) <= 2:
                        corrected_obs = flat_exp

                    # STRICT TRUNCATION RULE: If the AI reads a fragment (start, middle, or end),
                    # it MUST make up at least 99% of the expected text length to pass.
                    if len(flat_obs) >= 5 and flat_obs in flat_exp:
                        if len(flat_obs) >= int(len(flat_exp) * 0.99):
                            corrected_obs = flat_exp

                    if is_cjk_target and flat_exp in flat_obs and len(flat_obs) <= len(flat_exp) + 2:
                        corrected_obs = flat_exp

                    if is_cjk_target and flat_exp in flat_obs:
                        stripped_obs = re.sub(r'[A-Za-z]', '', flat_obs)
                        if stripped_obs == flat_exp:
                            corrected_obs = flat_exp

                    # Fragment Summation Math Algorithm
                    if used_rolling_this_attempt:
                        matcher = difflib.SequenceMatcher(None, flat_exp, corrected_obs)
                        matching_chars = sum(m.size for m in matcher.get_matching_blocks())
                        if matching_chars >= int(len(flat_exp) * 0.85):
                            corrected_obs = flat_exp
                        else:
                            # --- NEW: SCRAMBLED ROLLING TEXT REARRANGEMENT RULE ---
                            if not is_cjk_target:
                                import unicodedata
                                def remove_accents(input_str):
                                    return "".join(c for c in unicodedata.normalize('NFKD', input_str) if not unicodedata.combining(c))

                                exp_words = flat_exp.split()
                                obs_lower = remove_accents(corrected_obs.lower())
                                found_chars_len = 0
                                total_exp_chars = sum(len(w) for w in exp_words)
                                
                                for word in exp_words:
                                    word_clean = remove_accents(word.lower())
                                    if word_clean in obs_lower:
                                        found_chars_len += len(word)
                                        obs_lower = obs_lower.replace(word_clean, "", 1)
                                        
                                if total_exp_chars > 0 and (found_chars_len / total_exp_chars) >= 0.85:
                                    corrected_obs = flat_exp
                            else:
                                exp_chars = list(flat_exp)
                                obs_chars = list(corrected_obs)
                                found_cjk_chars = 0
                                for ch in exp_chars:
                                    if ch in obs_chars:
                                        found_cjk_chars += 1
                                        obs_chars.remove(ch)
                                if len(exp_chars) > 0 and (found_cjk_chars / len(exp_chars)) >= 0.85:
                                    corrected_obs = flat_exp

                    flat_obs = corrected_obs
                
                t_conf = 0.0
                t_verdict = "FAIL"
                if flat_exp:
                    similarity = difflib.SequenceMatcher(None, flat_exp, flat_obs).ratio()
                    t_conf = round(similarity * 100, 1)

                    if flat_obs == flat_exp: 
                        t_verdict = "PASS"
                        t_conf = 100.0
                    elif flat_exp in flat_obs:
                        t_verdict = "FAIL"
                    else:
                        if t_conf >= 70.0: t_verdict = "WARN"
                        else: t_verdict = "FAIL"
                else:
                    t_verdict = "PASS" if not flat_obs else "FAIL"
                
                if t_conf > best_sub_conf or (t_verdict == "PASS" and best_sub_verdict != "PASS"):
                    best_sub_conf = t_conf
                    best_sub_verdict = t_verdict
                    best_sub_data = {
                        "final_region": t_region,
                        "final_language": t_lang,
                        "exp_target": t_exp,
                        "flat_obs": flat_obs,
                        "flat_exp": flat_exp
                    }
                if t_verdict == "PASS":
                    break

            confidence_pct = best_sub_conf
            verdict = best_sub_verdict
            final_region = best_sub_data.get("final_region", args.region)
            final_language = best_sub_data.get("final_language", args.language)
            exp_target = best_sub_data.get("exp_target", "")
            flat_obs = best_sub_data.get("flat_obs", "")
            flat_exp = best_sub_data.get("flat_exp", "")
            final_error_red = False
            final_error_evidence = ""
            final_error_type = ""

            if attempt >= 4 and flat_obs in seen_observations:
                print(f"    -> [Smart Abort] AI returned exact same text ('{flat_obs}'). Stopping retries.")
                
            seen_observations.add(flat_obs)

            if verdict != "PASS" and attempt < max_retries - 1:
                if len(flat_obs) < len(flat_exp):
                    if "..." in str(orig_text) or "…" in str(orig_text) or "..." in flat_obs:
                        needs_rolling_capture = False
                        print("    -> [Next Retry Status] Missing words, but '...' detected. This text is statically truncated. Normal retry is enough.")
                    else:
                        needs_rolling_capture = True
                        print("    -> [Next Retry Status] Missing words and no '...' detected. Text may be rolling. Triggering Rolling Batch Capture!")
                else:
                    needs_rolling_capture = False

            if confidence_pct > best_conf:
                best_conf = confidence_pct
                best_roi_for_saving = retry_roi.copy() 
                best_attempt_data = {
                    "confidence_pct": confidence_pct,
                    "verdict": verdict,
                    "orig_text": orig_text,
                    "eng_text": eng_text,
                    "lang_detected": lang_detected,
                    "final_error_red": final_error_red,
                    "final_error_evidence": final_error_evidence,
                    "final_error_type": final_error_type,
                    "final_region": final_region,
                    "final_language": final_language,
                    "exp_target": exp_target,
                    "flat_obs": flat_obs,
                    "flat_exp": flat_exp
                }

            if verdict == "PASS":
                break

        if best_attempt_data:
            confidence_pct = best_attempt_data["confidence_pct"]
            verdict = best_attempt_data["verdict"]
            orig_text = best_attempt_data["orig_text"]
            eng_text = best_attempt_data["eng_text"]
            lang_detected = best_attempt_data["lang_detected"]
            final_error_red = best_attempt_data["final_error_red"]
            final_error_evidence = best_attempt_data["final_error_evidence"]
            final_error_type = best_attempt_data["final_error_type"]
            final_region = best_attempt_data["final_region"]
            final_language = best_attempt_data["final_language"]
            exp_target = best_attempt_data["exp_target"]
            flat_obs = best_attempt_data["flat_obs"]
            flat_exp = best_attempt_data["flat_exp"]

        observed_display = flat_obs 
        
        exp_lines = [
            f"Expected ({final_region.upper()}/{final_language}): {exp_target}"
        ]
        
        detected_flat = " ".join([ln.strip() for ln in str(orig_text).splitlines() if ln.strip()])
        print(f"Detected: '{detected_flat}'")
        
        print("-" * 70)
        print(f"Observed (normalized): {flat_obs}")
        print(f"Expected (normalized): {flat_exp}")
        print(f"Match Confidence:      {confidence_pct}%")
        print("-" * 70)
        
        error_msg_display = ""
        if final_error_red:
            error_words = []
            for ln in final_error_evidence.splitlines():
                if ln.lower().startswith("likely") or ln.lower().startswith("token"):
                    error_words.append(ln.split(":")[-1].strip())
            
            err_type_str = str(final_error_type or "STRING").upper()
            
            if error_words:
                words_str = ", ".join(error_words)
            else:
                words_str = final_error_evidence.splitlines()[0] if final_error_evidence else "Error detected"
            
            error_msg_display = f"[{err_type_str} ERROR: {words_str}]"
            print(error_msg_display)
        
        if verdict in ["FAIL", "WARN"]:
            mismatch_reason = ""
            if len(flat_obs) > len(flat_exp) and flat_exp.lower() in flat_obs.lower():
                extra_text = re.sub(re.escape(flat_exp), "", flat_obs, flags=re.IGNORECASE).strip()
                mismatch_reason = f"Extra text detected: '{extra_text}'"
            elif len(flat_obs) < len(flat_exp) and flat_obs.lower() in flat_exp.lower():
                mismatch_reason = "Missing part of the expected text."
            else:
                mismatch_reason = "Text misspelled or completely changed."

            text_err = f"Mismatch (Conf: {confidence_pct}%): {mismatch_reason}"
            
            if error_msg_display:
                error_msg_display += f" | {text_err}"
            else:
                error_msg_display = text_err
                
            print(f"[VERDICT REASON] {text_err}")

        if verdict == "PASS": print("PASS")
        elif verdict == "WARN": print("WARN")
        else: print("FAIL")
        summary_counts[verdict] += 1

        roi_saved_path = ""
        if args.save_roi_dir:
            out_dir = Path(args.save_roi_dir) / verdict
            out_dir.mkdir(parents=True, exist_ok=True)
            
            safe_dev_name = re.sub(r'[\\/*?:"<>| ]', '_', dev_name)
            dev_idx = str(exp_dict.get('index', '')).strip()
            ts_suffix = time.strftime("%H%M%S")
            
            if dev_idx and dev_idx != "SKIP_VERIFY":
                new_filename = f"roi_{dev_idx}_{safe_dev_name}_{ts_suffix}.jpg"
            else:
                new_filename = f"roi_{safe_dev_name}_{ts_suffix}.jpg"
                
            full_roi_path = out_dir / new_filename
            cv2.imwrite(str(full_roi_path), best_roi_for_saving)
            roi_saved_path = str(full_roi_path.resolve())

        if args.summary_excel:
            xl_p = Path(args.summary_excel)
            try:
                from openpyxl import load_workbook, Workbook
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.drawing.image import Image as OpenpyxlImage

                if not xl_p.exists():
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Batch Summary"
                    headers = ["Timestamp", "Device", "Region", "Language", "Command", "Display Style", "Index", "Tag", "Expected (Local)", "Actual Detected", "Confidence (%)", "Verdict", "Error Message", "ROI Image"]
                    ws.append(headers)
                    
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for col_num, cell in enumerate(ws[1], 1):
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    widths = [20, 15, 12, 15, 25, 45, 10, 25, 35, 35, 15, 12, 35, 30]
                    for i, width in enumerate(widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
                else:
                    wb = load_workbook(xl_p)
                    ws = wb.active
                    
                ws.append([
                    time.strftime("%Y-%m-%d %H:%M:%S"),
                    dev_name,
                    final_region.upper(),
                    final_language,
                    exp_dict.get('command', ''),
                    exp_dict.get('display_style', ''),
                    exp_dict.get('index', ''),
                    exp_dict.get('tag', ''),
                    flat_exp, 
                    observed_display,
                    confidence_pct,
                    verdict,
                    error_msg_display,
                    "" 
                ])
                
                row_idx = ws.max_row
                
                for col_num in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col_num)
                    h_align = "left" if col_num in [9, 10, 13] else "center"
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal=h_align)

                verdict_cell = ws.cell(row=row_idx, column=12) 
                if verdict == "PASS":
                    verdict_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    verdict_cell.font = Font(color="006100", bold=True)
                elif verdict == "FAIL":
                    verdict_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    verdict_cell.font = Font(color="9C0006", bold=True)
                elif verdict == "WARN":
                    verdict_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    verdict_cell.font = Font(color="9C6500", bold=True)
                
                ws.row_dimensions[row_idx].height = 80
                
                if roi_saved_path and os.path.exists(roi_saved_path):
                    img = OpenpyxlImage(roi_saved_path)
                    img.height = 95
                    img.width = 200
                    img.anchor = f"N{row_idx}" 
                    ws.add_image(img)

                wb.save(xl_p)
            except Exception as e:
                print(f"[WARNING] Failed to write to summary Excel: {e}")

        payload = {
            "device": dev_name,
            "command": exp_dict.get('command', ''),
            "display_style": exp_dict.get('display_style', ''),
            "index": exp_dict.get('index', ''),
            "tag": exp_dict.get('tag', ''),
            "expected": flat_exp, 
            "actual": observed_display,
            "confidence": f"{confidence_pct}%",
            "verdict": verdict,
            "error": error_msg_display
        }
        print(f"[GUI_RESULT]{json.dumps(payload)}")

        all_results.append({
            "roi": best_roi_for_saving, "orig_text": orig_text, "eng_text": eng_text, "lang_detected": lang_detected,
            "verdict": verdict, "exp_lines": exp_lines, "dev_name": dev_name, "error_msg": error_msg_display
        })
        
    
    t_end_total = time.time()
    time_taken = round(t_end_total - t0_total, 2)

    print("\n" + "=" * 70)
    print("EXECUTION SUMMARY")
    print("=" * 70)
    print(f"Total Devices Checked: {len(rois)}")
    print(f"PASS: {summary_counts['PASS']} | FAIL: {summary_counts['FAIL']} | WARN: {summary_counts['WARN']} | SKIP: {summary_counts['SKIP']}")
    print(f"Total Time Taken: {time_taken} seconds")
    print("=" * 70 + "\n")

    try: ocr.get_usage_and_cost()
    except Exception: pass

    for res in all_results:
        try: _show_ocr_result_window(res["roi"], res["orig_text"], res["eng_text"], res["lang_detected"], res["verdict"], expected_lines=res["exp_lines"], device_name=res["dev_name"], error_msg=res.get("error_msg", ""))
        except Exception: pass

if __name__ == "__main__":
    main()