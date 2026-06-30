import os
import queue
import subprocess
import sys
import threading
import tkinter as tk
import time
import tempfile
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk

import cv2
import json
import re
import yaml
import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

import sys
sys.coinit_flags = 2  # Forces COM initialization to STA mode to prevent Tkinter crashes

import runpy
from contextlib import redirect_stdout, redirect_stderr
import verify_string
import init_genai_session

import os
if getattr(sys, 'frozen', False):
    os.chdir(os.path.dirname(sys.executable))
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import main_msi_genai

class _QueueStream:
    def __init__(self, queue):
        self.queue = queue
    def write(self, text):
        if text:
            self.queue.put(text)
    def flush(self):
        pass

try:
    from pywinauto import Desktop
    from pywinauto.application import Application
    from pywinauto.keyboard import send_keys  
    HAS_PYWINAUTO = True
except ImportError:
    Desktop = None
    Application = None
    send_keys = None
    HAS_PYWINAUTO = False

_EVT_FINISHED = "__PROCESS_FINISHED__"
_EVT_COMMG_DONE = "__COMMG_DONE__"


def _settings_path() -> Path:
    base = Path(os.getenv("APPDATA") or Path.home())
    return base / "walkie_tracker_verify_string_gui.json"

def _load_settings() -> dict:
    p = _settings_path()
    try:
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}
    return {}

def _save_settings(data: dict) -> None:
    p = _settings_path()
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

def _probe_camera_ids(max_id: int = 5) -> list[tuple[str, str]]:
    cam_names = []
    if os.name == 'nt':
        try:
            cmd = 'powershell -ExecutionPolicy Bypass -Command "Get-PnpDevice -PresentOnly | Where-Object { $_.Class -eq \'Camera\' -or $_.Class -eq \'Image\' } | Select-Object -ExpandProperty FriendlyName"'
            res = subprocess.check_output(cmd, shell=True, text=True, creationflags=0x08000000)
            cam_names = [line.strip() for line in res.strip().split('\n') if line.strip()]
        except Exception:
            pass

    found = []
    for i in range(int(max_id) + 1):
        cap = None
        try:
            backend = getattr(cv2, "CAP_DSHOW", 0)
            cap = cv2.VideoCapture(i, backend)
            if not cap.isOpened():
                try: cap.release()
                except Exception: pass
                cap = cv2.VideoCapture(i)
                
            if not cap.isOpened():
                continue
                
            ok, frame = cap.read()
            if not ok or frame is None or getattr(frame, "size", 0) == 0:
                continue
            
            name = cam_names[i] if i < len(cam_names) else f"Camera {i}"
            found.append((str(i), name))
            
        except Exception:
            continue
        finally:
            try:
                if cap is not None: cap.release()
            except Exception: pass
    return found

def _norm_col(s: str) -> str:
    v = str(s or "").strip().lower().replace("_", " ")
    v = re.sub(r"[\(\)\[\]\{\}:,;/\\\-]+", " ", v)
    v = " ".join(v.split())
    if v.startswith("string "):
        v = v[len("string ") :].strip()
    if v.startswith("str "):
        v = v[len("str ") :].strip()
    return v

def _sheet_name_for_region(xls_path: str, region: str) -> str:
    wb = load_workbook(filename=xls_path, read_only=True, data_only=True)
    try:
        want = _norm_col(region)
        if want in ["en", "english", "global"]:
            want = "english"
        if want in ["latam", "lac", "lacr", "la cr"]:
            want = "lacr"

        for s in wb.sheetnames:
            if _norm_col(s) == want:
                return s
        for s in wb.sheetnames:
            if want in _norm_col(s):
                return s
        raise ValueError(f"Sheet for region '{region}' not found")
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _tag_options_from_excel(excel_path: str) -> list[str]:
    if not excel_path:
        return []
    p = Path(excel_path)
    if not p.exists():
        return []

    try:
        english_sheet = _sheet_name_for_region(str(p), "english")
    except Exception:
        english_sheet = "English"

    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        if english_sheet not in wb.sheetnames:
            return []
        ws = wb[english_sheet]

        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break

        tag_col_idx = None
        for i, h in enumerate(headers):
            n = _norm_col(h)
            if n in ["string tag", "stringtag", "tag"]:
                tag_col_idx = i
                break
        if tag_col_idx is None:
            return []

        tags = []
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if tag_col_idx >= len(row):
                continue
            v = row[tag_col_idx]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            try:
                if _norm_col(s) in ["string tag", "stringtag", "tag"]:
                    continue
            except Exception:
                pass
            if s not in seen:
                seen.add(s)
                tags.append(s)
        return tags
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _build_index_to_tag_map(excel_path: str) -> dict:
    if not excel_path:
        return {}
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        
        target = "english"
        sheet_name = None
        for s in xls.sheet_names:
            if _norm_col(s) in ["english", "en", "global"]:
                sheet_name = s
                break
            if target in _norm_col(s):
                sheet_name = s
                break
                
        if not sheet_name: 
            return {}
            
        df = pd.read_excel(xls, sheet_name=sheet_name, engine="openpyxl")
        
        idx_col = next((c for c in df.columns if _norm_col(c) == "index"), None)
        
        preferred, fallback = [], []
        for c in df.columns:
            low, n = str(c or "").strip().lower(), _norm_col(c)
            if ("tag" in low and "string" in low) or n in ["string tag", "stringtag"]: preferred.append(c)
            elif n == "tag" or "tag" in low: fallback.append(c)
        tag_col = preferred[0] if preferred else (fallback[0] if fallback else "")
        
        if not idx_col or not tag_col: 
            return {}
            
        mapping = {}
        for _, row in df.iterrows():
            idx_val = row[idx_col]
            tag_val = row[tag_col]
            
            if pd.isna(idx_val) or pd.isna(tag_val): 
                continue
                
            iv = str(idx_val).strip()
            tv = str(tag_val).strip()
            
            if tv.lower() == 'nan' or not tv: 
                continue
                
            if iv.endswith(".0"): 
                try: iv = str(int(float(iv)))
                except Exception: pass
            if iv.isdigit(): 
                try: iv = str(int(iv))
                except Exception: pass
                
            if iv not in mapping:
                mapping[iv] = tv
            
        return mapping
    except Exception:
        return {}

def _language_options_from_excel(excel_path: str, region: str) -> list[str]:
    if not excel_path:
        return []
    p = Path(excel_path)
    if not p.exists():
        return []

    try:
        sheet = _sheet_name_for_region(str(p), region)
    except Exception:
        return []
        
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break

        ignore = {
            "index", "string tag", "tag", "string category", "category",
            "version", "ver", "comment", "comments",
            "notes", "note", "description", "desc",
        }

        opts = []
        seen = set()
        for h in headers:
            n = _norm_col(h)
            if not n or n in ignore:
                continue
            label = str(h).strip() if h is not None else ""
            if n in [
                "japanese", "korean", "simplified chinese", "traditional chinese",
                "french", "spanish", "german", "italian", "polish",
                "russian", "turkish", "arabic", "hungarian", "hebrew",
                "czech", "portuguese", "english",
            ]:
                label = n.title()
            if label and label not in seen:
                seen.add(label)
                opts.append(label)

        return opts
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _default_excel_path() -> str:
    p = os.getenv("VERIFY_EXCEL", "").strip()
    if p:
        return p
    return ""

def _python_exe() -> str:
    return sys.executable or "python"

def _default_model_path() -> str:
    try:
        cfg_path = Path(__file__).resolve().parents[1] / "configs" / "settings.yaml"
        if not cfg_path.exists():
            return ""
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
        p = (((cfg.get("detector") or {}).get("path")) or "").strip()
        if not p:
            return ""
        pp = Path(p)
        if not pp.is_absolute():
            pp = (Path(__file__).resolve().parents[1] / pp).resolve()
        return str(pp)
    except Exception:
        return ""

def _resolve_path(p: str) -> str:
    v = (p or "").strip()
    if not v:
        return ""
    try:
        pp = Path(v)
        if not pp.is_absolute():
            pp = (Path(__file__).resolve().parents[1] / pp).resolve()
        return str(pp)
    except Exception:
        return v


class VerifyStringGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Verify String (Walkie-Tracker + Automation Integration)")
        self.root.minsize(1050, 900)

        self._auto_start_verify = False
        self._last_run_is_verification = True
        self.proc = None
        self.q = queue.Queue()
        self.last_result = ""
        self.last_expected = ""
        self.last_actual = ""       
        self.last_error_msg = ""    
        self._pending_expected_norm = False
        self._pending_actual_norm = False 
        self._batch_log_dir = None  

        self._settings = _load_settings()
        self.available_languages = ["English"]

        self.COMMG_PATH = r"C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Motorola\CommG_LTD\CommG_LTD.lnk"
        self.WINDOW_SEARCH_TERM = "CommuniGATOR"
        self.commg_handles = []
        self.type_lock = threading.Lock()
        
        self._commg_pending_queue = []
        self._commg_is_active_run = False
        self._is_paused = False 
        self.active_cmd_windows = []

        frm = tk.Frame(root, padx=10, pady=10)
        frm.pack(fill=tk.BOTH, expand=True)

        row = 0

        self.excel_var = tk.StringVar(value=str(self._settings.get("excel") or _default_excel_path()))
        self.tag_var = tk.StringVar(value=str(self._settings.get("tag") or ""))
        self.index_var = tk.StringVar(value=str(self._settings.get("index") or ""))
        
        self.index_var.trace_add("write", self._on_index_changed)
        
        tk.Label(frm, text="Excel (.xlsm/.xlsx)").grid(row=row, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.excel_var, width=60).grid(row=row, column=1, sticky="we", padx=(8, 8))
        tk.Button(frm, text="Browse...", command=self.browse_excel).grid(row=row, column=2, sticky="e")
        row += 1

        tk.Label(frm, text="String Tag").grid(row=row, column=0, sticky="w", pady=(6, 0))
        self.tag_combo = ttk.Combobox(frm, textvariable=self.tag_var, width=50, state="normal")
        self.tag_combo.grid(row=row, column=1, sticky="w", padx=(8, 0), pady=(6, 0))
        self.tag_combo.bind("<KeyRelease>", self._on_tag_typed)
        self.tag_combo.bind("<Escape>", self._on_tag_escape)
        self.tag_combo.bind("<FocusOut>", self._on_tag_focus_out)
        self.tag_combo.bind("<Down>", self._on_tag_down)
        row += 1

        tk.Label(frm, text="Index (optional)").grid(row=row, column=0, sticky="w", pady=(6, 0))
        tk.Entry(frm, textvariable=self.index_var, width=20).grid(row=row, column=1, sticky="w", padx=(8, 0), pady=(6, 0))
        row += 1

        self.camera_id_var = tk.StringVar(value=str(self._settings.get("camera_id") or "1"))
        self.save_log_var = tk.BooleanVar(value=bool(self._settings.get("save_log", False)))
        self.log_path_var = tk.StringVar(value=str(self._settings.get("log_path") or ""))
        
        self.enable_rolling_var = tk.BooleanVar(value=bool(self._settings.get("enable_rolling", False)))
        self.enable_truncation_var = tk.BooleanVar(value=bool(self._settings.get("enable_truncation", False)))
        self.enable_retries_var = tk.BooleanVar(value=bool(self._settings.get("enable_retries", False)))
        self.retry_count_var = tk.StringVar(value=str(self._settings.get("retry_count", "2")))
        self._log_fp = None
        self._log_session_dir = None

        extras = tk.Frame(frm)
        extras.grid(row=row, column=0, columnspan=3, sticky="we", pady=(8, 0))

        tk.Label(extras, text="Camera ID").pack(side=tk.LEFT, padx=(0, 2))
        self.camera_combo = ttk.Combobox(extras, textvariable=self.camera_id_var, width=35, state="readonly")
        self.camera_combo.pack(side=tk.LEFT)
        tk.Button(extras, text="Refresh", command=self.refresh_cameras).pack(side=tk.LEFT, padx=(6, 0))
        
        tk.Checkbutton(extras, text="Rolling Text", variable=self.enable_rolling_var).pack(side=tk.LEFT, padx=(12, 0))
        tk.Checkbutton(extras, text="Truncation (...)", variable=self.enable_truncation_var).pack(side=tk.LEFT, padx=(6, 0))
        tk.Checkbutton(extras, text="Custom Retries", variable=self.enable_retries_var).pack(side=tk.LEFT, padx=(12, 0))
        tk.Entry(extras, textvariable=self.retry_count_var, width=4).pack(side=tk.LEFT, padx=(2, 0))
        tk.Checkbutton(extras, text="Save log", variable=self.save_log_var).pack(side=tk.LEFT, padx=(12, 0))
        tk.Entry(extras, textvariable=self.log_path_var, width=28).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(extras, text="Browse...", command=self.browse_log).pack(side=tk.LEFT, padx=(6, 0))
        self.drive_folder_var = tk.StringVar(value=str(self._settings.get("drive_folder", "Walkie_Logs")))
        
        tk.Label(extras, text="Drive Folder Name:").pack(side=tk.LEFT, padx=(12, 0))
        tk.Entry(extras, textvariable=self.drive_folder_var, width=15).pack(side=tk.LEFT, padx=(6, 0))
        
        self.excel_filename_var = tk.StringVar(value=str(self._settings.get("excel_filename", "Batch_Summary_Report.xlsx")))
        tk.Label(extras, text="Excel Name:").pack(side=tk.LEFT, padx=(12, 0))
        tk.Entry(extras, textvariable=self.excel_filename_var, width=20).pack(side=tk.LEFT, padx=(6, 0))
        
        row += 1

        saved_model = str(self._settings.get("model_path") or "")
        if not saved_model.strip():
            saved_model = _default_model_path()
        self.model_path_var = tk.StringVar(value=_resolve_path(saved_model))
        
        tk.Label(frm, text="Model Path").grid(row=row, column=0, sticky="w", pady=(6, 0))
        tk.Entry(frm, textvariable=self.model_path_var, width=60).grid(row=row, column=1, sticky="we", padx=(8, 8), pady=(6, 0))
        tk.Button(frm, text="Browse...", command=self.browse_model).grid(row=row, column=2, sticky="e", pady=(6, 0))
        row += 1

        self.lf_commg = tk.LabelFrame(frm, text=" Automation Integration (CommG / CMD / PuTTY) ", padx=10, pady=5, fg="#00008B", font=('Arial', 10, 'bold'))
        self.lf_commg.grid(row=row, column=0, columnspan=3, sticky="we", pady=(10, 0))
        
        top_frame = tk.Frame(self.lf_commg)
        top_frame.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 5))

        old_enable = self._settings.get("integration_enable")
        if old_enable is None:
            old_enable = self._settings.get("commg_enable", False)
        self.integration_enable_var = tk.BooleanVar(value=bool(old_enable))
        tk.Checkbutton(top_frame, text="Enable Automation Sequence", variable=self.integration_enable_var, font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=(0, 15))

        self.integration_type_var = tk.StringVar(value=str(self._settings.get("integration_type", "CommG")))
        tk.Label(top_frame, text="Tool:").pack(side=tk.LEFT)
        tk.Radiobutton(top_frame, text="CommG", variable=self.integration_type_var, value="CommG").pack(side=tk.LEFT, padx=(5, 5))
        tk.Radiobutton(top_frame, text="CMD (Telnet)", variable=self.integration_type_var, value="CMD").pack(side=tk.LEFT, padx=(5, 5))
        tk.Radiobutton(top_frame, text="PuTTY", variable=self.integration_type_var, value="PuTTY").pack(side=tk.LEFT, padx=(5, 5))

        self.telnet_port_var = tk.StringVar(value=str(self._settings.get("telnet_port", "23")))
        tk.Label(top_frame, text="Port:").pack(side=tk.LEFT, padx=(10, 2))
        tk.Entry(top_frame, textvariable=self.telnet_port_var, width=5).pack(side=tk.LEFT)

        ip_frame = tk.Frame(self.lf_commg)
        ip_frame.grid(row=1, column=0, sticky="nw", padx=(0, 20))
        tk.Label(ip_frame, text="Target IPs").pack(anchor="w")
        self.ip_listbox = tk.Listbox(ip_frame, height=3, width=20)
        self.ip_listbox.pack(side=tk.LEFT, fill="y")
        
        saved_ips = self._settings.get("commg_ips", ["192.168.10.1", "192.168.10.2"])
        for ip in saved_ips:
            self.ip_listbox.insert(tk.END, ip)

        ip_btns = tk.Frame(ip_frame)
        ip_btns.pack(side=tk.LEFT, padx=(5, 0), fill="y")
        self.new_ip_entry = tk.Entry(ip_btns, width=15)
        self.new_ip_entry.pack(pady=(0, 2))
        tk.Button(ip_btns, text="Add IP", command=self._commg_add_ip).pack(fill="x", pady=1)
        tk.Button(ip_btns, text="Remove", command=self._commg_remove_ip).pack(fill="x", pady=1)

        cmd_frame = tk.Frame(self.lf_commg)
        cmd_frame.grid(row=1, column=1, sticky="nw")
        
        self.commg_mode_var = tk.StringVar(value=str(self._settings.get("commg_mode", "Batch")))
        
        tk.Radiobutton(cmd_frame, text="Single Command", variable=self.commg_mode_var, value="Single").grid(row=0, column=0, sticky="w")
        self.commg_custom_cmd_var = tk.StringVar(value=str(self._settings.get("commg_custom_cmd", "03001101")))
        tk.Entry(cmd_frame, textvariable=self.commg_custom_cmd_var, width=35).grid(row=0, column=1, padx=(5, 0))

        tk.Radiobutton(cmd_frame, text="Batch (CSV/Excel)", variable=self.commg_mode_var, value="Batch").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.commg_batch_file_var = tk.StringVar(value=str(self._settings.get("commg_batch_file", "")))
        tk.Entry(cmd_frame, textvariable=self.commg_batch_file_var, width=35).grid(row=1, column=1, padx=(5, 0), pady=(8, 0))
        tk.Button(cmd_frame, text="Browse...", command=self._commg_browse_batch).grid(row=1, column=2, padx=(5, 0), pady=(8, 0))
        
        if not HAS_PYWINAUTO:
            tk.Label(self.lf_commg, text="⚠️ pywinauto not installed. Automation unavailable.", fg="red").grid(row=2, column=0, columnspan=3, sticky="w")
            self.integration_enable_var.set(False)
            for child in self.lf_commg.winfo_children():
                try: child.configure(state='disabled')
                except Exception: pass
        row += 1

        self.lf_devices = tk.LabelFrame(frm, text=" Devices / Language Iterations ", padx=10, pady=5, fg="#00008B", font=('Arial', 10, 'bold'))
        self.lf_devices.grid(row=row, column=0, columnspan=3, sticky="we", pady=(10, 0))
        
        self.devices_grid_frame = tk.Frame(self.lf_devices)
        self.devices_grid_frame.pack(fill=tk.BOTH, expand=True)

        self.device_rows = []
        
        # --- FIX: Define the default column count before triggering the grid refresh ---
        self.num_lang_columns = 1
        # -------------------------------------------------------------------------------

        self.refresh_languages()
        self._sync_devices_to_ips(initial_load=True)
        row += 1

        btns = tk.Frame(frm)
        btns.grid(row=row, column=0, columnspan=3, sticky="we", pady=(10, 0))
        self.btn_cam_test = tk.Button(btns, text="Camera Test", command=self.run_camera_test, bg="#ADD8E6", font=('Arial', 10, 'bold'))
        self.btn_cam_test.pack(side=tk.LEFT, padx=(0, 15))
        
        self.btn_run = tk.Button(btns, text="Start", command=self.init_and_run, bg="#90EE90", font=('Arial', 10, 'bold'))
        self.btn_run.pack(side=tk.LEFT)
        
        self.btn_stop = tk.Button(btns, text="Stop", command=self.stop, bg="#FFCCCB", font=('Arial', 10, 'bold'))
        self.btn_stop.pack(side=tk.LEFT, padx=(8, 0))
        self.btn_close = tk.Button(btns, text="Close", command=self.close)
        self.btn_close.pack(side=tk.LEFT, padx=(8, 0))
        row += 1

        self.status_var = tk.StringVar(value="Idle")
        self.status_label = tk.Label(frm, textvariable=self.status_var, anchor="w", font=('Arial', 10, 'bold'))
        self.status_label.grid(row=row, column=0, columnspan=3, sticky="we", pady=(8, 0))
        row += 1

        self.notebook = ttk.Notebook(frm)
        self.notebook.grid(row=row, column=0, columnspan=3, sticky="nsew", pady=(10, 0))

        self.tab_log = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_log, text="Execution Log")
        
        self.output = tk.Text(self.tab_log, height=15, wrap=tk.WORD)
        log_scroll = ttk.Scrollbar(self.tab_log, command=self.output.yview)
        self.output.configure(yscrollcommand=log_scroll.set)
        self.output.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.output.tag_configure("pass", foreground="#1B5E20")
        self.output.tag_configure("fail", foreground="#B71C1C")
        self.output.tag_configure("warn", foreground="#E65100")
        self.output.tag_configure("cmd", foreground="#1565C0")
        self.output.tag_configure("error", foreground="#B71C1C")
        self.output.tag_configure("commg", foreground="#800080")

        self.tab_summary = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_summary, text="Results Summary")

        self.summary_top = tk.Frame(self.tab_summary)
        self.summary_top.pack(fill=tk.X, padx=5, pady=5)
        
        self.lbl_stats = tk.Label(self.summary_top, text="Total: 0 | PASS: 0 | FAIL: 0 | WARN: 0 | SKIP: 0", font=('Arial', 10, 'bold'))
        self.lbl_stats.pack(side=tk.LEFT)

        self.current_filter = "ALL"
        
        tk.Button(self.summary_top, text="Clear Summary", command=self.clear_summary_data).pack(side=tk.RIGHT, padx=(10, 2))
        tk.Button(self.summary_top, text="Show All", command=lambda: self.filter_tree("ALL")).pack(side=tk.RIGHT, padx=2)
        tk.Button(self.summary_top, text="Show Fails", bg="#FFCCCB", command=lambda: self.filter_tree("FAIL")).pack(side=tk.RIGHT, padx=2)
        tk.Button(self.summary_top, text="Show Warns", bg="#FFF3E0", command=lambda: self.filter_tree("WARN")).pack(side=tk.RIGHT, padx=2)
        tk.Button(self.summary_top, text="Show Passes", bg="#90EE90", command=lambda: self.filter_tree("PASS")).pack(side=tk.RIGHT, padx=2)

        tree_columns = ("device", "index", "tag", "expected", "actual", "verdict", "error")
        self.tree = ttk.Treeview(self.tab_summary, columns=tree_columns, show="headings", height=15)
        
        self.tree.heading("device", text="Device")
        self.tree.heading("index", text="Index")
        self.tree.heading("tag", text="Tag")
        self.tree.heading("expected", text="Expected")
        self.tree.heading("actual", text="Actual")
        self.tree.heading("verdict", text="Verdict")
        self.tree.heading("error", text="Error Details")

        self.tree.column("device", width=120, anchor=tk.CENTER)
        self.tree.column("index", width=80, anchor=tk.CENTER)
        self.tree.column("tag", width=120)
        self.tree.column("expected", width=200)
        self.tree.column("actual", width=200)
        self.tree.column("verdict", width=80, anchor=tk.CENTER)
        self.tree.column("error", width=250)

        tree_scroll = ttk.Scrollbar(self.tab_summary, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.tag_configure("PASS", background="#E8F5E9") 
        self.tree.tag_configure("FAIL", background="#FFEBEE") 
        self.tree.tag_configure("WARN", background="#FFF3E0") 
        self.tree.tag_configure("SKIP", background="#EEEEEE")

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        self.all_results_data = []

        frm.grid_columnconfigure(1, weight=1)
        frm.grid_rowconfigure(row, weight=1)

        try:
            cur = (self.camera_id_var.get() or "").strip()
            if cur:
                self.camera_combo["values"] = [cur]
        except Exception:
            pass
        self.refresh_tags()

        try:
            self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        except Exception:
            pass

        self.root.after(50, self._drain_queue)

    def _update_accumulated_time(self):
        # Only add time if we are active AND ADB is not running
        if hasattr(self, 'batch_start_time') and getattr(self, "_commg_is_active_run", False) and not getattr(self, "_is_adb_running", False):
            now = time.time()
            elapsed = now - self.batch_start_time
            self._settings["accumulated_time"] = self._settings.get("accumulated_time", 0.0) + elapsed
            self.batch_start_time = now
            self._persist_settings()
        elif getattr(self, "_is_adb_running", False) and hasattr(self, 'batch_start_time'):
            # Fast-forward the origin time so we don't count the ADB minutes
            self.batch_start_time = time.time()

    def _on_index_changed(self, *args):
        if getattr(self, "_commg_is_active_run", False): return
        
        idx = self.index_var.get().strip()
        if not idx or "," in idx: return
        
        if not getattr(self, "_index_to_tag_cache", None):
            try:
                self._index_to_tag_cache = _build_index_to_tag_map(self.excel_var.get().strip())
            except Exception:
                self._index_to_tag_cache = {}
                
        clean_idx = idx
        if clean_idx.isdigit(): clean_idx = str(int(clean_idx))
        
        tag = self._index_to_tag_cache.get(clean_idx, "")
        if tag and self.tag_var.get() != tag:
            self.tag_var.set(tag)

    def _commg_add_ip(self):
        new_ip = self.new_ip_entry.get().strip()
        if not new_ip:
            messagebox.showwarning("Empty Input", "Please enter an IP address before clicking Add.")
            return
            
        current_ips = list(self.ip_listbox.get(0, tk.END))
        if new_ip in current_ips:
            messagebox.showwarning("Duplicate IP", f"The IP address {new_ip} is already in the list!")
            return
            
        self.ip_listbox.insert(tk.END, new_ip)
        self.new_ip_entry.delete(0, tk.END)
        self._sync_devices_to_ips()

    def _commg_remove_ip(self):
        selected = self.ip_listbox.curselection()
        if not selected:
            messagebox.showinfo("No Selection", "Please select an IP address from the list to remove.")
            return
            
        ip_to_remove = self.ip_listbox.get(selected[0])
        confirm = messagebox.askyesno(
            "Confirm Removal", 
            f"Are you sure you want to remove this IP address?\n\n{ip_to_remove}"
        )
        
        if confirm:
            self.ip_listbox.delete(selected[0])
            self._sync_devices_to_ips()

    def _commg_browse_batch(self):
        p = filedialog.askopenfilename(
            title="Select Batch File",
            filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv"), ("All Files", "*.*")]
        )
        if p:
            self.commg_batch_file_var.set(p)

    def _commg_ping_ip(self, ip):
        try:
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            response = subprocess.call(['ping', '-n', '1', '-w', '2000', ip], startupinfo=startupinfo)
            return response == 0
        except Exception:
            return False

    def _robust_connection_check(self, ip, port, is_cmd=True):
        import time
        
        connected = False
        for attempt in range(3):
            self.q.put(f"[Connection Check] Ping {ip} (Attempt {attempt+1}/3)...\n")
            if self._commg_ping_ip(ip):
                if is_cmd:
                    import socket
                    try:
                        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        s.settimeout(2.0)
                        s.connect((ip, port))
                        s.close()
                        connected = True
                        break
                    except Exception as e:
                        self.q.put(f"[Connection Check] Ping OK, but target port {port} is closed/unreachable: {e}\n")
                else:
                    connected = True
                    break
            time.sleep(2.0)
            
        if connected:
            return "CONNECTED"
            
        self.q.put(f"[Connection Check] Connection lost for {ip}. Skipping device for this command...\n")
        return "SKIP"

    def _commg_find_handles(self):
        if not HAS_PYWINAUTO: return []
        handles = []
        desktop = Desktop(backend="uia")
        for win in desktop.windows():
            text = win.window_text()
            if text and self.WINDOW_SEARCH_TERM in text:
                handles.append(win.handle)
        return handles

    def _commg_ensure_connection(self):
        if not HAS_PYWINAUTO: return False
        if self.commg_handles:
            try:
                app = Application(backend="uia").connect(handle=self.commg_handles[0])
                app.window(handle=self.commg_handles[0]).exists()
                return True
            except:
                self.commg_handles = [] 

        self.q.put("[CommG] CommuniGATOR not found. Launching...\n")
        found_handles = self._commg_find_handles()
        
        if not found_handles:
            if not os.path.exists(self.COMMG_PATH):
                self.q.put(f"[CommG] ERROR: Cannot find shortcut at {self.COMMG_PATH}\n")
                return False
                
            os.startfile(self.COMMG_PATH) 
            time.sleep(3.5) 
            found_handles = self._commg_find_handles()
            
        if not found_handles:
            self.q.put("[CommG] ERROR: Software launched but couldn't attach.\n")
            return False

        self.commg_handles = found_handles[:1]
        self.q.put("[CommG] Successfully hooked into CommuniGATOR.\n")
        return True

    def _commg_send_command_thread_impl(self, payloads, batch_items):
        ips = list(self.ip_listbox.get(0, tk.END))
        if not ips:
            self.q.put("[CommG] No IPs configured!\n")
            self.q.put((_EVT_COMMG_DONE, "ABORT", batch_items))
            return

        if not self._commg_ensure_connection():
            self.q.put("[CommG] Failed to connect to CommuniGATOR.\n")
            self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
            return

        handle = self.commg_handles[0]
        try:
            app = Application(backend="uia").connect(handle=handle)
            main_win = app.window(handle=handle)
            toolbar = main_win.child_window(auto_id="59392", control_type="ToolBar")
            input_field = main_win.child_window(auto_id="1004", control_type="Edit")
            
            for i, ip in enumerate(ips):
                if i < len(self.device_rows) and self.device_rows[i].get("current_active_lang") == "SKIP_DEVICE":
                    continue
                    
                payload = payloads[i] if i < len(payloads) else payloads[-1]
                
                status = self._robust_connection_check(ip, 23, is_cmd=False)
                if status == "PAUSE" or status == "ABORT":
                    self.q.put("[CommG] Sequence paused by user.\n")
                    self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
                    return
                elif status == "SKIP":
                    continue

                try:
                    with self.type_lock:
                        main_win.set_focus()
                        
                        toolbar.button(0).click() 
                        time.sleep(1.0)
                        popup = app.top_window() 
                        popup.type_keys(ip + "{ENTER}", set_foreground=True)
                        time.sleep(1.5) 
                        
                        toolbar.button(1).click()
                        time.sleep(1.0)
                        
                        input_field.set_focus()
                        input_field.type_keys("^a{BACKSPACE}0121FF{ENTER}", set_foreground=True)
                        time.sleep(1.5)
                        
                        input_field.set_focus()
                        input_field.type_keys("^a{BACKSPACE}" + payload + "{ENTER}", set_foreground=True)
                        self.q.put(f"[CommG] Sent {payload} to {ip}\n")
                
                except Exception as inner_e:
                    self.q.put(f"[CommG] ERROR on {ip}: {inner_e}. Skipping to next.\n")
                    continue
                    
            # --- FASTER WAIT TIME ---
            self.q.put("[CommG] Waiting 2s for UI to update...\n")
            time.sleep(2.0) 
            
        except Exception as e:
            self.q.put(f"[CommG] FATAL ERROR: {e}\n")
            self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
            return

        self.q.put((_EVT_COMMG_DONE, None, None))

    def _cmd_telnet_thread(self, payloads, batch_items):
        import socket
        
        ips = [ip.strip() for ip in self.ip_listbox.get(0, tk.END) if ip.strip()]
        if not ips:
            self.q.put("[CMD] No IPs configured!\n")
            self.q.put((_EVT_COMMG_DONE, "ABORT", batch_items))
            return

        try:
            port = int(self.telnet_port_var.get().strip() or 23)
        except ValueError:
            port = 23
            
        if not hasattr(self, "active_sockets"):
            self.active_sockets = []

        for i, ip in enumerate(ips):
            if i < len(self.device_rows) and self.device_rows[i].get("current_active_lang") == "SKIP_DEVICE":
                continue
                
            payload = payloads[i] if i < len(payloads) else payloads[-1]
            
            status = self._robust_connection_check(ip, port, is_cmd=True)
            if status == "PAUSE" or status == "ABORT":
                self.q.put("[CMD] Sequence paused by user.\n")
                self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
                return
            elif status == "SKIP":
                continue

            try:
                self.q.put(f"[CMD] Opening background connection to {ip}...\n")
                
                s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                s.settimeout(5.0)
                s.connect((ip, port))
                
                time.sleep(1.0)
                
                s.sendall(f"{payload}\r\n".encode('ascii'))
                self.q.put(f"[CMD] Sent {payload} to {ip} (Session running invisibly)\n")
                
                self.active_sockets.append(s)
            
            except Exception as inner_e:
                self.q.put(f"[CMD] ERROR on {ip}: {inner_e}. Skipping to next.\n")
                continue

        # --- FASTER WAIT TIME ---
        self.q.put("[CMD] Waiting 2s for devices to process...\n")
        time.sleep(2.0) 
        self.q.put((_EVT_COMMG_DONE, None, None))

    def _putty_thread(self, payloads, batch_items):
        ips = [ip.strip() for ip in self.ip_listbox.get(0, tk.END) if ip.strip()]
        if not ips:
            self.q.put("[PuTTY] No IPs configured!\n")
            self.q.put((_EVT_COMMG_DONE, "ABORT", batch_items))
            return

        try:
            port = str(int(self.telnet_port_var.get().strip() or 23))
        except ValueError:
            port = "23"

        if not hasattr(self, "active_putty_apps"):
            self.active_putty_apps = {}

        putty_exe = None
        saved_putty = self._settings.get("putty_path", "")
        
        if os.path.exists(saved_putty):
            putty_exe = saved_putty
        else:
            putty_paths = [
                r"C:\Program Files\PuTTY\putty.exe",
                r"C:\Program Files (x86)\PuTTY\putty.exe"
            ]
            for p in putty_paths:
                if os.path.exists(p):
                    putty_exe = p
                    break

            if not putty_exe:
                self.q.put("[PuTTY] Searching for putty.exe...\n")
                putty_exe = filedialog.askopenfilename(title="Locate putty.exe", filetypes=[("Executable", "putty.exe")])
                if putty_exe:
                    self._settings["putty_path"] = putty_exe
                    _save_settings(self._settings)
                else:
                    self.q.put("[PuTTY] ERROR: putty.exe not found or selected.\n")
                    self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
                    return

        for i, ip in enumerate(ips):
            if i < len(self.device_rows) and self.device_rows[i].get("current_active_lang") == "SKIP_DEVICE":
                continue
                
            payload = payloads[i] if i < len(payloads) else payloads[-1]

            status = self._robust_connection_check(ip, int(port), is_cmd=True)
            if status == "PAUSE" or status == "ABORT":
                self.q.put("[PuTTY] Sequence paused by user.\n")
                self.q.put((_EVT_COMMG_DONE, "PAUSE", batch_items))
                return
            elif status == "SKIP":
                continue

            try:
                with self.type_lock:
                    if ip not in self.active_putty_apps:
                        self.q.put(f"[PuTTY] Launching Configuration GUI for {ip}...\n")
                        
                        app = Application(backend="uia").start(putty_exe)
                        config_win = app.window(title="PuTTY Configuration")
                        config_win.wait("ready", timeout=5)
                        
                        try:
                            config_win.child_window(title="Session", control_type="TreeItem").click_input()
                            time.sleep(0.2)
                        except: pass

                        try:
                            host_edit = config_win.child_window(title_re="Host Name.*", control_type="Edit")
                            host_edit.click_input()
                            host_edit.type_keys("^a{BACKSPACE}" + ip, with_spaces=True)
                        except:
                            config_win.set_focus()
                            config_win.type_keys("%n^a{BACKSPACE}" + ip)
                            
                        try:
                            port_edit = config_win.child_window(title="Port", control_type="Edit")
                            port_edit.click_input()
                            port_edit.type_keys("^a{BACKSPACE}" + port)
                        except:
                            config_win.set_focus()
                            config_win.type_keys("%p^a{BACKSPACE}" + port)
                            
                        try:
                            config_win.set_focus()
                            config_win.type_keys("{TAB}{TAB}rr")
                            time.sleep(0.3)
                        except Exception as e:
                            self.q.put(f"[PuTTY] Keyboard shortcut failed: {e}\n")

                        try:
                            open_btn = config_win.child_window(title="Open", control_type="Button")
                            open_btn.click_input()
                        except:
                            config_win.set_focus()
                            config_win.type_keys("{ENTER}")
                        
                        time.sleep(1.5)
                        term_win = app.window(title_re=r".*PuTTY.*")
                        term_win.wait("ready", timeout=5)
                        
                        self.active_putty_apps[ip] = (app, term_win)

                app, term_win = self.active_putty_apps[ip]
                term_win.set_focus()
                term_win.type_keys(payload + "{ENTER}", set_foreground=True)
                self.q.put(f"[PuTTY] Sent {payload} to {ip}\n")

            except Exception as inner_e:
                self.q.put(f"[PuTTY] ERROR on {ip}: {inner_e}. Skipping to next.\n")
                continue

        # --- FASTER WAIT TIME ---
        self.q.put("[PuTTY] Waiting 2s for devices to process...\n")
        time.sleep(2.0) 
        self.q.put((_EVT_COMMG_DONE, None, None))

    def _integration_send_command_thread(self, payloads, batch_items):
        integration_type = self.integration_type_var.get()
        if integration_type == "CMD":
            self._cmd_telnet_thread(payloads, batch_items)
        elif integration_type == "PuTTY":
            self._putty_thread(payloads, batch_items)
        else:
            self._commg_send_command_thread_impl(payloads, batch_items)
            
    def get_radio_serial(self, ip, port=8501):
        import socket
        import re
        self.q.put(f"[Telnet] Fetching serial for {ip} via port {port}...\n")
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5.0)
            s.connect((ip, port))
            s.sendall(b"ver\r\n")
            output = ""
            for _ in range(15):
                chunk = s.recv(4096).decode('utf-8', errors='ignore')
                output += chunk
                match = re.search(r"Serial Number\s*:\s*([A-Za-z0-9]+)", output)
                if match:
                    s.close()
                    serial = match.group(1).strip()
                    self.q.put(f"[Telnet] Found serial: {serial} for IP {ip}\n")
                    return serial
            s.close()
        except Exception as e:
            self.q.put(f"[Telnet Error] Failed to get serial for {ip}: {e}\n")
        return None

    def _start_language_iteration(self):
        self._update_accumulated_time()
        
        if self._current_lang_idx >= self._max_lang_idx:
            self._commg_is_active_run = False
            self.status_var.set("Automation Batch Complete (All Languages)")
            self.status_label.configure(fg="green")
            self._set_running(False)
            
            threading.Thread(target=self._write_final_excel_summary, daemon=True).start()

            total_elapsed = self._settings.get("accumulated_time", 0.0)
            m, s = divmod(total_elapsed, 60)
            h, m = divmod(m, 60)
            time_str = f"{int(h):02d}:{int(m):02d}:{int(s):02d}"

            p, f, w, skip = 0, 0, 0, 0
            for d in self.all_results_data:
                v = d.get("verdict", "")
                if v == "PASS": p += 1
                elif v == "FAIL": f += 1
                elif v == "WARN": w += 1
                elif v == "SKIP": skip += 1
            total_cmds = p + f + w + skip

            self.q.put(f"\n{'='*60}\n")
            self.q.put(f"BATCH SEQUENCE COMPLETED (All Languages)\n")
            self.q.put(f"Total Time Taken: {time_str}\n")
            self.q.put(f"Total Executions: {total_cmds} (PASS: {p}, FAIL: {f}, WARN: {w}, SKIP: {skip})\n")
            self.q.put(f"{'='*60}\n")
            return

        self.q.put(f"\n{'='*60}\n")
        self.q.put(f"[Batch] Starting Language Iteration {self._current_lang_idx + 1} of {self._max_lang_idx}\n")
        self.q.put(f"{'='*60}\n")

        self.status_var.set(f"Changing Languages (Iter {self._current_lang_idx + 1}/{self._max_lang_idx})...")
        self.status_label.configure(fg="blue")
        
        # --- FIX: Pause the stopwatch while ADB runs! ---
        self._update_accumulated_time()
        self._is_adb_running = True
        # ------------------------------------------------

        def worker():
            import ctypes
            import subprocess
            import time
            import os
            import socket
            
            try:
                safe_port = int(self.telnet_port_var.get().strip())
            except Exception:
                safe_port = 23
                
            try:
                is_admin = os.getuid() == 0
            except AttributeError:
                is_admin = ctypes.windll.shell32.IsUserAnAdmin() != 0
                
            if not is_admin:
                self.q.put("[FATAL ERROR] This script MUST be run as Administrator to use the Network Routing workaround!\n")
                self.q.put((_EVT_COMMG_DONE, "ABORT", []))
                return

            ips = list(self.ip_listbox.get(0, tk.END))
            creationflags = getattr(subprocess, 'CREATE_NO_WINDOW', 0) if os.name == 'nt' else 0

            self.q.put("\n[System] Performing deep ADB cleanup before changing languages...\n")
            try:
                subprocess.run(["adb", "kill-server"], creationflags=creationflags, timeout=5)
                time.sleep(2.0)
                subprocess.run(["adb", "start-server"], creationflags=creationflags, timeout=5)
            except Exception:
                pass

            try:
                import sys
                sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                import main as android_lang_changer
            except Exception as e:
                self.q.put(f"[LangChange Error] Could not load main.py: {e}\n")
                android_lang_changer = None

            for i, ip in enumerate(ips):
                r = self.device_rows[i] if i < len(self.device_rows) else None
                if r:
                    langs = self._get_device_langs(r)
                    
                    if self._current_lang_idx >= len(langs):
                        r["current_active_lang"] = "SKIP_DEVICE"
                        self.q.put(f"\n[Sequence] Device {ip} has finished all its languages. Skipping ADB Language Setup.\n")
                        continue
                        
                    target_lang = langs[self._current_lang_idx]
                    r["current_active_lang"] = target_lang

                    # --- FIX: BYPASS ADB IF USER SELECTED 'SKIP' ---
                    if target_lang == "SKIP_DEVICE":
                        self.q.put(f"\n[Sequence] Device {ip} is set to SKIP Iteration {self._current_lang_idx + 1}. Skipping ADB Setup.\n")
                        continue
                    # -----------------------------------------------
                else:
                    target_lang = "English"

                while True:
                    self.q.put(f"\n{'-'*60}\n")
                    self.q.put(f"[Sequence] Processing Device: {ip} | Target Language: {target_lang}\n")
                    self.q.put(f"{'-'*60}\n")

                    self.q.put(f"[Telnet] Switching {ip} to AP Mode...\n")
                    try:
                        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        s.settimeout(5.0)
                        s.connect((ip, safe_port))
                        s.sendall(b"0121FF\r\n")
                        time.sleep(1.0)
                        s.sendall(b"03011001\r\n")
                        s.close()
                        
                        self.q.put(f"[Wait] 30 seconds for Windows and {ip} to re-establish network...\n")
                        time.sleep(30.0)
                    except Exception as e:
                        self.q.put(f"[Warning] Telnet failed for {ip}: {e}. Device might already be in AP mode. Proceeding...\n")

                    self.q.put(f"[Network] Verifying if {ip} is reachable on the network...\n")
                    is_pingable = False
                    for ping_try in range(10):
                        if self._commg_ping_ip(ip):
                            self.q.put(f"[Network] {ip} is online!\n")
                            is_pingable = True
                            break
                        time.sleep(2.0)
                    
                    if not is_pingable:
                        self.q.put(f"[Network Warning] {ip} is not responding to pings. ADB might fail.\n")

                    self.q.put(f"[ADB] Clearing old connections...\n")
                    try:
                        subprocess.run(["adb", "disconnect"], creationflags=creationflags, capture_output=True, timeout=5)
                    except Exception:
                        pass
                    
                    self.q.put(f"[ADB] Connecting directly to {ip}:5500...\n")
                    try:
                        res = subprocess.run(["adb", "connect", f"{ip}:5500"], creationflags=creationflags, capture_output=True, text=True, timeout=5)
                        out_text = res.stdout.lower()
                    except Exception:
                        out_text = "failed"
                    
                    target_serial = f"{ip}:5500"
                    
                    if "cannot connect" in out_text or "failed" in out_text or "offline" in out_text:
                        self.q.put(f"[Network] Direct connect failed. Using 199.0.0.4 static route fallback...\n")
                        subprocess.run(["route", "delete", "199.0.0.4"], creationflags=creationflags, capture_output=True)
                        subprocess.run(["route", "add", "199.0.0.4", "mask", "255.255.255.255", ip], creationflags=creationflags, capture_output=True)
                        try:
                            subprocess.run(["adb", "connect", "199.0.0.4:5500"], creationflags=creationflags, timeout=5)
                        except Exception:
                            pass
                        target_serial = "199.0.0.4:5500"

                    self.q.put(f"[ADB] Waiting for device {target_serial} to authorize...\n")
                    device_ready = False
                    for _ in range(15):
                        try:
                            out = subprocess.check_output(["adb", "devices"], text=True, creationflags=creationflags, timeout=3)
                            if f"{target_serial}\tdevice" in out:
                                device_ready = True
                                break
                            if f"{target_serial}\toffline" in out or target_serial not in out:
                                subprocess.run(["adb", "disconnect", target_serial], creationflags=creationflags, capture_output=True, timeout=3)
                                time.sleep(1.0)
                                subprocess.run(["adb", "connect", target_serial], creationflags=creationflags, capture_output=True, timeout=3)
                        except subprocess.TimeoutExpired:
                            subprocess.run(["adb", "kill-server"], creationflags=creationflags, capture_output=True)
                            time.sleep(1.0)
                            subprocess.run(["adb", "start-server"], creationflags=creationflags, capture_output=True)
                        except Exception:
                            pass
                        time.sleep(2.0)
                    
                    if not device_ready:
                        self.q.put(f"[ADB Warning] Device {target_serial} is still offline. UI Automation may fail.\n")
                    else:
                            time.sleep(1.0)
                            try:
                                self.q.put("[ADB] Ensuring device screen is awake...\n")
                                
                                # 1. Force the screen to NEVER sleep while plugged into USB
                                subprocess.run(["adb", "-s", target_serial, "shell", "svc", "power", "stayon", "true"], creationflags=creationflags, timeout=3)
                                
                                # 2. Wake the screen (if it happened to be off)
                                subprocess.run(["adb", "-s", target_serial, "shell", "input", "keyevent", "224"], creationflags=creationflags, timeout=3)
                                time.sleep(1.0)
                                
                                # 3. Press 'Back' twice to clear any random popups (battery warnings, etc.)
                                subprocess.run(["adb", "-s", target_serial, "shell", "input", "keyevent", "4"], creationflags=creationflags, timeout=3)
                                subprocess.run(["adb", "-s", target_serial, "shell", "input", "keyevent", "4"], creationflags=creationflags, timeout=3)
                                time.sleep(0.5)
                                
                                # 4. Press 'Home' to ensure UI Automator starts cleanly from the home screen
                                subprocess.run(["adb", "-s", target_serial, "shell", "input", "keyevent", "3"], creationflags=creationflags, timeout=3)
                                time.sleep(1.0)
                                
                            except Exception as e:
                                self.q.put(f"[ADB Warning] Failed to send wake commands: {e}\n")

                    success = False
                    if android_lang_changer:
                        try:
                            device = android_lang_changer.get_device(target_serial)
                            if device:
                                self.q.put(f"[UI Auto] Executing language change on {ip} to '{target_lang}'...\n")
                                
                                def check_system_locale(target_lang_name):
                                    try:
                                        res = subprocess.run(
                                            ["adb", "-s", target_serial, "shell", "getprop", "persist.sys.locale"], 
                                            capture_output=True, text=True, creationflags=creationflags, timeout=5
                                        )
                                        current_locale = res.stdout.strip().lower()
                                        
                                        if not current_locale:
                                            res = subprocess.run(
                                                ["adb", "-s", target_serial, "shell", "settings", "get", "system", "system_locales"], 
                                                capture_output=True, text=True, creationflags=creationflags, timeout=5
                                            )
                                            current_locale = res.stdout.strip().lower()

                                        prefix_map = {
                                            "Czech": "cs", "Simplified Chinese": "zh", "Portuguese": "pt",
                                            "Spanish": "es", "Polish": "pl", "Italian": "it", "Turkish": "tr",
                                            "Hungarian": "hu", "English": "en", "Japanese": "ja",
                                            "Russian": "ru", "French": "fr", "German": "de", "Korean": "ko",
                                            "Traditional Chinese": "zh", "Arabic": "ar", "Hebrew": "iw"
                                        }
                                        prefix = prefix_map.get(target_lang_name, target_lang_name[:2].lower())
                                        
                                        if prefix not in current_locale: return False
                                        if target_lang_name == "Traditional Chinese" and not any(x in current_locale for x in ["tw", "hk", "hant"]): return False
                                        if target_lang_name == "Simplified Chinese" and not any(x in current_locale for x in ["cn", "hans"]): return False
                                        return True
                                    except Exception:
                                        return False

                                if check_system_locale(target_lang):
                                    self.q.put(f"[UI Auto] Device {ip} is already in '{target_lang}'. Skipping UI automation!\n")
                                    success = True
                                else:
                                    for auto_try in range(5):
                                        android_lang_changer.change_language(device, target_lang)
                                        time.sleep(3.0)
                                        if check_system_locale(target_lang):
                                            success = True
                                            self.q.put(f"[UI Auto] Success: Verified system locale changed to {target_lang}!\n")
                                            break
                                            
                                        self.q.put(f"[UI Auto] Retry {auto_try+1}/5: Drag finished, but system locale did not update. Retrying...\n")
                                        time.sleep(2.0)
                            else:
                                self.q.put(f"[ADB Error] UI Automator could not find device at {target_serial}.\n")
                        except Exception as e:
                            self.q.put(f"[UI Auto Error] Exception occurred: {e}\n")

                    if not success:
                        self.q.put(f"[UI Auto Warning] Language change to '{target_lang}' FAILED on {ip}.\n")
                        self.q.put("[UI Auto] Paused. Waiting for manual language confirmation...\n")
                        
                        retry = messagebox.askretrycancel(
                            "Language Change Failed",
                            f"Device {ip} could not automatically switch to '{target_lang}'.\n\n"
                            f"Click 'Retry' to restart the automation sequence (AP Mode -> ADB -> UI Auto) for this device.\n\n"
                            f"Click 'Cancel' to abort the entire batch."
                        )
                        if not retry:
                            self.q.put((_EVT_COMMG_DONE, "ABORT", []))
                            return
                        else:
                            self.q.put(f"[UI Auto] Retrying language sequence for {ip} from the start...\n")
                            continue

                    self.q.put(f"[ADB] Sending MUX return intent to {target_serial}...\n")
                    try:
                        subprocess.run(["adb", "-s", target_serial, "shell", "am", "broadcast", "-a", "msi.factorymuxrouter.intent.action.BP_ROUTE_MUX", "-n", "com.motorolasolutions.factorymuxrouter/.MuxRouterBroadcastReceiver"], creationflags=creationflags, timeout=5)
                    except Exception:
                        pass
                    
                    self.q.put(f"[Network] Disconnecting ADB and deleting route...\n")
                    try:
                        subprocess.run(["adb", "disconnect", target_serial], creationflags=creationflags, timeout=5)
                    except Exception:
                        pass
                    if target_serial == "199.0.0.4:5500":
                        subprocess.run(["route", "delete", "199.0.0.4"], creationflags=creationflags, capture_output=True)

                    self.q.put(f"[Wait] 15 seconds for {ip} to completely boot back to MUX Mode...\n")
                    time.sleep(15.0)
                    
                    break

            self.q.put("\n[Pre-Check] Verifying MUX connection on all devices before string verification...\n")
            all_online = False
            for attempt in range(5):
                all_online = True
                for ip in ips:
                    ser = self.get_radio_serial(ip, port=safe_port)
                    if not ser:
                        all_online = False
                if all_online:
                    self.q.put("[LangChange] Successfully verified MUX connection on all devices!\n")
                    break
                self.q.put(f"[Pre-Check] Checking network (Attempt {attempt+1}/5)...\n")
                time.sleep(5.0)

            if not all_online:
                retry = messagebox.askretrycancel(
                    "Action Required: Manual MUX Check",
                    "Some devices failed to return to MUX mode after the language change.\n\n"
                    "1. Please check the physical devices.\n"
                    "2. Swipe down on their screens and ensure USB is set to 'MUX'.\n"
                    "3. Click 'Retry' when they are fully connected to start the verification."
                )
                if not retry:
                    self.q.put((_EVT_COMMG_DONE, "ABORT", []))
                    return

            self.q.put("[LangChange] Preparation complete! Resuming batch verification...\n")
            self.q.put(("__LANG_CHANGE_DONE__", None))
            
        threading.Thread(target=worker, daemon=True).start()
        
    def _run_next_commg_step(self):
        if not getattr(self, "_commg_pending_queue", []):
            self._commg_is_active_run = False
            self._set_running(False)
            self.status_var.set("Automation Batch Complete")
            self.status_label.configure(fg="green")
            self._update_summary_ui(self.current_filter)
            return

        ips = list(self.ip_listbox.get(0, tk.END))
        num_ips = len(ips) if ips else 1
        mode = self.commg_mode_var.get()
        
        batch_items = []
        
        if mode == "Batch":
            if getattr(self, "_commg_pending_queue", []):
                cmd_item = self._commg_pending_queue.pop(0)
                for _ in range(num_ips):
                    batch_items.append(cmd_item)
        else:
            if getattr(self, "_commg_pending_queue", []):
                cmd_item = self._commg_pending_queue.pop(0)
                for _ in range(num_ips):
                    batch_items.append(cmd_item)

        if not batch_items:
            return

        self._current_batch_items = [cmd_item] 

        cmds = []
        tags = []
        indices = []

        for item in batch_items:
            cmd = item
            tag = ""
            if isinstance(item, tuple):
                cmd, tag = item
                
            cmds.append(str(cmd).strip())
            
            tag = str(tag).strip()
            if tag.lower() == 'nan':
                tag = ""
            
            if tag.upper() in ["SKIP", "NO_VERIFY", "NONE"]:
                tags.append("SKIP_VERIFY")
                indices.append("SKIP_VERIFY")
            elif tag:
                tags.append(tag)
                indices.append("")
            else:
                if ":" in cmd:
                    extracted_idx = cmd.split(":")[-1].strip()
                else:
                    extracted_idx = cmd.strip()
                
                clean_idx = extracted_idx
                if clean_idx.isdigit():
                    clean_idx = str(int(clean_idx))
                    
                if not getattr(self, "_index_to_tag_cache", None):
                    try:
                        self._index_to_tag_cache = _build_index_to_tag_map(self.excel_var.get().strip())
                    except Exception:
                        self._index_to_tag_cache = {}

                found_tag = self._index_to_tag_cache.get(clean_idx, "")
                tags.append(found_tag)
                indices.append(extracted_idx)

        self._current_batch_cmd = cmds[0]
        self.tag_var.set(",".join(tags))
        self.index_var.set(",".join(indices))
        self._current_batch_cmds = cmds
        tool = self.integration_type_var.get()
        
        active_ips = 0
        for i, ip in enumerate(ips):
            if i < len(self.device_rows) and self.device_rows[i].get("current_active_lang") == "SKIP_DEVICE":
                continue
            active_ips += 1

        self.q.put(f"\n[{tool}] Dispatched Command: {cmds[0]} to {active_ips} active device(s).\n", ("commg",))

        self.status_var.set(f"Automation Running: {cmds[0]}...")
        self.status_label.configure(fg="purple")
        self._set_running(True)

        threading.Thread(target=self._integration_send_command_thread, args=(cmds, batch_items), daemon=True).start()

    def _get_device_langs(self, row_dict):
        # Read exactly what is in the columns so Iterations align perfectly!
        out = []
        for v in row_dict.get("lang_vars", []):
            val = v.get().strip()
            if not val or val.upper() == "SKIP":
                out.append("SKIP_DEVICE")
            else:
                out.append(val)
        return out

    def _add_lang_column(self):
        self.num_lang_columns += 1
        for row in self.device_rows:
            row["lang_vars"].append(tk.StringVar(value="English"))
        self._regrid_device_rows()

    def _remove_lang_column(self):
        if self.num_lang_columns > 1:
            self.num_lang_columns -= 1
            for row in self.device_rows:
                if len(row["lang_vars"]) > self.num_lang_columns:
                    row["lang_vars"].pop()
            self._regrid_device_rows()
            
    def _update_dropdown_options(self, cb, current_var):
        """Dynamically filters the dropdown list so you can't pick taken languages (except SKIP)."""
        used_langs = set()
        
        # 1. Find all languages currently selected anywhere on the grid
        for r in self.device_rows:
            for v in r.get("lang_vars", []):
                val = v.get().strip()
                if val and val.upper() != "SKIP":
                    used_langs.add(val)
                    
        my_val = current_var.get().strip()
        valid_options = []
        
        # 2. Build a new list of allowed languages
        for lang in self.available_languages:
            # Allow it if it's "SKIP", if it's our currently selected value, or if nobody else is using it
            if lang.upper() == "SKIP" or lang == my_val or lang not in used_langs:
                valid_options.append(lang)
                
        cb["values"] = valid_options

    def _regrid_device_rows(self):
        for widget in self.devices_grid_frame.winfo_children():
            widget.destroy()

        # Headers
        tk.Label(self.devices_grid_frame, text="ID", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 10))
        tk.Label(self.devices_grid_frame, text="Name", font=("Arial", 9, "bold")).grid(row=0, column=1, sticky="w", padx=(0, 10))
        
        for c in range(self.num_lang_columns):
            tk.Label(self.devices_grid_frame, text=f"Iter {c+1}", font=("Arial", 9, "bold", "underline")).grid(row=0, column=2+c, sticky="w", padx=2)
            
        btn_frame = tk.Frame(self.devices_grid_frame)
        btn_frame.grid(row=0, column=2+self.num_lang_columns, sticky="w", padx=(10, 0))
        tk.Button(btn_frame, text="+ Add Lang", command=self._add_lang_column, padx=4, pady=0, bg="#ADD8E6").pack(side=tk.LEFT, padx=2)
        if self.num_lang_columns > 1:
            tk.Button(btn_frame, text="- Remove", command=self._remove_lang_column, padx=4, pady=0, bg="#FFCCCB").pack(side=tk.LEFT, padx=2)

        # Rows
        for idx, row in enumerate(self.device_rows):
            row_idx = idx + 1
            
            lbl_id = tk.Label(self.devices_grid_frame, text=f"D{row['id']}")
            lbl_id.grid(row=row_idx, column=0, sticky="w", padx=(0, 10), pady=2)
            
            ent_name = tk.Entry(self.devices_grid_frame, textvariable=row["var_name"], width=15)
            ent_name.grid(row=row_idx, column=1, sticky="ew", pady=2, padx=(0, 10))
            
            for c, l_var in enumerate(row["lang_vars"]):
                cb = ttk.Combobox(self.devices_grid_frame, textvariable=l_var, state="readonly", width=16)
                cb.grid(row=row_idx, column=2+c, sticky="w", padx=2, pady=2)
                
                # --- FIX: Hook up the dynamic dropdown filter! ---
                cb.configure(postcommand=lambda current_cb=cb, current_var=l_var: self._update_dropdown_options(current_cb, current_var))
                self._update_dropdown_options(cb, l_var) # Set initial values
                # -------------------------------------------------
                
        # Helper text
        tk.Label(self.devices_grid_frame, text="Devices auto-sync with IP addresses.", fg="gray", font=("Arial", 8)).grid(row=len(self.device_rows)+1, column=0, columnspan=4, sticky="w", pady=(8, 2))
        
    def _add_device_row(self, did: int, name: str = "", lang_str: str = "English"):
        var_name = tk.StringVar(value=str(name or ""))
        langs = [l.strip() for l in lang_str.split(",") if l.strip()]
        
        lang_vars = []
        for c in range(self.num_lang_columns):
            val = langs[c] if c < len(langs) else "English"
            lang_vars.append(tk.StringVar(value=val))

        row_dict = {
            "id": int(did),
            "var_name": var_name,
            "lang_vars": lang_vars,
        }
        self.device_rows.append(row_dict)

    def _renumber_device_rows(self):
        for idx, row in enumerate(self.device_rows):
            new_id = idx + 1
            row["id"] = int(new_id)
            widgets = row.get("widgets") or ()
            if len(widgets) >= 1:
                lbl_id = widgets[0]
                try:
                    lbl_id.configure(text=f"D{int(new_id)}")
                except Exception:
                    pass

    def _sync_devices_to_ips(self, initial_load=False):
        ips = list(self.ip_listbox.get(0, tk.END))
        saved_langs = self._settings.get("saved_device_languages", [])
        
        if initial_load:
            max_langs = 1
            for s_lang in saved_langs:
                parts = [p.strip() for p in s_lang.split(",") if p.strip()]
                if len(parts) > max_langs:
                    max_langs = len(parts)
            self.num_lang_columns = max_langs
        elif not hasattr(self, "num_lang_columns"):
            self.num_lang_columns = 1
            
        names_dict = {}
        langs_dict = {}
        if initial_load:
            try:
                cfgp = Path(__file__).resolve().parents[1] / "configs" / "device_profiles.json"
                if cfgp.exists():
                    with open(cfgp, "r", encoding="utf-8") as f:
                        obj = json.load(f) or {}
                    for d in obj.get("devices", []):
                        names_dict[int(d.get("id"))] = d.get("name", "")
                        langs_dict[int(d.get("id"))] = d.get("language", "English")
            except Exception: pass

        while len(self.device_rows) > len(ips):
            self.device_rows.pop()

        while len(self.device_rows) < len(ips):
            next_id = len(self.device_rows) + 1
            name = names_dict.get(next_id, f"Device {next_id}") if initial_load else f"Device {next_id}"
            
            if initial_load and len(saved_langs) >= next_id:
                lang = saved_langs[next_id - 1]
            else:
                lang = langs_dict.get(next_id, "English") if initial_load else "English"
                
            self._add_device_row(next_id, name, lang)

        self._regrid_device_rows()
        self._renumber_device_rows()

    def browse_excel(self):
        p = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel Files", "*.xlsm *.xlsx *.xls"), ("All Files", "*.*")],
        )
        if p:
            self.excel_var.set(p)
            self._index_to_tag_cache = {}  
            try:
                self.refresh_languages()
                self.refresh_tags()
            except Exception: pass

    def browse_model(self):
        p = filedialog.askopenfilename(
            title="Select YOLO model weights",
            filetypes=[("PyTorch Weights", "*.pt"), ("All Files", "*.*")],
        )
        if p: self.model_path_var.set(_resolve_path(p))

    def browse_log(self):
        d = filedialog.askdirectory(title="Select Log Folder")
        if d:
            normalized_path = str(Path(d).resolve())
            self.log_path_var.set(normalized_path)

    def _append(self, s: str, tags=None):
        if tags:
            self.output.insert(tk.END, s, tags)
        else:
            self.output.insert(tk.END, s)
        self.output.see(tk.END)

    def _append_cmd(self, s: str):
        self.output.insert(tk.END, s, ("cmd",))
        self.output.see(tk.END)

    def filter_tree(self, filter_val):
        self.current_filter = filter_val
        self._update_summary_ui(filter_val)

    def _update_summary_ui(self, filter_val="ALL"):
        for item in self.tree.get_children():
            self.tree.delete(item)

        p, f, w, s = 0, 0, 0, 0
        for d in self.all_results_data:
            v = d.get("verdict", "")
            if v == "PASS": p += 1
            elif v == "FAIL": f += 1
            elif v == "WARN": w += 1
            elif v == "SKIP": s += 1

            if filter_val == "ALL" or filter_val == v:
                self.tree.insert("", tk.END, values=(
                    d.get("device"), d.get("index"), d.get("tag"),
                    d.get("expected"), d.get("actual"), v, d.get("error")
                ), tags=(v,))

        tot = len(self.all_results_data)
        
        total_elapsed = self._settings.get("accumulated_time", 0.0)
        # --- FIX: Don't tick the clock if ADB is running ---
        if hasattr(self, 'batch_start_time') and getattr(self, "_commg_is_active_run", False) and not getattr(self, "_is_adb_running", False):
            total_elapsed += (time.time() - self.batch_start_time)

        if total_elapsed > 60:
            mins = int(total_elapsed // 60)
            secs = total_elapsed % 60
            time_taken_str = f"{mins}m {secs:.1f}s"
        else:
            time_taken_str = f"{total_elapsed:.1f}s"
                
        self.lbl_stats.config(text=f"Total: {tot} | PASS: {p} | FAIL: {f} | WARN: {w} | SKIP: {s} | Time: {time_taken_str}")
    def clear_summary_data(self):
        self.all_results_data.clear()
        self._update_summary_ui("ALL")

    def on_tree_double_click(self, event):
        item = self.tree.selection()
        if not item: return
        item = item[0]
        values = self.tree.item(item, "values")
        if not values: return

        top = tk.Toplevel(self.root)
        top.title(f"Details - {values[0]}")
        top.geometry("600x400")

        txt = tk.Text(top, wrap=tk.WORD, font=("Arial", 11), padx=10, pady=10)
        txt.pack(fill=tk.BOTH, expand=True)

        content = f"DEVICE: {values[0]}\n"
        content += f"INDEX: {values[1]}\n"
        content += f"TAG: {values[2]}\n"
        content += f"VERDICT: {values[5]}\n"
        content += "-" * 50 + "\n"
        content += f"EXPECTED:\n{values[3]}\n"
        content += "-" * 50 + "\n"
        content += f"ACTUAL:\n{values[4]}\n"
        if values[6] and values[6].strip():
            content += "-" * 50 + "\n"
            content += f"ERROR DETAILS:\n{values[6]}\n"

        txt.insert(tk.END, content)
        txt.config(state=tk.DISABLED)
        
    def _load_previous_results(self, excel_path):
        """Restores the UI summary table and calculates EXACT time from timestamps."""
        if not os.path.exists(excel_path): return
        try:
            from openpyxl import load_workbook
            from datetime import datetime
            import time
            
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            
            headers = []
            for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c).strip().lower() if c else "" for c in r]
                break
                
            if not headers: return
            
            def get_idx(name, fallback):
                return headers.index(name) if name in headers else fallback
                
            col_map = {
                "timestamp": get_idx("timestamp", 0),
                "device": get_idx("device", 1),
                "region": get_idx("region", 2),
                "language": get_idx("language", 3),
                "command": get_idx("command", 4),
                "display_style": get_idx("display style", 5),
                "index": get_idx("index", 6),
                "tag": get_idx("tag", 7),
                "expected": get_idx("expected (local)", 8),
                "actual": get_idx("actual detected", 9),
                "verdict": get_idx("verdict", 11),
                "error": get_idx("error message", 12),
            }
            
            self.all_results_data = []
            timestamps = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) <= col_map["verdict"]: continue
                v = str(row[col_map["verdict"]] or "").strip().upper()
                if v not in ["PASS", "FAIL", "WARN", "SKIP"]: continue 
                
                ts_val = str(row[col_map["timestamp"]] or "").strip()
                if ts_val and "SUMMARY" not in ts_val.upper():
                    timestamps.append(ts_val)
                
                # --- Perfectly restores both the UI visuals and the hidden backend data ---
                self.all_results_data.append({
                    "device": str(row[col_map["device"]] or ""),
                    "region": str(row[col_map["region"]] or ""),
                    "language": str(row[col_map["language"]] or ""),
                    "command": str(row[col_map["command"]] or ""),
                    "display_style": str(row[col_map["display_style"]] or ""),
                    "index": str(row[col_map["index"]] or ""),
                    "tag": str(row[col_map["tag"]] or ""),
                    "expected": str(row[col_map["expected"]] or ""),
                    "actual": str(row[col_map["actual"]] or ""),
                    "verdict": v,
                    "error": str(row[col_map["error"]] or "")
                })
            wb.close()
            
            parsed_dates = []
            for ts in timestamps:
                try: parsed_dates.append(datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"))
                except Exception: pass
            
            parsed_dates.sort()
            true_avg_sec = 36.75 
            if len(parsed_dates) > 1:
                diffs = [(parsed_dates[i] - parsed_dates[i-1]).total_seconds() for i in range(1, len(parsed_dates))]
                normal_diffs = [d for d in diffs if d <= 90.0] 
                if normal_diffs: true_avg_sec = sum(normal_diffs) / len(normal_diffs)

            exact_elapsed = true_avg_sec * len(self.all_results_data)
            self._settings["accumulated_time"] = exact_elapsed
            self.batch_start_time = time.time() 
            
            # --- This single line takes all the restored data and visually draws it in the GUI table! ---
            self._update_summary_ui(self.current_filter)
            
            self.q.put(f"[GUI] Successfully restored {len(self.all_results_data)} past results into the UI!\n", ("pass",))
        except Exception as e:
            self.q.put(f"[GUI ERROR] Failed to load previous results into UI: {e}\n", ("error",))

    def _append_line_with_result_color(self, line: str):
        stripped = (line or "").strip()
        
        if "[GUI_RESULT]" in stripped:
            try:
                json_str = stripped.split("[GUI_RESULT]")[1].strip()
                data = json.loads(json_str)
                self.all_results_data.append(data)
                self._update_summary_ui(self.current_filter)
            except Exception: 
                pass
            return

        low = stripped.lower()

        if stripped.startswith("Expected (") and "):" in stripped:
            try: self.last_expected = stripped.split("):", 1)[1].strip()
            except Exception: pass
        elif stripped == "Expected (normalized):":
            self._pending_expected_norm = True
        elif self._pending_expected_norm and stripped and stripped not in ["PASS", "FAIL", "WARN"]:
            try: self.last_expected = stripped
            except Exception: pass
            self._pending_expected_norm = False

        is_error = False
        if stripped.startswith("Traceback"): is_error = True
        elif "[error]" in low or "[gui error]" in low: is_error = True
        elif low.startswith("error:") or low.startswith("exception"): is_error = True
        elif "typeerror" in low or "valueerror" in low or "runtimeerror" in low: is_error = True
        elif "❌" in stripped: is_error = True

        if stripped == "PASS":
            self.last_result = "PASS"
            self.output.insert(tk.END, line, ("pass",))
        elif stripped == "FAIL":
            self.last_result = "FAIL"
            self.output.insert(tk.END, line, ("fail",))
            try:
                if (self.last_expected or "").strip():
                    exp_line = f"Expected: {self.last_expected}\n"
                    self.output.insert(tk.END, exp_line)
            except Exception: pass
        elif stripped == "WARN":
            self.last_result = "WARN"
            self.output.insert(tk.END, line, ("warn",))
        elif is_error:
            self.output.insert(tk.END, line, ("error",))
        elif "[CommG]" in stripped or "[CMD]" in stripped or "[PuTTY]" in stripped:
            self.output.insert(tk.END, line, ("commg",))
        else:
            self.output.insert(tk.END, line)
            
        self.output.see(tk.END)
        
        is_gui_msg = stripped.startswith("[CommG]") or stripped.startswith("[CMD]") or stripped.startswith("[PuTTY]") or stripped.startswith("[GUI")
        if "RADIO STRING VERIFICATION - DETECTED" in stripped:
            self._is_recording_log = True
            try:
                if self._log_fp is not None:
                    self._log_fp.write("=" * 70 + "\n")
            except Exception: pass

        try:
            if self._log_fp is not None and getattr(self, "_is_recording_log", False) and not is_gui_msg:
                self._log_fp.write(line)
                if not line.endswith("\n"):
                    self._log_fp.write("\n")
                self._log_fp.flush()
        except Exception: pass

    def _current_settings(self) -> dict:
        camera_name = (self.camera_id_var.get() or "").strip()
        camera_id = getattr(self, "_camera_map", {}).get(camera_name, camera_name)
        
        saved_langs = [",".join(self._get_device_langs(r)) for r in self.device_rows]
        
        return {
            "excel": (self.excel_var.get() or "").strip(),
            "tag": (self.tag_var.get() or "").strip(),
            "index": (self.index_var.get() or "").strip(),
            "model_path": (self.model_path_var.get() or "").strip(),
            "camera_id": camera_id,
            "save_log": bool(self.save_log_var.get()),
            "enable_rolling": bool(self.enable_rolling_var.get()),     
            "enable_truncation": bool(self.enable_truncation_var.get()), 
            "enable_retries": bool(self.enable_retries_var.get()),
            "retry_count": (self.retry_count_var.get() or "2").strip(),
            "log_path": (self.log_path_var.get() or "").strip(),
            "drive_folder": (self.drive_folder_var.get() or "Walkie_Logs").strip(), 
            "excel_filename": (self.excel_filename_var.get() or "Batch_Summary_Report.xlsx").strip(), 
            "integration_enable": bool(self.integration_enable_var.get()),
            "integration_type": (self.integration_type_var.get() or "CommG"),
            "telnet_port": (self.telnet_port_var.get() or "23"),
            "commg_enable": bool(self.integration_enable_var.get()), 
            "commg_ips": list(self.ip_listbox.get(0, tk.END)),
            "commg_mode": (self.commg_mode_var.get() or "Batch"),
            "commg_custom_cmd": (self.commg_custom_cmd_var.get() or "").strip(),
            "commg_batch_file": (self.commg_batch_file_var.get() or "").strip(),
            "putty_path": self._settings.get("putty_path", ""),
            "saved_device_languages": saved_langs,
            "accumulated_time": self._settings.get("accumulated_time", 0.0)
        }

    def _persist_settings(self):
        self._settings = self._current_settings()
        _save_settings(self._settings)

    def _on_close(self, skip_prompt=False):
        if not skip_prompt:
            if not messagebox.askyesno("Confirm Exit", "Are you sure you want to close the application?"):
                return
        
        try: 
            self.stop(skip_prompt=True)
        except Exception: 
            pass

        try: self._persist_settings()
        except Exception: pass
        try:
            if self._log_fp is not None:
                self._log_fp.close()
                self._log_fp = None
        except Exception: pass
        try: self._log_session_dir = None
        except Exception: pass
        
        self.root.destroy()

    def refresh_cameras(self):
        try:
            self.btn_run.configure(state=tk.DISABLED)
        except Exception: pass
        self.root.update()

        cam_data = _probe_camera_ids(max_id=8)
        
        if not cam_data: 
            cam_data = [("0", "Default Camera"), ("1", "External Camera"), ("2", "Virtual Camera")]
            
        self._camera_map = {}
        display_names = []
        
        for cid, cname in cam_data:
            if cname in self._camera_map:
                cname = f"{cname} (ID: {cid})"
            self._camera_map[cname] = cid
            display_names.append(cname)
            
        try:
            self.camera_combo["values"] = display_names
            cur = (self.camera_id_var.get() or "").strip()
            
            if cur in display_names:
                self.camera_id_var.set(cur)
            elif display_names:
                self.camera_id_var.set(display_names[0])
                
            messagebox.showinfo("Camera Search", f"Successfully completed - Found {len(display_names)} camera(s)!")
            
        except Exception: 
            pass
        finally:
            try:
                self.btn_run.configure(state=tk.NORMAL)
            except Exception: pass

    def refresh_languages(self):
        excel = (self.excel_var.get() or "").strip()
        opts = []
        if excel:
            try: 
                opts1 = _language_options_from_excel(excel, "apac")
                opts2 = _language_options_from_excel(excel, "emea")
                opts3 = _language_options_from_excel(excel, "lacr")
                opts = list(set(opts1 + opts2 + opts3))
            except Exception:
                pass
        if not opts:
            opts = ["Japanese", "Korean", "Simplified Chinese", "Traditional Chinese", "French", "Spanish", "German", "Italian", "Polish", "Russian", "Turkish", "Arabic", "Hungarian", "Hebrew", "Czech", "Portuguese", "English"]

        # --- FIX: ADD "SKIP" TO THE TOP OF THE LIST ---
        self.available_languages = ["SKIP"] + sorted(opts)
        
        for r in self.device_rows:
            for v in r.get("lang_vars", []):
                if not v.get() and len(self.available_languages) > 1:
                    v.set(self.available_languages[1]) # Default to the first real language, not SKIP
        if hasattr(self, "devices_grid_frame"):
            self._regrid_device_rows()
                
    def _open_lang_selector(self, row_dict):
        top = tk.Toplevel(self.root)
        top.title(f"Languages - D{row_dict['id']}")
        top.transient(self.root)
        top.grab_set()
        
        x = self.root.winfo_x() + 100
        y = self.root.winfo_y() + 100
        top.geometry(f"280x350+{x}+{y}")
        
        tk.Label(top, text="Select languages\n(Disabled items are in use by other devices)", pady=10, font=("Arial", 9, "bold")).pack()
        
        other_selected = set()
        for r in self.device_rows:
            if r is not row_dict:
                langs = self._get_device_langs(r)
                other_selected.update(langs)
                
        current_selected = set([l.strip() for l in row_dict["var_lang"].get().split(",") if l.strip()])
        
        check_vars = {}
        frame_checks = tk.Frame(top, padx=10, pady=5)
        frame_checks.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(frame_checks, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_checks, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for lang in self.available_languages:
            var = tk.BooleanVar(value=(lang in current_selected))
            check_vars[lang] = var
            cb = tk.Checkbutton(scrollable_frame, text=lang, variable=var, font=("Arial", 10))
            
            if lang in other_selected:
                cb.configure(state=tk.DISABLED)
                
            cb.pack(anchor="w", pady=2)
            
        def on_ok():
            selected = [lang for lang, var in check_vars.items() if var.get()]
            row_dict["var_lang"].set(", ".join(selected))
            top.destroy()
            
        btn_frame = tk.Frame(top, pady=10)
        btn_frame.pack()
        tk.Button(btn_frame, text="OK", command=on_ok, width=12, bg="#90EE90", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="Cancel", command=top.destroy, width=12, bg="#FFCCCB", font=("Arial", 9, "bold")).pack(side=tk.RIGHT, padx=10)

    def refresh_tags(self):
        excel = (self.excel_var.get() or "").strip()
        try: tags = _tag_options_from_excel(excel)
        except Exception: tags = []
        try: self._all_tags = tags
        except Exception: pass
        try: self.tag_combo["values"] = tags
        except Exception: pass

    def _on_tag_typed(self, _evt=None):
        try: all_tags = list(getattr(self, "_all_tags", []) or [])
        except Exception: all_tags = []

        if not all_tags: return

        try: typed_raw = self.tag_combo.get()
        except Exception: typed_raw = self.tag_var.get()
        typed = (typed_raw or "").strip().lower()
        if not typed: filtered = all_tags
        else: filtered = [t for t in all_tags if typed in str(t).lower()]

        try: self.tag_combo["values"] = filtered
        except Exception: pass

    def _unpost_tag_dropdown(self):
        try: self.tag_combo.tk.call("ttk::combobox::Unpost")
        except Exception: pass

    def _on_tag_escape(self, _evt=None):
        self._unpost_tag_dropdown()

    def _on_tag_focus_out(self, _evt=None):
        self._unpost_tag_dropdown()

    def _on_tag_down(self, _evt=None):
        try:
            if not list(self.tag_combo["values"] or []): self.refresh_tags()
        except Exception: pass
        try: self._on_tag_typed()
        except Exception: pass
        try: self.tag_combo.tk.call("ttk::combobox::Post", str(self.tag_combo))
        except Exception: pass

    def clear(self):
        self.output.delete("1.0", tk.END)
        self.last_result = ""
        self.status_var.set("Idle")
        try: self.status_label.configure(fg="black")
        except Exception: pass

    def close(self):
        self._on_close()

    def _set_running(self, running: bool):
        try:
            if running:
                self.btn_run.configure(state=tk.DISABLED)
                self.btn_stop.configure(state=tk.NORMAL)
            else:
                self.btn_run.configure(state=tk.NORMAL)
                self.btn_stop.configure(state=tk.DISABLED)
        except Exception: pass

    def _build_cmd(self):
        excel = self.excel_var.get().strip()
        if not excel: raise ValueError("Excel path is required")

        tag = self.tag_var.get().strip()
        idx = self.index_var.get().strip()

        if not tag and not idx: raise ValueError("Provide either String Tag or Index")

        region = "Multiple"
        langs = []
        for r in self.device_rows:
            if "current_active_lang" in r and r["current_active_lang"]:
                langs.append(r["current_active_lang"])
            else:
                langs.append(",".join(self._get_device_langs(r)) or "English")
        language = ",".join(langs) if langs else "English"

        script_path = Path(__file__).resolve().parent / "verify_string.py"
        cmd = [_python_exe(), str(script_path), "--excel", excel, "--region", region, "--language", language]

        if tag: cmd += ["--tag", tag]
        if idx: cmd += ["--index", idx]
        
        if self.enable_rolling_var.get():
            cmd += ["--enable-rolling"]
        if self.enable_truncation_var.get():
            cmd += ["--enable-truncation"]
        if self.enable_retries_var.get():
            cmd += ["--enable-retries"]
            retry_val = self.retry_count_var.get().strip()
            if retry_val.isdigit():
                cmd += ["--max-retries", retry_val]

        if getattr(self, "_commg_is_active_run", False) and hasattr(self, "_current_batch_cmds"):
            cmd += ["--command", ",".join(self._current_batch_cmds)]
        elif self.commg_mode_var.get() == "Single" and self.integration_enable_var.get():
            cmd += ["--command", self.commg_custom_cmd_var.get().strip()]

        model_path = _resolve_path(self.model_path_var.get())
        try: self.model_path_var.set(model_path)
        except Exception: pass
        if model_path: cmd += ["--model-path", model_path]

        camera_name = self.camera_id_var.get().strip()
        if camera_name:
            camera_id = getattr(self, "_camera_map", {}).get(camera_name, camera_name)
            cmd += ["--camera-id", str(camera_id)]

        return cmd

    def _run_subprocess(self, cmd, *, is_verification: bool = True):
        if hasattr(self, "proc_thread") and self.proc_thread and self.proc_thread.is_alive():
            messagebox.showinfo("Running", "A process is already running. Wait for it to finish.")
            return

        try: self._last_run_is_verification = bool(is_verification)
        except Exception: self._last_run_is_verification = True

        self.last_result = ""
        self.last_expected = ""   
        self.last_actual = ""     
        self.last_error_msg = ""  
        self._is_recording_log = False 
        
        self._process_gen = getattr(self, "_process_gen", 0) + 1
        current_gen = self._process_gen
        
        self.status_var.set("Running Verify...")
        try: self.status_label.configure(fg="black")
        except Exception: pass
        self._set_running(True)

        try:
            if self._log_fp is not None: self._log_fp.close()
        except Exception: pass
        self._log_fp = None

        if self.save_log_var.get() and bool(is_verification):
            if getattr(self, "_commg_is_active_run", False) and getattr(self, "_batch_log_dir", None):
                self._log_session_dir = self._batch_log_dir
            else:
                d = (self.log_path_var.get() or "").strip()
                if not d:
                    d = str(Path.cwd())
                    self.log_path_var.set(d)

                try:
                    dp = Path(d)
                    if dp.suffix.lower() in [".log", ".txt"]: dp = dp.parent
                    d = str(dp)
                except Exception: pass

                sess = time.strftime("%Y%m%d_%H%M%S")
                self._log_session_dir = str(Path(d) / f"verified_{sess}")
                try: Path(self._log_session_dir).mkdir(parents=True, exist_ok=True)
                except Exception: self._log_session_dir = d

            ts = time.strftime("%Y%m%d_%H%M%S")
            if getattr(self, "_commg_is_active_run", False):
                name = "batch_execution_full.log"
            else:
                safe_tag = (self.tag_var.get() or "").strip()
                safe_tag = "".join([c for c in safe_tag if c.isalnum() or c in ["_", "-"]])[:40]
                name = f"verify_string_{ts}.log" if not safe_tag else f"verify_string_{safe_tag}_{ts}.log"
                
            p = str(Path(self._log_session_dir) / name)
            try:
                Path(p).parent.mkdir(parents=True, exist_ok=True)
                self._log_fp = open(p, "a", encoding="utf-8")
            except Exception: self._log_fp = None
            self._current_log_file_name = name
            self._current_log_full_path = p

        try:
            self.output.insert(tk.END, "$ " + " ".join(cmd) + "\n\n", ("cmd",))
            self.output.see(tk.END)
        except Exception: pass

        try:
            if self.save_log_var.get() and bool(is_verification) and self._log_session_dir:
                cmd = list(cmd) + ["--save-roi-dir", self._log_session_dir]
                
                filename = self.excel_filename_var.get().strip() or "Batch_Summary_Report.xlsx"
                if not filename.lower().endswith(".xlsx"): filename += ".xlsx"
                
                excel_path = str(Path(self._log_session_dir) / filename)
                cmd = list(cmd) + ["--summary-excel", excel_path]
        except Exception as e:
            self.q.put(f"[GUI ERROR] Failed to initialize log paths: {e}\n")
            self._log_fp = None

        env = os.environ.copy()
        try:
            devices = []
            skipped_indices = getattr(self, "_current_skipped_indices", [])
            
            for i, r in enumerate(self.device_rows):
                did = int(r.get("id"))
                nm = str(r.get("var_name").get() if r.get("var_name") else "").strip()
                
                # --- FIX: Fetch the exact language currently active in the column ---
                if "current_active_lang" in r:
                    lang = r["current_active_lang"]
                else:
                    lang = ",".join(self._get_device_langs(r)) or "English"
                
                if i in skipped_indices:
                    cmd_str = self._current_batch_cmds[0] if hasattr(self, "_current_batch_cmds") and self._current_batch_cmds else "-"
                    idx_str = self.index_var.get().split(",")[0] if self.index_var.get() else ""
                    tag_str = self.tag_var.get().split(",")[0] if self.tag_var.get() else ""
                    
                    payload = {
                        "device": nm, 
                        "command": cmd_str, 
                        "display_style": "-", 
                        "index": idx_str, 
                        "tag": tag_str, 
                        "expected": "-", 
                        "actual": "-", 
                        "verdict": "SKIP", 
                        "error": "Device Offline / Ping Failed",
                        "language": lang
                    }
                    self.all_results_data.append(payload)
                    self._append(f"\n[GUI] {nm} was skipped (Device Offline / Ping Failed).\n", ("warn",))
                else:
                    # If language is SKIP_DEVICE, YOLO tracks it but doesn't waste tokens verifying it!
                    devices.append({"id": did, "name": nm, "language": lang})
            
            self._current_skipped_indices = []
            self._update_summary_ui(self.current_filter)
            
            env["WALKIE_DEVICE_PROFILES_JSON"] = json.dumps({"devices": devices}, ensure_ascii=False)
            cfgp = Path(__file__).resolve().parents[1] / "configs" / "device_profiles.json"
            cfgp.parent.mkdir(parents=True, exist_ok=True)
            with open(cfgp, "w", encoding="utf-8") as f:
                json.dump({"devices": devices}, f, indent=2, ensure_ascii=False)
        except Exception: pass

        env["PYTHONUNBUFFERED"] = "1"

        def _stream_process():
            try:
                creationflags = subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
                self.current_process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    env=env,
                    bufsize=1,
                    creationflags=creationflags
                )
                
                for line in iter(self.current_process.stdout.readline, b''):
                    decoded = line.decode("utf-8", errors="replace")
                    self.q.put(decoded)
                    
            except Exception as e:
                self.q.put(f"[GUI ERROR] Failed to run AI subprocess: {e}\n")
            finally:
                if hasattr(self, 'current_process') and self.current_process:
                    try:
                        self.current_process.stdout.close()
                        self.current_process.wait()
                        rc = self.current_process.returncode
                    except Exception:
                        rc = -1
                else:
                    rc = -1
                    
                self.q.put((_EVT_FINISHED, rc, current_gen))

        self.proc_thread = threading.Thread(target=_stream_process, daemon=True)
        self.proc_thread.start()
        
    def run_camera_test(self):
        env = os.environ.copy()
        camera_name = self.camera_id_var.get().strip()
        camera_id = 0
        if camera_name:
            camera_id = getattr(self, "_camera_map", {}).get(camera_name, camera_name)
            env["WALKIE_CAMERA_ID"] = str(camera_id)

        try:
            devices = []
            for r in self.device_rows:
                did = int(r.get("id"))
                nm = str(r.get("var_name").get() if r.get("var_name") else "").strip()
                lang = ",".join(self._get_device_langs(r)) or "English"
                devices.append({"id": did, "name": nm, "language": lang})
            env["WALKIE_DEVICE_PROFILES_JSON"] = json.dumps({"devices": devices}, ensure_ascii=False)
            cfgp = Path(__file__).resolve().parents[1] / "configs" / "device_profiles.json"
            cfgp.parent.mkdir(parents=True, exist_ok=True)
            with open(cfgp, "w", encoding="utf-8") as f:
                json.dump({"devices": devices}, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.q.put(f"[GUI WARNING] Failed to pass device profiles: {e}\n", ("warn",))
            
        self.q.put(f"\n[INFO] Launching Live Camera Preview... (Camera: {camera_name})\n")
        self.q.put("[INFO] 💡 LEFT-CLICK a box to select it, then LEFT-CLICK another box to SWAP their device names!\n")
        
        def _cam_worker():
            old_argv = sys.argv
            model_p = self.model_path_var.get().strip()
            
            sys.argv = ['verify_string', '--preview', '--excel', 'dummy.xlsx', '--region', 'APAC', '--language', 'dummy']
            sys.argv.extend(['--camera-id', str(camera_id)])
            
            if model_p:
                sys.argv.extend(['--model-path', model_p])
            
            old_env = os.environ.copy()
            os.environ.update(env)
            
            q_stream = _QueueStream(self.q)
            
            try:
                with redirect_stdout(q_stream), redirect_stderr(q_stream):
                    runpy.run_module('verify_string', run_name="__main__")
            except SystemExit:
                pass
            except Exception as e:
                self.q.put(f"[GUI ERROR] Failed to launch camera test: {e}\n", ("error",))
            finally:
                sys.argv = old_argv
                os.environ.clear()
                os.environ.update(old_env)

        threading.Thread(target=_cam_worker, daemon=True).start()

    def init_genai(self):
        script_path = Path(__file__).resolve().parent / "init_genai_session.py"
        cmd = [_python_exe(), str(script_path)]
        self._run_subprocess(cmd, is_verification=False)

    def init_and_run(self):
        self._needs_queue_reset = False
        is_resume_mode = False
        
        has_paused = getattr(self, "_is_paused", False) and getattr(self, "_commg_pending_queue", [])
        
        if has_paused:
            choice1 = messagebox.askyesnocancel(
                "Paused State Detected",
                "A paused automation state was found.\n\n"
                "• YES: Resume EXACTLY where it left off.\n"
                "• NO: Discard paused state and configure a new start.\n"
                "• CANCEL: Abort."
            )
            if choice1 is None:
                return
            elif choice1 is True:
                self._is_paused = False
                self._commg_is_active_run = True
                self._auto_start_verify = True
                self._is_resuming_batch = True 
                self._summary_written = False
                is_resume_mode = True
                
                # --- NEW: Explicitly lock the vault door on the queue! ---
                self._protect_resume_queue = True 
                # ---------------------------------------------------------
                
                self._setup_logs_and_prefetch(is_resume_mode)
                return
            else:
                self._is_paused = False
                self._commg_pending_queue = []
                self.clear_summary_data()
        else:
            self.clear_summary_data()
            
        choice2 = messagebox.askyesnocancel(
            "Batch Configuration",
            "How do you want to start the batch?\n\n"
            "• YES: Resume from a Custom Index & Language.\n"
            "• NO: Start Fresh from the beginning.\n"
            "• CANCEL: Abort."
        )
        
        if choice2 is None:
            return
        elif choice2 is True:
            from tkinter import simpledialog
            target_idx = simpledialog.askstring("Resume", "Enter the Index to resume from:")
            if not target_idx: return
            self._resume_target_idx = target_idx.strip()
            is_resume_mode = True
        else:
            self._resume_target_idx = None
            is_resume_mode = False
            self._settings["accumulated_time"] = 0.0
            self._persist_settings()
            
        self.btn_run.configure(state=tk.DISABLED)
        self.batch_start_time = time.time()
        self._summary_written = False
        self._full_batch_commands = [] 
        
        max_l = 1
        for r in self.device_rows:
            try:
                langs = self._get_device_langs(r)
                if len(langs) > max_l: max_l = len(langs)
            except Exception: pass
        self._max_lang_idx = max_l
        
        if is_resume_mode and self._max_lang_idx > 1:
            from tkinter import simpledialog
            lang_iter = simpledialog.askinteger(
                "Language Batch", 
                f"You have {self._max_lang_idx} language iterations queued.\n\n"
                f"Which language iteration number (1 to {self._max_lang_idx}) do you want to start from?", 
                initialvalue=1, minvalue=1, maxvalue=self._max_lang_idx
            )
            if not lang_iter:
                self.btn_run.configure(state=tk.NORMAL)
                return
            self._current_lang_idx = lang_iter - 1
        else:
            self._current_lang_idx = 0
            
        self._starting_lang_idx = self._current_lang_idx
        
        if self.integration_enable_var.get() and HAS_PYWINAUTO:
            if not list(self.ip_listbox.get(0, tk.END)):
                messagebox.showwarning("No IPs", "Please add at least one IP address to the list before starting!")
                self._set_running(False) 
                return

            if self.commg_mode_var.get() == "Single":
                c = self.commg_custom_cmd_var.get().strip()
                if not c:
                    messagebox.showerror("Error", "Please enter a Custom Command.")
                    self._set_running(False) 
                    return
                self._commg_pending_queue = [c]
                self._full_batch_commands = list(self._commg_pending_queue) 
            else:
                bf = self.commg_batch_file_var.get().strip()
                if not bf or not os.path.exists(bf):
                    messagebox.showerror("Error", "Please select a valid Batch File.")
                    self._set_running(False) 
                    return
                try:
                    import pandas as pd
                    if bf.lower().endswith('.csv'): df = pd.read_csv(bf, header=None)
                    else: df = pd.read_excel(bf, header=None)
                    
                    cmds = df.iloc[:, 0].dropna().astype(str).tolist()
                    self._commg_pending_queue = [c.strip() for c in cmds if c.strip()]
                    
                    if len(df.columns) > 1:
                        tags = df.iloc[:, 1].fillna("").astype(str).tolist()
                        self._commg_pending_queue = list(zip(self._commg_pending_queue, tags))
                    
                    self._full_batch_commands = list(self._commg_pending_queue)

                    if not self._full_batch_commands:
                        messagebox.showwarning("Warning", "No commands found in the batch file.")
                        self._set_running(False) 
                        return
                        
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to load Batch File: {e}")
                    self._set_running(False) 
                    return
                    
            if not getattr(self, "_commg_pending_queue", []):
                messagebox.showwarning("Warning", "No commands found in the batch file.")
                self._set_running(False) 
                return
            
            if getattr(self, "_resume_target_idx", None):
                target_idx = self._resume_target_idx
                found_pos = -1
                for i, item in enumerate(self._full_batch_commands):
                    cmd = item[0] if isinstance(item, tuple) else item
                    idx_val = cmd.split(":")[-1].strip() if ":" in cmd else cmd.strip()
                    try: clean_idx = str(int(idx_val)) if idx_val.isdigit() else idx_val
                    except Exception: clean_idx = idx_val
                    try: clean_target = str(int(target_idx)) if target_idx.isdigit() else target_idx
                    except Exception: clean_target = target_idx

                    if clean_idx == clean_target:
                        found_pos = i
                        break

                if found_pos != -1:
                    self._commg_pending_queue = self._full_batch_commands[found_pos:]
                    self.q.put(f"[GUI] Resuming Language Iteration {self._current_lang_idx + 1} from index {target_idx}. Skipped {found_pos} items.\n", ("warn",))
                else:
                    messagebox.showerror("Not Found", f"Index '{target_idx}' not found in the batch.")
                    self._set_running(False)
                    self._resume_target_idx = None
                    return
                self._resume_target_idx = None
                
            # --- FIX: Lock the queue when using Custom Index Resume! ---
            self._is_resuming_batch = True
            self._protect_resume_queue = True
            # -----------------------------------------------------------
            
            self._commg_is_active_run = True
            self._auto_start_verify = True
            
            self._setup_logs_and_prefetch(is_resume_mode)
        else:
            self.notebook.select(self.tab_summary)
            self._commg_pending_queue = []
            self._commg_is_active_run = False
            self._batch_log_dir = None
            self._setup_logs_and_prefetch(is_resume_mode)

    def _setup_logs_and_prefetch(self, is_resume_mode):
        self.notebook.select(self.tab_summary)
        
        if self._commg_is_active_run:
            tool = self.integration_type_var.get()
            self.q.put(f"[{tool}] Initialized batch with {len(self._commg_pending_queue)} items.\n", ("commg",))

        if self.save_log_var.get():
            d = (self.log_path_var.get() or "").strip()
            if not d:
                d = str(Path.cwd())
                self.log_path_var.set(d)
            
            # --- FIX 1: Don't create a new folder if we are just unpausing! ---
            if is_resume_mode and getattr(self, "_batch_log_dir", None) and Path(self._batch_log_dir).exists():
                pass # Keep using the exact same local folder
            else:
                sess = time.strftime("%Y%m%d_%H%M%S")
                self._batch_log_dir = str(Path(d) / f"batch_run_{sess}")
                try:
                    Path(self._batch_log_dir).mkdir(parents=True, exist_ok=True)
                except Exception:
                    self._batch_log_dir = d

            if is_resume_mode:
                import shutil
                drive_folder_name = self.drive_folder_var.get().strip() or "Walkie_Logs"
                filename = self.excel_filename_var.get().strip() or "Batch_Summary_Report.xlsx"
                if not filename.lower().endswith(".xlsx"): filename += ".xlsx"
                
                base_name = filename[:-5]
                
                target_dir = Path("G:/My Drive") / drive_folder_name / base_name
                if not target_dir.exists():
                    target_dir = Path("G:/Shared drives") / drive_folder_name / base_name
                
                # --- FIX: Protect against data wipes AND load previous data ---
                local_resume_path = Path(self._batch_log_dir) / filename
                if local_resume_path.exists():
                    self.q.put(f"[GUI] Resuming using existing local cache. No need to download from Drive.\n", ("pass",))
                    self._load_previous_results(local_resume_path) # <-- RESTORES UI DATA
                elif target_dir.exists():
                    synced = 0
                    for drive_file in target_dir.glob("*.xlsx"):
                        correct_name = drive_file.name.replace("_LIVE_UPDATE", "")
                        local_path = Path(self._batch_log_dir) / correct_name
                        try:
                            shutil.copy2(drive_file, local_path)
                            synced += 1
                        except Exception as e:
                            self.q.put(f"[GUI ERROR] Download blocked by Windows/Drive: {e}\n", ("error",))
                            
                    if synced > 0:
                        self.q.put(f"[GUI] Fetched {synced} existing Excel file(s) from Drive to resume.\n", ("pass",))
                        self._load_previous_results(local_resume_path) # <-- RESTORES UI DATA
                    else:
                        self.q.put(f"[GUI FATAL ERROR] Could not fetch old files! Stopping sequence to prevent wiping your Google Drive data.\n", ("error",))
                        self._set_running(False)
                        return
                else:
                     self.q.put(f"[GUI WARNING] No previous Drive folder found. Starting fresh.\n", ("warn",))
                     
            else:
                base_name = self.excel_filename_var.get().strip() or "Batch_Summary_Report"
                if base_name.lower().endswith(".xlsx"):
                    base_name = base_name[:-5]
                import re
                base_name = re.sub(r'_NewRun_\d{8}_\d{6}$', '', base_name)
                new_filename = f"{base_name}_NewRun_{sess}.xlsx"
                self.excel_filename_var.set(new_filename)
                self.q.put(f"[GUI] Normal Start: Using new Excel filename to avoid overwriting Drive: {new_filename}\n", ("pass",))

        self.status_var.set("Fetching Serials (MUX Mode)...")
        self.status_label.configure(fg="blue")
        self.btn_run.configure(state=tk.DISABLED)
        # ... (the rest of the function remains the same starting from safe_ips)
        
        safe_ips = list(self.ip_listbox.get(0, tk.END))
        try:
            safe_port = int(self.telnet_port_var.get().strip())
        except ValueError:
            safe_port = 23
            
        def prefetch_worker():
            ips = [ip.strip() for ip in safe_ips if ip.strip()]
            if not hasattr(self, "_serial_cache"):
                self._serial_cache = {}
                
            if ips:
                self.q.put("\n" + "="*60 + "\n")
                self.q.put("[Pre-Check] Grabbing Serial Numbers while in MUX Mode...\n")
                
            while True:
                needs_recovery = False
                all_cached = True
                
                for ip in ips:
                    if ip not in self._serial_cache:
                        ser = self.get_radio_serial(ip, port=safe_port)
                        if ser:
                            self._serial_cache[ip] = ser
                        else:
                            self.q.put(f"[Pre-Check Warning] {ip} timed out. It is likely stuck in AP Mode.\n")
                            needs_recovery = True
                            all_cached = False
                            
                if all_cached:
                    break 
                    
                if needs_recovery:
                    self.q.put("[Pre-Check] Attempting to auto-recover devices to MUX mode via ADB Intent...\n")
                    try:
                        import subprocess, time
                        out = subprocess.check_output(["adb", "devices"], text=True)
                        recovered_any = False
                        for line in out.splitlines():
                            if "\tdevice" in line:
                                dev_ser = line.split("\t")[0]
                                self.q.put(f"[ADB] Sending MUX intent to {dev_ser}...\n")
                                creationflags = getattr(subprocess, 'CREATE_NO_WINDOW', 0) if os.name == 'nt' else 0
                                subprocess.run(["adb", "-s", dev_ser, "shell", "am", "broadcast", "-a", "msi.factorymuxrouter.intent.action.BP_ROUTE_MUX", "-n", "com.motorolasolutions.factorymuxrouter/.MuxRouterBroadcastReceiver"], creationflags=creationflags)
                                recovered_any = True
                        
                        if recovered_any:
                            self.q.put("[Pre-Check] ⏳ Waiting 15 seconds for USB to reconnect in MUX Mode...\n")
                            time.sleep(15.0)
                            self.q.put("\n[Pre-Check] Retrying Serial Fetch...\n")
                            continue 
                        else:
                            self.q.put("[ADB Error] No devices found in AP mode to send the intent to.\n")
                    except Exception as e:
                        self.q.put(f"[ADB Error] Auto-recovery failed: {e}\n")

                self.q.put("[Pre-Check] Paused. Waiting for manual restart...\n")
                retry = messagebox.askretrycancel(
                    "Action Required: Manual MUX Check",
                    "The devices failed to return to MUX mode automatically.\n\n"
                    "1. Please MANUALLY SHUTDOWN and POWER ON the devices (Cold Boot).\n"
                    "2. Wait for them to fully turn on and reconnect to Windows.\n"
                    "3. Click 'Retry' below."
                )
                
                if retry:
                    self.q.put("\n[Pre-Check] Retrying Serial Fetch...\n")
                    continue 
                
                self.q.put("[Pre-Check] Aborted by user.\n")
                self.q.put((_EVT_COMMG_DONE, "ABORT", []))
                return

            self.q.put(("__PREFETCH_DONE__", None))
            
        self.init_genai()
        threading.Thread(target=prefetch_worker, daemon=True).start()

    def run(self):
        self.btn_run.configure(state=tk.DISABLED)
        if not getattr(self, '_commg_is_active_run', False) or not hasattr(self, 'batch_start_time'):
            self.batch_start_time = time.time()
            self._summary_written = False 
        
        try: cmd = self._build_cmd()
        except Exception as e:
            messagebox.showerror("Invalid input", str(e))
            self._commg_is_active_run = False
            self._set_running(False)
            return

        try: self._persist_settings()
        except Exception: pass

        self._run_subprocess(cmd, is_verification=True)

    def stop(self, skip_prompt=False):
        self._update_accumulated_time()
        
        if not skip_prompt:
            dialog = tk.Toplevel(self.root)
            dialog.title("Process Control")
            dialog.transient(self.root) 
            dialog.grab_set()           
            
            try:
                x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 175
                y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 70
                dialog.geometry(f"350x140+{x}+{y}")
            except Exception:
                dialog.geometry("350x140")
                
            tk.Label(dialog, text="Process is currently running.\nWhat would you like to do?", font=('Arial', 10)).pack(pady=15)
            
            action = {"result": "continue"}
            
            def on_continue():
                action["result"] = "continue"
                dialog.destroy()
                
            def on_stop():
                if messagebox.askyesno("Confirm Stop", "Are you sure to stop?", parent=dialog):
                    action["result"] = "stop"
                    dialog.destroy()
                    
            def on_retry():
                from tkinter import simpledialog
                import time
                import queue
                from pathlib import Path
                
                target_idx = simpledialog.askstring("Retry", "Enter the index number to retry from\n(Searches from BOTTOM first):", parent=dialog)
                if not target_idx:
                    return
                target_idx = target_idx.strip()

                found_pos = -1
                full_batch = getattr(self, "_full_batch_commands", [])
                
                for i in range(len(full_batch) - 1, -1, -1):
                    item = full_batch[i]
                    cmd = item[0] if isinstance(item, tuple) else item
                    idx_val = cmd.split(":")[-1].strip() if ":" in cmd else cmd.strip()
                    
                    try: clean_idx = str(int(idx_val)) if idx_val.isdigit() else idx_val
                    except Exception: clean_idx = idx_val
                    
                    try: clean_target = str(int(target_idx)) if target_idx.isdigit() else target_idx
                    except Exception: clean_target = target_idx
                    
                    if clean_idx == clean_target:
                        found_pos = i
                        break
                
                if found_pos != -1:
                    if hasattr(self, "current_process") and self.current_process:
                        try:
                            self._process_gen = getattr(self, "_process_gen", 0) + 1 
                            self.current_process.kill()
                            self._append("\n[GUI] Process interrupted for retry. Cleaning up trailing logs...\n", ("warn",))
                        except Exception:
                            pass
                    
                    while not self.q.empty():
                        try:
                            self.q.get_nowait()
                        except queue.Empty:
                            break

                    self._commg_pending_queue = full_batch[found_pos:]
                    self._current_batch_items = []
                    
                    ips = list(self.ip_listbox.get(0, tk.END))
                    num_ips = len(ips) if ips else 1
                    current_lang_idx = getattr(self, "_current_lang_idx", 0)
                    expected_valid_size = int((current_lang_idx * len(full_batch) * num_ips) + (found_pos * num_ips))
                    
                    self.all_results_data = self.all_results_data[:expected_valid_size]
                    self._update_summary_ui(self.current_filter)
                    
                    action["result"] = "retry"
                    dialog.destroy()
                else:
                    messagebox.showerror("Not Found", f"Index '{target_idx}' not found in the original batch.", parent=dialog)

            btn_frame = tk.Frame(dialog)
            btn_frame.pack(fill=tk.BOTH, expand=True)
            
            tk.Button(btn_frame, text="Continue", command=on_continue, width=8, bg="#90EE90", font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=15)
            tk.Button(btn_frame, text="Retry", command=on_retry, width=8, bg="#FFF3E0", font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=15)
            tk.Button(btn_frame, text="Stop", command=on_stop, width=8, bg="#FFCCCB", font=('Arial', 9, 'bold')).pack(side=tk.RIGHT, padx=15)
            
            dialog.protocol("WM_DELETE_WINDOW", on_continue)
            
            self.root.wait_window(dialog)
            
            if action["result"] == "continue":
                return
            elif action["result"] == "retry":
                if getattr(self, "active_sockets", []):
                    for s in self.active_sockets:
                        try:
                            s.sendall(b"STR_TEST:CLOSE\r\n")
                            time.sleep(0.2)
                            s.close()
                        except Exception: pass
                    self.active_sockets = []

                if getattr(self, "active_putty_apps", {}):
                    for ip, (app, term_win) in list(self.active_putty_apps.items()):
                        try:
                            term_win.type_keys("STR_TEST:CLOSE{ENTER}")
                            time.sleep(0.5)
                            app.kill()
                        except Exception: pass
                    self.active_putty_apps = {}

                self.root.after(1500, self._run_next_commg_step)
                return
                
        if getattr(self, "_commg_is_active_run", False):
            self._is_paused = True
            self.status_var.set("Automation Paused")
            self.status_label.configure(fg="#E65100")
            self.q.put("\n[GUI] Sequence stopped manually. State saved. Click 'Start' to resume.\n", ("warn",))
            
            if hasattr(self, "_current_batch_items") and self._current_batch_items:
                self._commg_pending_queue = self._current_batch_items + getattr(self, "_commg_pending_queue", [])
                self._current_batch_items = []
        else:
            self._commg_pending_queue = []
            
        self._commg_is_active_run = False
        self._set_running(False)
        
        if getattr(self, "active_sockets", []):
            for s in self.active_sockets:
                try:
                    s.sendall(b"STR_TEST:CLOSE\r\n")
                    time.sleep(0.2)
                    s.close()
                except Exception: pass
            self.active_sockets = []

        if getattr(self, "active_putty_apps", {}):
            for ip, (app, term_win) in list(self.active_putty_apps.items()):
                try:
                    term_win.type_keys("STR_TEST:CLOSE{ENTER}")
                    time.sleep(0.5)
                    app.kill()
                except Exception: pass
            self.active_putty_apps = {}

        if hasattr(self, "current_process") and self.current_process:
            try:
                self._process_gen = getattr(self, "_process_gen", 0) + 1 
                self.current_process.kill()
                self._append("\n[GUI] Background AI process force-stopped.\n", ("warn",))
            except Exception:
                pass
        elif hasattr(self, "proc_thread") and self.proc_thread and self.proc_thread.is_alive():
            self._append("\n[WARNING] Note: Background script is still finishing its current execution loop.\n")
            
    def _write_final_excel_summary(self):
        if getattr(self, "_summary_written", False):
            return
        self._summary_written = True
        
        if not self.save_log_var.get() or not getattr(self, "_log_session_dir", None):
            return
            
        filename = self.excel_filename_var.get().strip() or "Batch_Summary_Report.xlsx"
        if not filename.lower().endswith(".xlsx"): filename += ".xlsx"
        
        base_name = filename[:-5]
        raw_xl_p = Path(self._log_session_dir) / filename
        
        if not raw_xl_p.exists():
            return
            
        try:
            import math
            import io
            from datetime import datetime
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.utils import get_column_letter
            import pandas as pd
            import time
            import re
            
            self._append(f"\n[GUI] Starting final Excel cleanup, duplicate removal, and splitting by Language...\n", ("warn",))
            
            wb_in = load_workbook(raw_xl_p)
            ws_in = wb_in.active
            
            image_dict = {}
            for img in ws_in._images:
                orig_row = None
                try:
                    if hasattr(img, 'anchor'):
                        if hasattr(img.anchor, '_from'): orig_row = img.anchor._from.row + 1
                        elif hasattr(img.anchor, 'row'): orig_row = img.anchor.row + 1
                        elif isinstance(img.anchor, str):
                            m = re.search(r'\d+', img.anchor)
                            if m: orig_row = int(m.group())
                except Exception: pass
                
                if orig_row is not None:
                    img_bytes = None
                    try:
                        if hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                            img.ref.seek(0); img_bytes = img.ref.read()
                        elif hasattr(img, '_data'):
                            data = img._data() if callable(img._data) else img._data
                            if hasattr(data, 'read'): data.seek(0); img_bytes = data.read()
                            elif isinstance(data, bytes): img_bytes = data
                    except Exception: pass
                    if img_bytes:
                        image_dict[orig_row] = {"bytes": img_bytes, "width": img.width, "height": img.height}
                        
            max_row = ws_in.max_row
            last_row_val = str(ws_in.cell(row=max_row, column=1).value or "").strip()
            total_data_rows = max_row - 1
            if "SUMMARY" in last_row_val.upper():
                total_data_rows -= 1
                
            headers = [str(ws_in.cell(row=1, column=c).value).strip().lower() for c in range(1, ws_in.max_column + 1)]
            dev_col = headers.index("device") + 1 if "device" in headers else 2
            lang_col = headers.index("language") + 1 if "language" in headers else 4
            idx_col = headers.index("index") + 1 if "index" in headers else 7
            time_col = headers.index("timestamp") + 1 if "timestamp" in headers else 1
            
            latest_row_map = {}
            rows_with_no_index = []
            timestamps = []
            
            for r in range(2, total_data_rows + 2):
                idx_val = str(ws_in.cell(row=r, column=idx_col).value or "").strip()
                dev_val = str(ws_in.cell(row=r, column=dev_col).value or "").strip()
                lang_val = str(ws_in.cell(row=r, column=lang_col).value or "").strip()
                ts_val = str(ws_in.cell(row=r, column=time_col).value or "").strip()
                
                if ts_val and "SUMMARY" not in ts_val.upper():
                    timestamps.append(ts_val)
                
                if idx_val and idx_val.lower() != 'nan' and idx_val != 'none' and "skip" not in idx_val.lower():
                    unique_key = (idx_val, dev_val, lang_val)
                    latest_row_map[unique_key] = r
                else:
                    rows_with_no_index.append(r)
                    
            rows_to_keep_globally = sorted(list(latest_row_map.values()) + rows_with_no_index)
            removed_count = total_data_rows - len(rows_to_keep_globally)
            self._append(f"[GUI] Removed {removed_count} crashed/duplicate rows.\n")
            
            # --- EXACT TIME CALCULATION ---
            parsed_dates = []
            for ts in timestamps:
                try: parsed_dates.append(datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"))
                except Exception: pass
            
            parsed_dates.sort()
            true_avg_sec = 36.75 
            if len(parsed_dates) > 1:
                diffs = [(parsed_dates[i] - parsed_dates[i-1]).total_seconds() for i in range(1, len(parsed_dates))]
                normal_diffs = [d for d in diffs if d <= 90.0]
                if normal_diffs: true_avg_sec = sum(normal_diffs) / len(normal_diffs)

            # FIX: Define total_elapsed here so the Global Summary knows the total time!
            total_elapsed = true_avg_sec * len(rows_to_keep_globally)

            # --- SMART BIN PACKING (Group by Language, Max 80MB) ---
            lang_groups = {}
            for r in rows_to_keep_globally:
                lang_val = str(ws_in.cell(row=r, column=lang_col).value or "Unknown").strip()
                lang_groups.setdefault(lang_val, []).append(r)

            file_size_mb = raw_xl_p.stat().st_size / (1024 * 1024)
            max_mb = 80.0
            est_mb_per_row = file_size_mb / float(max(1, total_data_rows))
            wb_in.close()
            
            final_folder = Path(self._log_session_dir) / base_name
            final_folder.mkdir(parents=True, exist_ok=True)
            
            part_files_rows = []
            current_part_rows = []
            current_part_mb = 0.0
            
            # Pack languages into files without splitting them
            for lang, rows in lang_groups.items():
                lang_mb = len(rows) * est_mb_per_row
                
                # If adding this language exceeds 80MB AND the current part isn't empty, start a new part!
                if current_part_mb + lang_mb > max_mb and len(current_part_rows) > 0:
                    part_files_rows.append(current_part_rows)
                    current_part_rows = list(rows)
                    current_part_mb = lang_mb
                else:
                    current_part_rows.extend(rows)
                    current_part_mb += lang_mb
                    
            if current_part_rows:
                part_files_rows.append(current_part_rows)
                
            num_files_needed = len(part_files_rows)
            part_files_generated = []
            
            global_p = global_f = global_w = global_s = 0
            
            for i, chunk_kept_orig_rows in enumerate(part_files_rows):
                chunk_kept_set = set(chunk_kept_orig_rows)
                
                wb_part = load_workbook(raw_xl_p)
                ws_part = wb_part.active
                ws_part._images = []
                
                if "SUMMARY" in str(ws_part.cell(row=ws_part.max_row, column=1).value or "").upper():
                    ws_part.delete_rows(ws_part.max_row)
                    
                rows_to_delete = sorted([r for r in range(2, total_data_rows + 2) if r not in chunk_kept_set])
                if rows_to_delete:
                    blocks = []
                    start_r = end_r = rows_to_delete[0]
                    for r in rows_to_delete[1:]:
                        if r == end_r + 1: end_r = r
                        else:
                            blocks.append((start_r, end_r))
                            start_r = end_r = r
                    blocks.append((start_r, end_r))
                    
                    for start_b, end_b in reversed(blocks):
                        ws_part.delete_rows(start_b, end_b - start_b + 1)
                        
                current_row = 2
                for orig_r in chunk_kept_orig_rows:
                    if orig_r in image_dict:
                        img_info = image_dict[orig_r]
                        try:
                            new_img = OpenpyxlImage(io.BytesIO(img_info["bytes"]))
                            new_img.width = img_info["width"]
                            new_img.height = img_info["height"]
                            
                            col_idx = 13
                            r_idx = current_row - 1
                            marker_from = AnchorMarker(col=col_idx, colOff=0, row=r_idx, rowOff=0)
                            marker_to = AnchorMarker(col=col_idx + 1, colOff=0, row=r_idx + 1, rowOff=0)
                            new_img.anchor = TwoCellAnchor(editAs='twoCell', _from=marker_from, to=marker_to)
                            ws_part.add_image(new_img)
                        except Exception: pass
                    current_row += 1
                    
                p = f = w = s = 0
                for r in range(2, ws_part.max_row + 1):
                    verdict = str(ws_part.cell(row=r, column=12).value or "").strip().upper()
                    if verdict == "PASS": p += 1
                    elif verdict == "FAIL": f += 1
                    elif verdict == "WARN": w += 1
                    elif verdict == "SKIP": s += 1
                    
                global_p += p; global_f += f; global_w += w; global_s += s
                
                processed_count = p + f + w
                part_time = true_avg_sec * processed_count
                m, sec = divmod(part_time, 60)
                h, m = divmod(m, 60)
                time_str = f"{int(h):02d}h {int(m):02d}m {int(sec):02d}s" if h > 0 else f"{int(m):02d}m {int(sec):02d}s"
                
                summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                summary_font = Font(bold=True)
                
                part_sum_title = f"SUMMARY (Part {i+1})" if num_files_needed > 1 else "SUMMARY"
                ws_part.append([part_sum_title, f"Time: {time_str}", f"PASS: {p}", f"FAIL: {f}", f"WARN: {w}", f"SKIP: {s}"])
                r_idx = ws_part.max_row
                ws_part.row_dimensions[r_idx].height = 25
                for col_num in range(1, 15):
                    cell = ws_part.cell(row=r_idx, column=col_num)
                    cell.fill = summary_fill; cell.font = summary_font; cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if i == num_files_needed - 1:
                    m_g, sec_g = divmod(total_elapsed, 60)
                    h_g, m_g = divmod(m_g, 60)
                    g_time_str = f"{int(h_g):02d}h {int(m_g):02d}m {int(sec_g):02d}s" if h_g > 0 else f"{int(m_g):02d}m {int(sec_g):02d}s"
                    
                    ws_part.append(["GLOBAL SUMMARY", f"Total Time: {g_time_str}", f"PASS: {global_p}", f"FAIL: {global_f}", f"WARN: {global_w}", f"SKIP: {global_s}"])
                    r_idx = ws_part.max_row
                    ws_part.row_dimensions[r_idx].height = 30
                    global_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                    for col_num in range(1, 15):
                        cell = ws_part.cell(row=r_idx, column=col_num)
                        cell.fill = global_fill; cell.font = summary_font; cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                out_name = f"{base_name}_part_{i+1}.xlsx" if num_files_needed > 1 else f"{base_name}_cleaned.xlsx"
                out_path = final_folder / out_name
                wb_part.save(out_path)
                wb_part.close()
                part_files_generated.append(out_path)
                self._append(f"[GUI] Saved cleanly split part: {out_name}\n", ("pass",))

            # ==============================================================================
            # EXECUTIVE ANALYTICS GENERATION
            # ==============================================================================
            self._append(f"[GUI] Generating Executive Analytics Dashboard...\n", ("warn",))
            
            def categorize_root_cause(error_msg):
                msg = str(error_msg).lower()
                if msg == 'nan' or msg.strip() == '': return 'Passed (No Error)'
                if 'truncat' in msg: return 'Screen Truncation (Text too long)'
                if 'roll' in msg: return 'Rolling Text Sync Issue'
                if 'space' in msg: return 'Spacing or Punctuation Mismatch'
                if 'extra' in msg: return 'Extra Words/Prefixes on Screen'
                if 'typo' in msg or 'misread' in msg or 'incorrect character' in msg: return 'OCR Character Confusion / Font Issue'
                if 'completely different' in msg: return 'Completely Wrong Text / Translation Bug'
                if 'missing' in msg: return 'Missing Words/Letters'
                return 'Other Mismatch'

            def calculate_stats_analytics(df, group_col):
                if df.empty: return pd.DataFrame(columns=[group_col, 'Total', 'Passes', 'Fails', 'Warnings', 'Pass_Pct', 'Fail_Pct', 'Warn_Pct'])
                stats = df.groupby(group_col)['Verdict'].value_counts().unstack(fill_value=0)
                for col in ['PASS', 'FAIL', 'WARN']:
                    if col not in stats.columns: stats[col] = 0
                stats['Total'] = stats['PASS'] + stats['FAIL'] + stats['WARN']
                stats['Passes'] = stats['PASS']
                stats['Fails'] = stats['FAIL']
                stats['Warnings'] = stats['WARN']
                stats['Pass_Pct'] = (stats['Passes'] / stats['Total']) * 100
                stats['Fail_Pct'] = (stats['Fails'] / stats['Total']) * 100
                stats['Warn_Pct'] = (stats['Warnings'] / stats['Total']) * 100
                stats = stats.reset_index().sort_values(by='Total', ascending=False)
                if not stats.empty:
                    grand_total = pd.DataFrame({group_col: ['GRAND TOTAL'], 'Total': [stats['Total'].sum()], 'Passes': [stats['Passes'].sum()], 'Fails': [stats['Fails'].sum()], 'Warnings': [stats['Warnings'].sum()]})
                    t_tot = grand_total['Total'].iloc[0]
                    grand_total['Pass_Pct'] = (grand_total['Passes'] / t_tot * 100) if t_tot > 0 else 0
                    grand_total['Fail_Pct'] = (grand_total['Fails'] / t_tot * 100) if t_tot > 0 else 0
                    grand_total['Warn_Pct'] = (grand_total['Warnings'] / t_tot * 100) if t_tot > 0 else 0
                    stats = pd.concat([stats, grand_total], ignore_index=True)
                return stats[[group_col, 'Total', 'Passes', 'Fails', 'Warnings', 'Pass_Pct', 'Fail_Pct', 'Warn_Pct']]

            def autofit_columns_analytics(writer, df, sheet_name):
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(df.columns):
                    series = df[col]
                    max_len = min(max((series.astype(str).map(len).max(), len(str(col)))) + 2, 50)
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

            def inject_dashboard_analytics(writer, df_raw, sheet_name, title):
                if df_raw.empty: return
                verdicts = df_raw['Verdict'].astype(str).str.strip().str.upper()
                passes = (verdicts == 'PASS').sum()
                fails = (verdicts == 'FAIL').sum()
                warns = (verdicts == 'WARN').sum()
                total_valid = passes + fails + warns
                if total_valid == 0: return
                ws = writer.sheets[sheet_name]
                ws.cell(row=1, column=10, value="Verdict").font = Font(bold=True)
                ws.cell(row=1, column=11, value="Count").font = Font(bold=True)
                ws.cell(row=2, column=10, value="PASS")
                ws.cell(row=2, column=11, value=passes)
                ws.cell(row=3, column=10, value="FAIL")
                ws.cell(row=3, column=11, value=fails)
                ws.cell(row=4, column=10, value="WARN")
                ws.cell(row=4, column=11, value=warns)
                manual_review_count = fails + warns
                manual_review_pct = (manual_review_count / total_valid) * 100 if total_valid > 0 else 0
                ws.cell(row=6, column=10, value="QA MANUAL REVIEW REQUIRED:").font = Font(bold=True, color="9C0006")
                ws.cell(row=6, column=11, value=f"{manual_review_count} items ({manual_review_pct:.1f}%)").font = Font(bold=True, color="9C0006")
                pie = PieChart()
                labels = Reference(ws, min_col=10, min_row=2, max_row=4)
                data = Reference(ws, min_col=11, min_row=1, max_row=4)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.title = f"{title} Verdict Breakdown"
                ws.add_chart(pie, "M1")
                max_row = ws.max_row
                for col in range(1, 9):
                    cell = ws.cell(row=max_row, column=col)
                    cell.font = Font(bold=True); cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

            def inject_bar_chart_analytics(writer, df_reasons, sheet_name):
                if df_reasons.empty: return
                ws = writer.sheets[sheet_name]
                bar = BarChart()
                bar.type = "col"
                bar.style = 10
                bar.title = "Root Cause Failure Analysis"
                bar.y_axis.title = 'Number of Errors'
                bar.x_axis.title = 'Error Category'
                data = Reference(ws, min_col=2, min_row=1, max_row=len(df_reasons)+1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_reasons)+1)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                bar.width = 20; bar.height = 10
                ws.add_chart(bar, "D2")

            def generate_historical_snapshot_analytics(df_merged, normal_lang_stats):
                verdicts = df_merged['Verdict'].astype(str).str.strip().str.upper()
                passes = (verdicts == 'PASS').sum()
                fails = (verdicts == 'FAIL').sum()
                warns = (verdicts == 'WARN').sum()
                total = passes + fails + warns
                total_langs = df_merged['Language'].nunique() if 'Language' in df_merged.columns else 0
                pass_pct = round((passes / total) * 100, 1) if total > 0 else 0
                fail_pct = round((fails / total) * 100, 1) if total > 0 else 0
                
                actual_mins = 0.0
                avg_sec_per_string = 0.0
                if 'Timestamp' in df_merged.columns:
                    dates = pd.to_datetime(df_merged['Timestamp'], errors='coerce').dropna().sort_values()
                    if not dates.empty and len(dates) > 1:
                        diffs = dates.diff().dt.total_seconds().dropna()
                        diffs = diffs.clip(upper=90.0)
                        total_active_seconds = diffs.sum() + 30.0 
                        actual_mins = round(total_active_seconds / 60.0, 1)
                        if total > 0: avg_sec_per_string = round(total_active_seconds / total, 1)
                
                worst_lang = "None"
                if not normal_lang_stats.empty and 'Language' in normal_lang_stats.columns:
                    lang_only = normal_lang_stats[normal_lang_stats['Language'] != 'GRAND TOTAL']
                    if not lang_only.empty:
                        worst_lang_row = lang_only.sort_values(by='Fail_Pct', ascending=False).iloc[0]
                        worst_lang = f"{worst_lang_row['Language']} ({worst_lang_row['Fail_Pct']:.1f}% Fail)"
                
                snapshot = pd.DataFrame({
                    "Run_Date": [time.strftime("%Y-%m-%d %H:%M")],
                    "Total_Languages_Tested": [total_langs],
                    "Total_Strings_Verified": [total],
                    "Overall_Pass_%": [pass_pct],
                    "Overall_Fail_%": [fail_pct],
                    "Worst_Language": [worst_lang],
                    "Avg_Time_Per_String_Sec": [avg_sec_per_string],
                    "Actual_Total_Time_Mins": [actual_mins]
                })
                return snapshot

            all_dfs = []
            for fp in part_files_generated:
                df_part = pd.read_excel(str(fp), engine='openpyxl')
                if 'ROI Image' in df_part.columns: df_part = df_part.drop(columns=['ROI Image'])
                df_part = df_part[~df_part.iloc[:, 0].astype(str).str.contains("SUMMARY", na=False, case=False)]
                df_part = df_part[df_part.iloc[:, 0] != "Timestamp"]
                all_dfs.append(df_part)

            df_merged = pd.concat(all_dfs, ignore_index=True).dropna(how='all')
            df_merged['Verdict'] = df_merged['Verdict'].astype(str).str.strip().str.upper()
            df_merged = df_merged[df_merged['Verdict'] != 'SKIP']
            
            def get_type(error_msg):
                msg = str(error_msg).lower()
                if "rolling" in msg: return "Rolling"
                elif "truncated" in msg: return "Truncation"
                else: return "Normal"
                    
            df_merged['Type'] = df_merged['Error Message'].apply(get_type) if 'Error Message' in df_merged.columns else "Normal"
            df_merged['Failure_Root_Cause'] = df_merged['Error Message'].apply(categorize_root_cause)
            
            df_errors_only = df_merged[df_merged['Verdict'].isin(['FAIL', 'WARN'])]
            root_cause_stats = df_errors_only['Failure_Root_Cause'].value_counts().reset_index()
            root_cause_stats.columns = ['Error Category', 'Number of Occurrences']
            
            df_normal = df_merged[df_merged['Type'] == 'Normal']
            df_rolling = df_merged[df_merged['Type'] == 'Rolling']
            df_trunc = df_merged[df_merged['Type'] == 'Truncation']
            
            all_lang_stats = calculate_stats_analytics(df_merged, 'Language')
            normal_lang_stats = calculate_stats_analytics(df_normal, 'Language')
            device_stats = calculate_stats_analytics(df_merged, 'Device')
            error_stats = calculate_stats_analytics(df_merged, 'Type')
            
            snapshot_df = generate_historical_snapshot_analytics(df_merged, normal_lang_stats)
            
            analytics_filename = f"{base_name}_Executive_Analytics.xlsx"
            analytics_out_path = final_folder / analytics_filename
            
            with pd.ExcelWriter(str(analytics_out_path), engine='openpyxl') as writer:
                snapshot_df.to_excel(writer, sheet_name='Historical_Snapshot', index=False)
                autofit_columns_analytics(writer, snapshot_df, 'Historical_Snapshot')
                
                all_lang_stats.to_excel(writer, sheet_name='Lang_Stats_All', index=False)
                autofit_columns_analytics(writer, all_lang_stats, 'Lang_Stats_All')
                inject_dashboard_analytics(writer, df_merged, 'Lang_Stats_All', "Overall Language")
                
                root_cause_stats.to_excel(writer, sheet_name='Error_Root_Cause_Analysis', index=False)
                autofit_columns_analytics(writer, root_cause_stats, 'Error_Root_Cause_Analysis')
                inject_bar_chart_analytics(writer, root_cause_stats, 'Error_Root_Cause_Analysis')
                
                device_stats.to_excel(writer, sheet_name='Device_Hardware_Stats', index=False)
                autofit_columns_analytics(writer, device_stats, 'Device_Hardware_Stats')
                inject_dashboard_analytics(writer, df_merged, 'Device_Hardware_Stats', "Device Hardware")
                
                error_stats.to_excel(writer, sheet_name='Error_Type_Breakdown', index=False)
                autofit_columns_analytics(writer, error_stats, 'Error_Type_Breakdown')
                
                df_normal.to_excel(writer, sheet_name='Raw_Data_Normal', index=False)
                autofit_columns_analytics(writer, df_normal, 'Raw_Data_Normal')
                
                if not df_rolling.empty:
                    df_rolling.to_excel(writer, sheet_name='Raw_Data_Rolling', index=False)
                    autofit_columns_analytics(writer, df_rolling, 'Raw_Data_Rolling')
                    
                if not df_trunc.empty:
                    df_trunc.to_excel(writer, sheet_name='Raw_Data_Truncation', index=False)
                    autofit_columns_analytics(writer, df_trunc, 'Raw_Data_Truncation')
            
            self._append(f"[GUI] Saved Executive Analytics Dashboard: {analytics_filename}\n", ("pass",))

            drive_folder_name = self.drive_folder_var.get().strip()
            if drive_folder_name:
                import shutil
                target_dir = Path("G:/My Drive") / drive_folder_name / base_name
                target_dir.mkdir(parents=True, exist_ok=True)
                
                for f in final_folder.glob("*.xlsx"):
                    try: shutil.copy2(f, target_dir / f.name)
                    except Exception: pass
                self._append(f"[GUI] Synced final folder to Google Drive: {drive_folder_name}/{base_name}\n", ("pass",))
                
        except Exception as e:
            self._append(f"\n[GUI ERROR] Failed to process final Excel: {e}\n", ("error",))
            
    def _drain_queue(self):
        try:
            while True:
                s = self.q.get_nowait()
                
                if isinstance(s, tuple) and len(s) == 2 and s[0] == "__PREFETCH_DONE__":
                    self.q.put("[Pre-Check] Setup complete. Waiting for GenAI initialization...\n")
                    self.status_var.set("Init GenAI...")
                    continue

                if isinstance(s, tuple) and len(s) == 2 and s[0] == "__LANG_CHANGE_DONE__":
                    self.q.put("[Batch] Languages changed. Starting String Verification...\n")
                    
                    # --- FIX: Resume the stopwatch! ---
                    self._is_adb_running = False
                    self.batch_start_time = time.time()
                    # ----------------------------------
                    
                    self._update_accumulated_time()
                    
                    if getattr(self, "_protect_resume_queue", False):
                        self.q.put("[GUI] Protected sliced resume queue from being overwritten!\n", ("pass",))
                        self._protect_resume_queue = False 
                    else:
                        if getattr(self, "_full_batch_commands", []):
                            self._commg_pending_queue = list(self._full_batch_commands)
                            self.q.put(f"[GUI] Master queue loaded ({len(self._commg_pending_queue)} items).\n", ("pass",))
                        else:
                            self.q.put("[GUI FATAL ERROR] Master command memory is missing. Cannot proceed.\n", ("error",))
                            self._set_running(False)
                            return
                    
                    self._current_lang_idx += 1
                    self._run_next_commg_step()
                    continue

                if isinstance(s, tuple) and len(s) >= 2 and s[0] == _EVT_COMMG_DONE:
                    status = s[1]
                    batch_items = s[2] if len(s) > 2 else []
                    
                    if status in ("ABORT", "PAUSE"):
                        self._commg_is_active_run = False
                        if status == "PAUSE":
                            self._is_paused = True
                            if batch_items:
                                self._commg_pending_queue = batch_items + getattr(self, "_commg_pending_queue", [])
                            self.status_var.set("Automation Sequence Paused")
                            self.status_label.configure(fg="#E65100")
                        else:
                            self._commg_pending_queue = []
                            self.status_var.set("Automation Sequence Aborted")
                            self.status_label.configure(fg="red")
                            
                        self._set_running(False)
                        if getattr(self, "_summary_written", False) == False:
                            threading.Thread(target=self._write_final_excel_summary, daemon=True).start()
                    else:
                        if not getattr(self, "_commg_is_active_run", False):
                            continue
                        
                        all_skip = True
                        for t, idx in zip(self.tag_var.get().split(","), self.index_var.get().split(",")):
                            if t.strip() != "SKIP_VERIFY" and idx.strip() != "SKIP_VERIFY":
                                all_skip = False
                                break
                                
                        if all_skip:
                            self.q.put("[GUI] Skipping verification phase as requested for all.\n", ("commg",))
                            
                            ips = list(self.ip_listbox.get(0, tk.END))
                            skip_records = []
                            cmds_for_skip = getattr(self, "_current_batch_cmds", [])
                            for i, (t, idx) in enumerate(zip(self.tag_var.get().split(","), self.index_var.get().split(","))):
                                dev_nm = ips[i] if i < len(ips) else f"Device {i+1}"
                                cmd_str = cmds_for_skip[i] if i < len(cmds_for_skip) else "-"
                                dev_lang = ",".join(self._get_device_langs(self.device_rows[i])) if i < len(self.device_rows) else "English"
                                
                                style_str = "-"
                                try:
                                    mapping = {34: "EMERGENCY_NOTICE_CENTER_DISP_STYLE", 38: "FEATURE_HOME_SCREEN_SUBFEATURE_DISP_STYLE", 39: "FEATURE_HOME_SCREEN_FEATURE_DISP_STYLE", 40: "FEATURE_HOME_SCREEN_FIRST_LINE_SELECTED_FEATURE_DISP_STYLE", 41: "FEATURE_HOME_SCREEN_SYSTEM_DISP_STYLE", 50: "FEATURE_GOOD_NOTICE_SYSTEM_DISP_STYLE", 51: "FEATURE_GOOD_NOTICE_FEATURE_DISP_STYLE", 52: "FEATURE_BAD_NOTICE_SYSTEM_DISP_STYLE", 53: "FEATURE_BAD_NOTICE_FEATURE_DISP_STYLE", 54: "FEATURE_NEUTRAL_NOTICE_SYSTEM_DISP_STYLE", 55: "FEATURE_NEUTRAL_NOTICE_FEATURE_DISP_STYLE"}
                                    parts = cmd_str.split(":")
                                    if len(parts) >= 3:
                                        style_str = mapping.get(int(parts[2]), f"UNKNOWN_{parts[2]}")
                                except: pass
                                
                                payload = {
                                    "device": dev_nm, 
                                    "command": cmd_str, 
                                    "display_style": style_str, 
                                    "index": idx, 
                                    "tag": t, 
                                    "expected": "-", 
                                    "actual": "-", 
                                    "verdict": "SKIP", 
                                    "error": "",
                                    "language": dev_lang
                                }
                                self.all_results_data.append(payload)
                                skip_records.append(payload)
                                
                            self._update_summary_ui(self.current_filter)
                            
                            if getattr(self, "_log_session_dir", None):
                                filename_base = self.excel_filename_var.get().strip() or "Batch_Summary_Report"
                                if filename_base.lower().endswith(".xlsx"): 
                                    filename_base = filename_base[:-5]
                                    
                                for rec in skip_records:
                                    dev_lang = rec["language"]
                                    safe_lang = re.sub(r'[\\/*?:"<>| ]', '_', dev_lang)
                                    
                                    part_num = 1
                                    while True:
                                        candidate_path = Path(self._log_session_dir) / f"{filename_base}_{safe_lang}_part{part_num}.xlsx"
                                        if not candidate_path.exists():
                                            xl_p = candidate_path
                                            break
                                        if candidate_path.stat().st_size / (1024 * 1024) < 80.0:
                                            xl_p = candidate_path
                                            break
                                        part_num += 1

                                    try:
                                        from openpyxl import load_workbook, Workbook
                                        from openpyxl.styles import PatternFill, Font, Alignment
                                        if not xl_p.exists():
                                            wb = Workbook()
                                            ws = wb.active
                                            ws.title = "Batch Summary"
                                            headers = ["Timestamp", "Device", "Region", "Language", "Command", "Display Style", "Index", "Tag", "Expected (Local)", "Actual Detected", "Confidence (%)", "Verdict", "Error Message", "ROI Image"]
                                            ws.append(headers)
                                            
                                            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                                            header_font = Font(color="FFFFFF", bold=True)
                                            for col_num, cell in enumerate(ws[1], 1):
                                                cell.fill = header_fill
                                                cell.font = header_font
                                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                            
                                            widths = [20, 15, 12, 15, 25, 45, 10, 25, 35, 35, 15, 12, 35, 30]
                                            for idx_col, width in enumerate(widths, 1):
                                                ws.column_dimensions[ws.cell(row=1, column=idx_col).column_letter].width = width
                                        else:
                                            wb = load_workbook(xl_p)
                                            ws = wb.active
                                            
                                        ws.append([
                                            time.strftime("%Y-%m-%d %H:%M:%S"),
                                            rec["device"],
                                            "Multiple",
                                            dev_lang,
                                            rec.get("command", "-"),
                                            rec.get("display_style", "-"),
                                            rec["index"], 
                                            rec["tag"],
                                            "-", "-", "-", "SKIP", "", ""
                                        ])
                                        
                                        row_idx = ws.max_row
                                        for col_num in range(1, 15):
                                            cell = ws.cell(row=row_idx, column=col_num)
                                            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                                            
                                        verdict_cell = ws.cell(row=row_idx, column=12)
                                        verdict_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                        verdict_cell.font = Font(color="333333", bold=True)
                                        ws.row_dimensions[row_idx].height = 80
                                            
                                        wb.save(xl_p)
                                    except Exception:
                                        pass

                            if self._commg_pending_queue:
                                self._run_next_commg_step()
                            else:
                                self._start_language_iteration() 
                        else:
                            self.run()
                    continue

                if isinstance(s, tuple) and len(s) >= 2 and s[0] == _EVT_FINISHED:
                    rc = s[1]
                    proc_gen = s[2] if len(s) > 2 else -1
                    
                    if hasattr(self, "_process_gen") and proc_gen != -1 and proc_gen != self._process_gen:
                        continue 
                        
                    self._current_batch_items = [] 
                        
                    is_verify = bool(self._last_run_is_verification)
                    
                    if is_verify:
                        self._update_accumulated_time()
                        
                       # --- FIX: ATOMIC GOOGLE DRIVE TRANSFER ---
                        try:
                            if self.save_log_var.get() and getattr(self, "_log_session_dir", None):
                                import shutil
                                import zipfile
                                import time
                                import os
                                
                                filename = self.excel_filename_var.get().strip() or "Batch_Summary_Report.xlsx"
                                if not filename.lower().endswith(".xlsx"): filename += ".xlsx"
                                raw_xl_p = Path(self._log_session_dir) / filename
                                
                                drive_folder_name = self.drive_folder_var.get().strip() or "Walkie_Logs"
                                base_name = filename[:-5]
                                
                                drive_base = Path("G:/My Drive") 
                                if not (drive_base / drive_folder_name).exists():
                                    drive_base = Path("G:/Shared drives")
                                    
                                target_dir = drive_base / drive_folder_name / base_name
                                target_dir.mkdir(parents=True, exist_ok=True)
                                
                                current_sync_time = time.time()
                                last_sync_time = getattr(self, "_last_drive_sync_time", 0)
                                
                                if raw_xl_p.exists():
                                    if zipfile.is_zipfile(str(raw_xl_p)):
                                        
                                        # ONLY sync every 3 minutes so Google Drive Web has time to render
                                        if current_sync_time - last_sync_time >= 180.0:
                                            target_file = target_dir / f"{base_name}_LIVE_UPDATE.xlsx"
                                            
                                            # 1. Copy as a hidden .tmp file so Google Drive Desktop ignores it
                                            temp_target = target_dir / f"{base_name}_UPLOADING.tmp"
                                            shutil.copy2(raw_xl_p, temp_target)
                                            
                                            # 2. Instantly rename it to .xlsx. This forces Drive to upload the file WHOLE!
                                            os.replace(temp_target, target_file)
                                            
                                            self._last_drive_sync_time = current_sync_time
                                            self.q.put(f"[Live Sync] Safely uploaded to Drive (3-min throttle): {target_file.name}\n", ("pass",))
                                    else:
                                        self.q.put(f"[Live Sync Blocked] Local Excel is corrupted! Blocked upload to protect Google Drive.\n", ("error",))
                                else:
                                    self.q.put(f"[Live Sync Error] Local Excel file not found at {raw_xl_p}!\n", ("error",))
                            else:
                                self.q.put("[Live Sync Info] Skipped Drive Sync. Either 'Save log' is unchecked, or no session exists.\n", ("warn",))
                        except Exception as e:
                            self.q.put(f"[GUI ERROR] Live Sync to Google Drive Failed: {e}\n", ("error",))
                        # ----------------------------------------------------------------------------------

                    if is_verify and getattr(self, "active_sockets", []):
                        self.q.put("[CMD] Verification complete. Sending close command and terminating sessions...\n")
                        for soc in self.active_sockets:
                            try:
                                soc.sendall(b"STR_TEST:CLOSE\r\n")
                                time.sleep(0.2)
                                soc.close()
                            except Exception:
                                pass
                        self.active_sockets = []

                    if is_verify and getattr(self, "active_putty_apps", {}):
                        self.q.put("[PuTTY] Verification complete. Terminating sessions...\n")
                        for ip, (app, term_win) in list(self.active_putty_apps.items()):
                            try:
                                term_win.type_keys("STR_TEST:CLOSE{ENTER}")
                                time.sleep(0.5)
                                app.kill()
                            except Exception:
                                pass
                        self.active_putty_apps = {}

                    will_autostart = False
                    try:
                        will_autostart = bool(self._auto_start_verify) and (not is_verify) and int(rc or 0) == 0
                    except Exception:
                        will_autostart = False

                    if self._commg_is_active_run:
                        if is_verify:
                            if self._commg_pending_queue:
                                self._run_next_commg_step()
                            else:
                                self._start_language_iteration()
                            continue
                        elif will_autostart:
                            self._auto_start_verify = False 
                            self._is_resuming_batch = False 
                            
                            # --- FIX: ALWAYS trigger language check, even on resume ---
                            self._start_language_iteration()
                            # ----------------------------------------------------------
                            continue

                    if will_autostart:
                        self.status_var.set("Init GenAI finished, starting verification...")
                    else:
                        if is_verify: 
                            if getattr(self, "_is_paused", False): msg = "Automation Paused"
                            else: msg = "Finished" if rc in [0, None] else f"Finished (exit={rc})"
                        else: 
                            if getattr(self, "_is_paused", False): msg = "Automation Paused"
                            else: msg = "Init GenAI finished" if rc in [0, None] else f"Init GenAI finished (exit={rc})"
                        self.status_var.set(msg)

                        try:
                            if rc not in [0, None]: self.status_label.configure(fg="#B71C1C")
                            else: self.status_label.configure(fg="black")
                        except Exception: pass

                        if not will_autostart: 
                            self._set_running(False)
                            self._update_summary_ui(self.current_filter)
                            if is_verify:
                                threading.Thread(target=self._write_final_excel_summary, daemon=True).start()

                    try:
                        if self._log_fp is not None:
                            self._log_fp.flush()
                            self._log_fp.close()
                            self._log_fp = None
                    except Exception: pass

                    try:
                        if will_autostart:
                            self._auto_start_verify = False
                            self.run()
                    except Exception:
                        self._auto_start_verify = False
                    continue

                if isinstance(s, str):
                    for line in s.splitlines(True):
                        self._append_line_with_result_color(line)
                else:
                    self._append(str(s))
        except queue.Empty:
            pass

        self.root.after(50, self._drain_queue)

def main():
    root = tk.Tk()
    app = VerifyStringGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()