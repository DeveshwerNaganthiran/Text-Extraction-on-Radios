import os
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit()

def main():
    root = tk.Tk()
    root.withdraw() 
    
    print("Select the folder that contains your .log file and PASS/FAIL/WARN subfolders...")
    folder_path = filedialog.askdirectory(title="Select Log & Image Folder")
    
    if not folder_path:
        print("No folder selected. Exiting.")
        return
        
    folder = Path(folder_path)
    log_files = list(folder.glob("*.log"))
    
    if not log_files:
        print("No .log file found in that folder! Please select the correct folder.")
        return
        
    log_file = log_files[0]
    print(f"Reading backup data from plain text log: {log_file.name}")
    
    # 1. Extract the raw text data
    results = []
    current_data = {}
    
    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
        
    for line in lines:
        line = line.strip()
        
        dev_match = re.search(r"Device:\s*(.*?)\s*\(Extracting", line)
        if dev_match:
            current_data = {"device": dev_match.group(1).strip()}
            continue
            
        if line.startswith("Command:"): current_data["command"] = line.split("Command:", 1)[1].strip()
        elif line.startswith("Display Style:"): current_data["display_style"] = line.split("Display Style:", 1)[1].strip()
        elif line.startswith("Index:"): current_data["index"] = line.split("Index:", 1)[1].strip()
        elif line.startswith("Tag:"): current_data["tag"] = line.split("Tag:", 1)[1].strip()
        elif line.startswith("Language:"):
            parts = line.split("|")
            current_data["language"] = parts[0].replace("Language:", "").strip()
            if len(parts) > 1 and "Region:" in parts[1]:
                current_data["region"] = parts[1].replace("Region:", "").strip()
        elif line.startswith("Expected (normalized):"): current_data["expected"] = line.split("Expected (normalized):", 1)[1].strip()
        elif line.startswith("Observed (normalized):"): current_data["actual"] = line.split("Observed (normalized):", 1)[1].strip()
        elif line.startswith("Match Confidence:"): current_data["confidence"] = line.split("Match Confidence:", 1)[1].strip()
        elif line.startswith("[VERDICT REASON]"): current_data["error"] = line.split("[VERDICT REASON]", 1)[1].strip()
        elif line.startswith("[") and ("ERROR:" in line): current_data["error"] = line
        elif line in ["PASS", "FAIL", "WARN", "SKIP"]:
            if current_data.get("device"):
                current_data["verdict"] = line
                results.append(current_data.copy())
                current_data = {}
                
    print(f"Found {len(results)} saved rows in the log file. Rebuilding Excel...")
    
    # 2. Gather images and sort them CHRONOLOGICALLY (oldest first) to match log order
    available_images = list(folder.rglob("*.jpg"))
    available_images.sort(key=lambda x: x.stat().st_mtime)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Summary"
    
    headers = ["Timestamp", "Device", "Region", "Language", "Command", "Display Style", "Index", "Tag", "Expected (Local)", "Actual Detected", "Confidence (%)", "Verdict", "Error Message", "ROI Image"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
        
    widths = [20, 15, 12, 15, 25, 45, 10, 25, 35, 35, 15, 12, 35, 30]
    for idx_col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx_col).column_letter].width = width

    # 4. Inject data and securely lock images
    for i, data in enumerate(results, 2):
        dev = data.get("device", "")
        idx = data.get("index", "")
        verdict = data.get("verdict", "").upper()
        region = data.get("region", "Multiple").upper()
        
        safe_dev = re.sub(r'[\\/*?:"<>| ]', '_', dev)
        matched_img = None
        
        if verdict != "SKIP":
            if idx and idx != "SKIP_VERIFY":
                pattern = re.compile(rf"^roi_{re.escape(str(idx))}_{re.escape(safe_dev)}_\d+\.jpg$")
            else:
                pattern = re.compile(rf"^roi_{re.escape(safe_dev)}_\d+\.jpg$")
            
            for img in available_images:
                # --- FIX: Ensure the image is INSIDE the folder that matches the verdict! ---
                if img.parent.name.upper() != verdict:
                    continue
                
                if pattern.match(img.name):
                    matched_img = img
                    break
                    
        if matched_img:
            ts = datetime.fromtimestamp(matched_img.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        else:
            ts = "Recovered_From_Log"
        
        ws.append([
            ts, dev, region, data.get("language", ""), data.get("command", ""), data.get("display_style", ""),
            idx, data.get("tag", ""), data.get("expected", ""), data.get("actual", ""),
            data.get("confidence", "").replace("%", ""), verdict, data.get("error", ""), ""
        ])
        
        for col_num in range(1, 15):
            cell = ws.cell(row=i, column=col_num)
            h_align = "left" if col_num in [9, 10, 13] else "center"
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal=h_align)

        v_cell = ws.cell(row=i, column=12)
        if verdict == "PASS": v_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"); v_cell.font = Font(color="006100", bold=True)
        elif verdict == "FAIL": v_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"); v_cell.font = Font(color="9C0006", bold=True)
        elif verdict == "WARN": v_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"); v_cell.font = Font(color="9C6500", bold=True)
            
        ws.row_dimensions[i].height = 80
        
        if matched_img:
            available_images.remove(matched_img)
            try:
                img_obj = OpenpyxlImage(str(matched_img))
                
                orig_w, orig_h = img_obj.width, img_obj.height
                target_h = 450 if orig_h > 300 else 120
                target_w = int(orig_w * (target_h / float(orig_h)))
                
                img_obj.height = target_h; img_obj.width = target_w
                ws.row_dimensions[i].height = int(target_h * 0.75)
                
                current_w = ws.column_dimensions['N'].width
                needed_w = target_w / 7.0
                if current_w is None or needed_w > current_w:
                    ws.column_dimensions['N'].width = needed_w
                
                marker_from = AnchorMarker(col=13, colOff=0, row=i-1, rowOff=0)
                marker_to = AnchorMarker(col=14, colOff=0, row=i, rowOff=0)
                img_obj.anchor = TwoCellAnchor(editAs='twoCell', _from=marker_from, to=marker_to)
                
                ws.add_image(img_obj)
            except Exception as e:
                print(f"Failed to attach image {matched_img.name}: {e}")

    # FIX: Lock auto-filters strictly to the dimensions of the generated table
    ws.auto_filter.ref = ws.dimensions

    out_path = folder / "RECOVERED_Batch_Summary.xlsx"
    wb.save(out_path)
    print(f"\n{'='*50}")
    print(f"SUCCESS! Fully recovered Excel file saved to:\n{out_path}")
    print(f"{'='*50}\n")

if __name__ == "__main__":
    main()